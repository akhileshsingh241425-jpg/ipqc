from flask import Blueprint, request, jsonify, send_file
from app.models.database import db
from sqlalchemy import text
import requests
import os
import json
import io
import re
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

ai_assistant_bp = Blueprint('ai_assistant', __name__)

# AI API Configuration
GROQ_API_KEY = os.environ.get('GROQ_API_KEY', '')
GROQ_API_URL = 'https://api.groq.com/openai/v1/chat/completions'

# External Barcode Tracking API

# External Barcode Tracking API
BARCODE_TRACKING_API = 'https://umanmrp.in/api/get_barcode_tracking.php'

# ============================================
# UTILITY FUNCTIONS - Julian Date & Barcode
# ============================================

def parse_barcode(barcode):
    """
    Parse barcode to extract all information
    Format: GS04875KG3022500075
    Returns: {prefix, model, plant, julian_date, year, serial, manufacturing_date}
    """
    if not barcode or len(barcode) < 18:
        return None
    
    try:
        return {
            'barcode': barcode,
            'prefix': barcode[0:2],           # GS
            'model': barcode[2:7],            # 04875
            'plant': barcode[7:9],            # KG
            'julian_date': int(barcode[9:12]), # 302
            'year': int('20' + barcode[12:14]), # 2025
            'serial': barcode[14:],           # 00075
            'manufacturing_date': julian_to_date(int(barcode[9:12]), int('20' + barcode[12:14]))
        }
    except:
        return None

def julian_to_date(julian_day, year):
    """Convert Julian day number to calendar date"""
    try:
        base_date = datetime(year, 1, 1)
        target_date = base_date + timedelta(days=julian_day - 1)
        return target_date.strftime('%Y-%m-%d')
    except:
        return None

def date_to_julian(date_str):
    """Convert date string (YYYY-MM-DD) to Julian day"""
    try:
        date_obj = datetime.strptime(date_str, '%Y-%m-%d')
        return date_obj.timetuple().tm_yday
    except:
        return None

def get_julian_from_barcode(barcode):
    """Extract Julian date from barcode"""
    if len(barcode) >= 12:
        try:
            return int(barcode[9:12])
        except:
            return None
    return None

def get_all_mrp_data(company):
    """Get all MRP data for a company with full details"""
    mrp_party_name = get_mrp_party_name(company)
    try:
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=60
        )
        if response.status_code == 200:
            data = response.json()
            if data.get('status') == 'success':
                return {'success': True, 'data': data.get('data', [])}
        return {'success': False, 'data': []}
    except Exception as e:
        return {'success': False, 'error': str(e), 'data': []}

def extract_binning_from_ro(running_order):
    """Extract binning (I1, I2, I3) from running_order field"""
    if not running_order:
        return None
    match = re.search(r'i-?(\d+)', running_order, re.IGNORECASE)
    if match:
        return f"I{match.group(1)}"
    return None

def extract_ro_from_ro(running_order):
    """Extract R-O (R-1, R-2, R-3) from running_order field"""
    if not running_order:
        return None
    match = re.search(r'(R-\d+)', running_order, re.IGNORECASE)
    if match:
        return match.group(1).upper()
    return None

# Company Name Mapping (Database name -> MRP API name)
# EXACT names from production server
COMPANY_NAME_MAPPING = {
    # Rays Power (Production Server Name)
    'Rays Power': 'RAYS POWER INFRA PRIVATE LIMITED',
    'rays power': 'RAYS POWER INFRA PRIVATE LIMITED',
    'Rays power': 'RAYS POWER INFRA PRIVATE LIMITED',
    'RAYS POWER': 'RAYS POWER INFRA PRIVATE LIMITED',
    
    # Larsen & Toubro (Production Server Name)
    'Larsen & Toubro': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'larsen & toubro': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'LARSEN & TOUBRO': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    'L&T': 'LARSEN & TOUBRO LIMITED, CONSTRUCTION',
    
    # Sterlin and Wilson (Production Server Name - Note: "Sterlin" not "Sterling")
    'Sterlin and Wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'sterlin and wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'STERLIN AND WILSON': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
    'Sterling and Wilson': 'STERLING AND WILSON RENEWABLE ENERGY LIMITED',
}

def get_mrp_party_name(db_company_name):
    """Get the MRP API party name from database company name"""
    # Try exact match first
    if db_company_name in COMPANY_NAME_MAPPING:
        return COMPANY_NAME_MAPPING[db_company_name]
    # Try lowercase match
    lower_name = db_company_name.lower()
    for key, value in COMPANY_NAME_MAPPING.items():
        if key.lower() == lower_name:
            return value
    # Return original if no mapping found
    return db_company_name

# Cache for external API data (5 minute cache)
_external_cache = {}
_cache_timeout = 300  # 5 minutes

def check_mix_packing(company):
    """
    Check for mix packing in pallets using MASTER FTR DATABASE binning data
    Mix packing = Single pallet contains multiple binning types (I1+I2, I2+I3, etc.)
    Returns list of pallets with mix packing issues
    
    IMPORTANT: Uses actual binning from ftr_master_serials table, NOT from MRP running_order field
    """
    import re
    
    # Default to Rays Power if no company specified
    if not company:
        company = 'Rays Power'
    
    try:
        # Step 1: Get company_id from database
        company_result = db.session.execute(text(
            "SELECT id FROM companies WHERE company_name LIKE :name"
        ), {'name': f'%{company.split()[0]}%'})
        company_row = company_result.fetchone()
        
        if not company_row:
            return {'success': False, 'error': f'Company {company} not found in database'}
        
        company_id = company_row[0]
        
        # Step 2: Fetch MRP data (pallet info + barcodes)
        mrp_party_name = get_mrp_party_name(company)
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=60
        )
        
        if response.status_code != 200:
            return {'success': False, 'error': 'MRP API failed'}
        
        mrp_data = response.json()
        all_barcodes = mrp_data.get('data', [])
        
        # Step 3: ⚡ OPTIMIZED - Fetch ALL binnings in ONE query
        all_serials = [b.get('barcode') for b in all_barcodes if b.get('barcode')]
        
        if not all_serials:
            return {'success': False, 'error': 'No barcodes found in MRP data'}
        
        # Single query to get all binnings from Master FTR database
        binning_result = db.session.execute(text("""
            SELECT serial_number, binning, class_status 
            FROM ftr_master_serials 
            WHERE company_id = :cid 
            AND serial_number IN :serials
        """), {'cid': company_id, 'serials': tuple(all_serials)})
        
        # Step 4: Create instant lookup dictionary
        binning_lookup = {row[0]: {'binning': row[1], 'class_status': row[2]} for row in binning_result.fetchall()}
        
        # Step 5: Group barcodes by pallet with ACTUAL binning from database
        pallets = {}  # {pallet_no: {'barcodes': [], 'binnings': set(), 'binning_details': []}}
        
        for b in all_barcodes:
            pallet_no = b.get('pallet_no', '')
            if not pallet_no:
                continue
            
            barcode = b.get('barcode', '')
            if not barcode:
                continue
            
            # Get ACTUAL binning from Master FTR database (instant lookup - no DB query)
            db_data = binning_lookup.get(barcode, {'binning': 'Not in Master FTR', 'class_status': 'Unknown'})
            actual_binning = db_data['binning'] if db_data['binning'] else 'Unknown'
            class_status = db_data['class_status'] if db_data['class_status'] else 'OK'
            
            if pallet_no not in pallets:
                pallets[pallet_no] = {
                    'barcodes': [],
                    'binnings': set(),
                    'binning_details': []
                }
            
            pallets[pallet_no]['barcodes'].append(barcode)
            pallets[pallet_no]['binnings'].add(actual_binning)
            pallets[pallet_no]['binning_details'].append({
                'barcode': barcode,
                'binning': actual_binning,
                'class_status': class_status,
                'mrp_running_order': b.get('running_order', '')
            })
        
        # Step 5: Find pallets with mix packing (more than 1 binning type)
        mix_packed_pallets = []
        
        for pallet_no, data in pallets.items():
            # Count ALL binning types - including Unknown and Not in Master FTR
            # These are ALSO different categories and indicate issues
            all_binnings = {b for b in data['binnings'] if b}
            
            # ❌ MIX PACKING if more than 1 binning type found
            if len(all_binnings) > 1:
                # Get binning breakdown with actual barcodes
                binning_groups = {}  # {binning: [barcodes]}
                for detail in data['binning_details']:
                    bin_val = detail['binning']
                    if bin_val not in binning_groups:
                        binning_groups[bin_val] = []
                    binning_groups[bin_val].append(detail['barcode'])
                
                mix_packed_pallets.append({
                    'pallet_no': pallet_no,
                    'binnings': sorted(list(all_binnings)),  # Show ALL binning types
                    'binning_groups': binning_groups,  # Show which barcodes have which binning
                    'total_modules': len(data['barcodes']),
                    'all_binning_details': data['binning_details']  # Full details
                })
        
        # Sort by pallet number
        mix_packed_pallets.sort(key=lambda x: str(x['pallet_no']))
        
        # Build answer
        answer_parts = []
        answer_parts.append(f"**🔍 {company} - Mix Packing Check (Master FTR Database)**\n")
        answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
        answer_parts.append(f"📦 **Total Pallets Checked:** {len(pallets):,}")
        answer_parts.append(f"⚠️ **Mix Packed Pallets:** {len(mix_packed_pallets):,}")
        answer_parts.append(f"✅ **Clean Pallets:** {len(pallets) - len(mix_packed_pallets):,}")
        
        if mix_packed_pallets:
            answer_parts.append(f"\n\n**❌ MIX PACKING FOUND (Based on Master FTR Binning):**\n")
            answer_parts.append(f"⚠️ NOTE: 'Not in Master FTR' means barcode exists in MRP but not in uploaded Master FTR data\n")
            
            for p in mix_packed_pallets[:20]:  # Show first 20
                binnings_str = " + ".join(p['binnings'])
                
                answer_parts.append(f"\n🔸 **Pallet {p['pallet_no']}** ({p['total_modules']} modules)")
                answer_parts.append(f"   ❌ **Mixed Binnings:** {binnings_str}")
                
                # Show barcode breakdown by binning
                for binning, barcodes in sorted(p['binning_groups'].items()):
                    if binning in ['Unknown', 'Not in Master FTR']:
                        continue
                    answer_parts.append(f"   • **{binning}**: {len(barcodes)} modules")
                    # Show first 3 barcodes of each binning type
                    answer_parts.append(f"      {', '.join(barcodes[:3])}")
            
            if len(mix_packed_pallets) > 20:
                answer_parts.append(f"\n\n... and {len(mix_packed_pallets) - 20} more pallets with mix packing")
        else:
            answer_parts.append(f"\n\n✅ **Perfect! No mix packing found. All pallets have uniform binning.**")
        
        return {
            'success': True,
            'has_answer': True,
            'answer': "\n".join(answer_parts),
            'mix_packed_count': len(mix_packed_pallets),
            'total_pallets': len(pallets),
            'mix_packed_pallets': [{
                'pallet_no': p['pallet_no'],
                'binnings': p['binnings'],
                'binnings_by_type': p['binning_groups'],  # For Excel - {binning: [barcodes]}
                'total_modules': p['total_modules']
            } for p in mix_packed_pallets]
        }
        
    except Exception as e:
        print(f"Error checking mix packing: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'has_answer': False, 'error': str(e)}

# ============================================
# QUALITY CHECK FUNCTIONS
# ============================================

def check_duplicate_barcodes(company):
    """Check for duplicate barcodes in MRP system"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    barcodes = mrp_result.get('data', [])
    barcode_counts = {}
    
    for b in barcodes:
        bc = b.get('barcode', '')
        if bc:
            if bc not in barcode_counts:
                barcode_counts[bc] = []
            barcode_counts[bc].append({
                'pallet': b.get('pallet_no', ''),
                'status': 'Dispatched' if b.get('dispatch_party') else 'Packed'
            })
    
    duplicates = {k: v for k, v in barcode_counts.items() if len(v) > 1}
    
    answer_parts = [f"**🔍 {company} - Duplicate Barcode Check**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Barcodes Checked:** {len(barcode_counts):,}")
    answer_parts.append(f"⚠️ **Duplicates Found:** {len(duplicates):,}")
    
    if duplicates:
        answer_parts.append(f"\n\n**❌ DUPLICATE BARCODES:**\n")
        for bc, locations in list(duplicates.items())[:20]:
            pallets = [f"Pallet {l['pallet']} ({l['status']})" for l in locations]
            answer_parts.append(f"🔸 **{bc}**")
            answer_parts.append(f"   Found in: {', '.join(pallets)}")
    else:
        answer_parts.append(f"\n\n✅ **No duplicates found!**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'duplicates': duplicates}

def check_binning_mismatch(company):
    """Check for binning mismatch between Database and MRP"""
    # Get company_id
    company_result = db.session.execute(text(
        "SELECT id FROM companies WHERE company_name LIKE :name"
    ), {'name': f'%{company.split()[0]}%'})
    company_row = company_result.fetchone()
    if not company_row:
        return {'has_answer': False, 'error': 'Company not found'}
    
    company_id = company_row[0]
    
    # Get database binning
    db_result = db.session.execute(text("""
        SELECT serial_number, binning FROM ftr_master_serials 
        WHERE company_id = :cid AND binning IS NOT NULL
    """), {'cid': company_id})
    
    db_binning = {row[0]: row[1] for row in db_result.fetchall()}
    
    # Get MRP data
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    mismatches = []
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        mrp_binning = extract_binning_from_ro(b.get('running_order', ''))
        db_bin = db_binning.get(barcode)
        
        if db_bin and mrp_binning and db_bin != mrp_binning:
            mismatches.append({
                'barcode': barcode,
                'db_binning': db_bin,
                'mrp_binning': mrp_binning,
                'pallet': b.get('pallet_no', '')
            })
    
    answer_parts = [f"**🔍 {company} - Binning Mismatch Check**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"⚠️ **Mismatches Found:** {len(mismatches):,}")
    
    if mismatches:
        answer_parts.append(f"\n\n**❌ BINNING MISMATCHES:**\n")
        for m in mismatches[:20]:
            answer_parts.append(f"🔸 **{m['barcode']}**")
            answer_parts.append(f"   Database: {m['db_binning']} | MRP: {m['mrp_binning']} | Pallet: {m['pallet']}")
    else:
        answer_parts.append(f"\n\n✅ **No binning mismatches found!**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'mismatches': mismatches}

def check_rejected_packed(company):
    """Check if any rejected modules got packed"""
    # Get company_id
    company_result = db.session.execute(text(
        "SELECT id FROM companies WHERE company_name LIKE :name"
    ), {'name': f'%{company.split()[0]}%'})
    company_row = company_result.fetchone()
    if not company_row:
        return {'has_answer': False, 'error': 'Company not found'}
    
    company_id = company_row[0]
    
    # Get rejected serials from database
    rejected_result = db.session.execute(text("""
        SELECT serial_number FROM ftr_master_serials 
        WHERE company_id = :cid AND class_status = 'REJECTED'
    """), {'cid': company_id})
    
    rejected_serials = set(row[0] for row in rejected_result.fetchall())
    
    # Get MRP packed data
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    rejected_packed = []
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        if barcode in rejected_serials:
            rejected_packed.append({
                'barcode': barcode,
                'pallet': b.get('pallet_no', ''),
                'status': 'Dispatched' if b.get('dispatch_party') else 'Packed'
            })
    
    answer_parts = [f"**🔴 {company} - Rejected Modules Check**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Rejected in DB:** {len(rejected_serials):,}")
    answer_parts.append(f"⚠️ **Rejected but Packed:** {len(rejected_packed):,}")
    
    if rejected_packed:
        answer_parts.append(f"\n\n**❌ CRITICAL: REJECTED MODULES PACKED!**\n")
        for r in rejected_packed[:20]:
            answer_parts.append(f"🔴 **{r['barcode']}** - Pallet {r['pallet']} ({r['status']})")
    else:
        answer_parts.append(f"\n\n✅ **No rejected modules found in packed!**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'rejected_packed': rejected_packed}

def check_missing_in_mrp(company):
    """Check barcodes in Database but not in MRP"""
    # Get company_id
    company_result = db.session.execute(text(
        "SELECT id FROM companies WHERE company_name LIKE :name"
    ), {'name': f'%{company.split()[0]}%'})
    company_row = company_result.fetchone()
    if not company_row:
        return {'has_answer': False, 'error': 'Company not found'}
    
    company_id = company_row[0]
    
    # Get assigned serials from database
    db_result = db.session.execute(text("""
        SELECT serial_number, pdi_number FROM ftr_master_serials 
        WHERE company_id = :cid AND status = 'assigned'
    """), {'cid': company_id})
    
    db_serials = {row[0]: row[1] for row in db_result.fetchall()}
    
    # Get MRP data
    mrp_result = get_all_mrp_data(company)
    mrp_barcodes = set(b.get('barcode', '') for b in mrp_result.get('data', []))
    
    missing = []
    for serial, pdi in db_serials.items():
        if serial not in mrp_barcodes:
            missing.append({'barcode': serial, 'pdi': pdi})
    
    answer_parts = [f"**🔍 {company} - Missing in MRP Check**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Assigned in DB:** {len(db_serials):,}")
    answer_parts.append(f"📦 **Found in MRP:** {len(mrp_barcodes):,}")
    answer_parts.append(f"⚠️ **Missing in MRP:** {len(missing):,}")
    
    if missing:
        answer_parts.append(f"\n\n**📋 Sample Missing Barcodes:**\n")
        for m in missing[:15]:
            answer_parts.append(f"   • {m['barcode']} (PDI: {m['pdi']})")
        if len(missing) > 15:
            answer_parts.append(f"\n   ... and {len(missing) - 15} more")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'missing': missing}

def check_extra_in_mrp(company):
    """Check barcodes in MRP but not in Database"""
    # Get company_id
    company_result = db.session.execute(text(
        "SELECT id FROM companies WHERE company_name LIKE :name"
    ), {'name': f'%{company.split()[0]}%'})
    company_row = company_result.fetchone()
    if not company_row:
        return {'has_answer': False, 'error': 'Company not found'}
    
    company_id = company_row[0]
    
    # Get all serials from database
    db_result = db.session.execute(text("""
        SELECT serial_number FROM ftr_master_serials WHERE company_id = :cid
    """), {'cid': company_id})
    
    db_serials = set(row[0] for row in db_result.fetchall())
    
    # Get MRP data
    mrp_result = get_all_mrp_data(company)
    
    extra = []
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        if barcode and barcode not in db_serials:
            extra.append({
                'barcode': barcode,
                'pallet': b.get('pallet_no', ''),
                'running_order': b.get('running_order', '')
            })
    
    answer_parts = [f"**🔍 {company} - Extra in MRP Check**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total in DB:** {len(db_serials):,}")
    answer_parts.append(f"⚠️ **Extra in MRP (Not in DB):** {len(extra):,}")
    
    if extra:
        answer_parts.append(f"\n\n**📋 Sample Extra Barcodes:**\n")
        for e in extra[:15]:
            answer_parts.append(f"   • {e['barcode']} | Pallet: {e['pallet']} | {e['running_order']}")
        if len(extra) > 15:
            answer_parts.append(f"\n   ... and {len(extra) - 15} more")
    else:
        answer_parts.append(f"\n\n✅ **All MRP barcodes exist in database!**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'extra': extra}

# ============================================
# JULIAN DATE BASED FUNCTIONS
# ============================================

def query_by_julian_date(company, julian_date, running_order=None, binning=None, status=None):
    """Query modules by Julian date with optional filters"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    filtered = []
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        bc_julian = get_julian_from_barcode(barcode)
        
        if bc_julian != julian_date:
            continue
        
        ro = b.get('running_order', '') or ''
        bc_ro = extract_ro_from_ro(ro)
        bc_binning = extract_binning_from_ro(ro)
        is_dispatched = bool(b.get('dispatch_party'))
        
        # Apply filters
        if running_order and bc_ro != running_order.upper():
            continue
        if binning and bc_binning != binning.upper():
            continue
        if status == 'dispatched' and not is_dispatched:
            continue
        if status == 'pending' and is_dispatched:
            continue
        
        filtered.append({
            'barcode': barcode,
            'running_order': bc_ro,
            'binning': bc_binning,
            'pallet': b.get('pallet_no', ''),
            'dispatched': is_dispatched
        })
    
    # Count stats
    total = len(filtered)
    dispatched = len([f for f in filtered if f['dispatched']])
    pending = total - dispatched
    
    # Build answer
    filter_str = f"Julian {julian_date}"
    if running_order:
        filter_str += f" | {running_order}"
    if binning:
        filter_str += f" | {binning}"
    
    mfg_date = julian_to_date(julian_date, 2025)
    
    answer_parts = [f"**📅 {company} - {filter_str}**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"🗓️ **Manufacturing Date:** {mfg_date}")
    answer_parts.append(f"📊 **Total Modules:** {total:,}")
    answer_parts.append(f"🚚 **Dispatched:** {dispatched:,}")
    answer_parts.append(f"⏳ **Remaining:** {pending:,}")
    
    # Breakdown by R-O
    ro_counts = {}
    bin_counts = {}
    for f in filtered:
        if f['running_order']:
            ro_counts[f['running_order']] = ro_counts.get(f['running_order'], 0) + 1
        if f['binning']:
            bin_counts[f['binning']] = bin_counts.get(f['binning'], 0) + 1
    
    if ro_counts:
        answer_parts.append(f"\n\n**🏭 Running Order Breakdown:**")
        for ro, count in sorted(ro_counts.items()):
            answer_parts.append(f"   {ro}: {count:,}")
    
    if bin_counts:
        answer_parts.append(f"\n\n**🏷️ Binning Breakdown:**")
        for b, count in sorted(bin_counts.items()):
            answer_parts.append(f"   {b}: {count:,}")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': filtered}

def get_julian_dates_list(company):
    """Get all Julian dates available for a company"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    julian_stats = {}
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        julian = get_julian_from_barcode(barcode)
        if julian:
            if julian not in julian_stats:
                julian_stats[julian] = {'total': 0, 'dispatched': 0}
            julian_stats[julian]['total'] += 1
            if b.get('dispatch_party'):
                julian_stats[julian]['dispatched'] += 1
    
    answer_parts = [f"**📅 {company} - All Julian Dates**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Julian Dates:** {len(julian_stats)}\n")
    
    for julian in sorted(julian_stats.keys()):
        stats = julian_stats[julian]
        pending = stats['total'] - stats['dispatched']
        date_str = julian_to_date(julian, 2025)
        status_icon = "✅" if pending == 0 else "⏳"
        answer_parts.append(f"{status_icon} **{julian}** ({date_str}): {stats['total']:,} total, {pending:,} pending")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'julian_stats': julian_stats}

def get_oldest_pending_julian(company):
    """Get oldest Julian date with pending modules"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    julian_pending = {}
    for b in mrp_result.get('data', []):
        if b.get('dispatch_party'):
            continue  # Skip dispatched
        
        barcode = b.get('barcode', '')
        julian = get_julian_from_barcode(barcode)
        if julian:
            julian_pending[julian] = julian_pending.get(julian, 0) + 1
    
    if not julian_pending:
        return {'has_answer': True, 'answer': f"✅ **{company}** - No pending modules!"}
    
    oldest = min(julian_pending.keys())
    oldest_date = julian_to_date(oldest, 2025)
    
    answer_parts = [f"**⏰ {company} - Oldest Pending Julian**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📅 **Oldest Julian:** {oldest} ({oldest_date})")
    answer_parts.append(f"⏳ **Pending Count:** {julian_pending[oldest]:,}")
    
    # Age calculation
    today = datetime.now()
    mfg_date = datetime.strptime(oldest_date, '%Y-%m-%d')
    age_days = (today - mfg_date).days
    answer_parts.append(f"📆 **Age:** {age_days} days old")
    
    if age_days > 30:
        answer_parts.append(f"\n⚠️ **WARNING:** Modules older than 30 days!")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts)}

# ============================================
# BARCODE SPECIFIC FUNCTIONS  
# ============================================

def get_barcode_full_status(barcode):
    """Get complete status of a single barcode"""
    # Clean barcode - remove spaces and convert to uppercase
    barcode = barcode.strip().upper()
    
    parsed = parse_barcode(barcode)
    if not parsed:
        return {'has_answer': True, 'answer': f"❌ Invalid barcode format: {barcode}\n\nBarcode should be 18 characters starting with 'GS', e.g., GS04875KG3022500075"}
    
    # Find in all companies' MRP
    found_in_mrp = None
    search_companies = [
        ('Rays Power', 'RAYS POWER INFRA PRIVATE LIMITED'),
        ('Larsen & Toubro', 'LARSEN & TOUBRO LIMITED, CONSTRUCTION'),
        ('Sterlin and Wilson', 'STERLING AND WILSON RENEWABLE ENERGY LIMITED')
    ]
    
    for company_name, mrp_party in search_companies:
        try:
            response = requests.post(
                BARCODE_TRACKING_API,
                json={'party_name': mrp_party},
                timeout=30
            )
            if response.status_code == 200:
                data = response.json()
                if data.get('status') == 'success':
                    for b in data.get('data', []):
                        if b.get('barcode', '').upper() == barcode:
                            found_in_mrp = {
                                **b, 
                                'company': company_name,
                                'mrp_party': mrp_party
                            }
                            break
            if found_in_mrp:
                break
        except Exception as e:
            print(f"Error searching {company_name}: {str(e)}")
            continue
    
    # Check in local database
    db_row = None
    try:
        db_result = db.session.execute(text("""
            SELECT m.serial_number, m.status, m.pdi_number, m.binning, m.class_status, m.pmax,
                   c.company_name 
            FROM ftr_master_serials m
            JOIN companies c ON m.company_id = c.id
            WHERE m.serial_number = :barcode
        """), {'barcode': barcode})
        db_row = db_result.fetchone()
    except Exception as e:
        print(f"Database error: {str(e)}")
    
    # Build answer
    answer_parts = [f"**🔍 Barcode: {barcode}**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    # Parsed info from barcode itself
    answer_parts.append(f"**📋 Barcode Info:**")
    answer_parts.append(f"   🏭 Model: {parsed['model']}")
    answer_parts.append(f"   🏢 Plant: {parsed['plant']}")
    answer_parts.append(f"   📅 Julian: {parsed['julian_date']}")
    answer_parts.append(f"   🗓️ Mfg Date: {parsed['manufacturing_date']}")
    answer_parts.append(f"   🔢 Serial: {parsed['serial']}")
    
    # MRP Status (Production/Packing Info) - Most important
    answer_parts.append(f"\n**📦 Production Status (MRP):**")
    if found_in_mrp:
        company = found_in_mrp.get('company', 'Unknown')
        running_order = found_in_mrp.get('running_order', 'N/A')
        pallet_no = found_in_mrp.get('pallet_no', 'N/A')
        pack_date = found_in_mrp.get('date', 'N/A')
        status = found_in_mrp.get('status', 'N/A')
        dispatch_party = found_in_mrp.get('dispatch_party')
        
        # Extract binning from running_order (e.g., "R-3 i-3" -> "I3")
        mrp_binning = extract_binning_from_ro(running_order)
        mrp_ro = extract_ro_from_ro(running_order)
        
        answer_parts.append(f"   ✅ **Found in MRP System**")
        answer_parts.append(f"   🏭 **Company:** {company}")
        answer_parts.append(f"   🏷️ **Binning:** {mrp_binning or 'N/A'}")
        answer_parts.append(f"   🔢 **Running Order:** {mrp_ro or running_order}")
        answer_parts.append(f"   📦 **Pallet No:** {pallet_no}")
        answer_parts.append(f"   📅 **Pack Date:** {pack_date}")
        
        if dispatch_party:
            answer_parts.append(f"\n   🚚 **STATUS: DISPATCHED**")
            answer_parts.append(f"   📤 Dispatched to: {dispatch_party}")
        else:
            answer_parts.append(f"\n   📦 **STATUS: PACKED** (Not Dispatched)")
    else:
        answer_parts.append(f"   ❌ **Not found in MRP**")
        answer_parts.append(f"   Module is NOT packed yet or barcode is incorrect")
    
    # Database status (FTR Data)
    answer_parts.append(f"\n**💾 FTR Database:**")
    if db_row:
        answer_parts.append(f"   ✅ Found in Database")
        answer_parts.append(f"   Company: {db_row.company_name}")
        answer_parts.append(f"   Status: {db_row.status or 'N/A'}")
        answer_parts.append(f"   PDI: {db_row.pdi_number or 'Not Assigned'}")
        answer_parts.append(f"   DB Binning: {db_row.binning or 'N/A'}")
        answer_parts.append(f"   Class: {db_row.class_status or 'OK'}")
        if db_row.pmax:
            answer_parts.append(f"   Pmax: {db_row.pmax}W")
    else:
        answer_parts.append(f"   ❌ Not in FTR Database")
    
    # Summary at the end
    answer_parts.append(f"\n━━━━━━━━━━━━━━━━━━━━━━━━━━━━")
    if found_in_mrp:
        binning = extract_binning_from_ro(found_in_mrp.get('running_order', ''))
        company = found_in_mrp.get('company', 'Unknown')
        if found_in_mrp.get('dispatch_party'):
            answer_parts.append(f"**✅ Summary: {company} | {binning or 'N/A'} | 🚚 DISPATCHED**")
        else:
            answer_parts.append(f"**✅ Summary: {company} | {binning or 'N/A'} | 📦 PACKED**")
    else:
        answer_parts.append(f"**⚠️ Module not found in packing system**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts)}

# ============================================
# PALLET FUNCTIONS
# ============================================

def get_pallet_full_details(company, pallet_no):
    """Get complete details of a pallet"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    pallet_barcodes = []
    for b in mrp_result.get('data', []):
        if str(b.get('pallet_no', '')) == str(pallet_no):
            parsed = parse_barcode(b.get('barcode', ''))
            pallet_barcodes.append({
                'barcode': b.get('barcode', ''),
                'running_order': extract_ro_from_ro(b.get('running_order', '')),
                'binning': extract_binning_from_ro(b.get('running_order', '')),
                'julian': parsed['julian_date'] if parsed else None,
                'dispatched': bool(b.get('dispatch_party')),
                'date': b.get('date', '')
            })
    
    if not pallet_barcodes:
        return {'has_answer': True, 'answer': f"❌ Pallet {pallet_no} not found for {company}"}
    
    # Analyze pallet
    binnings = set(b['binning'] for b in pallet_barcodes if b['binning'])
    ros = set(b['running_order'] for b in pallet_barcodes if b['running_order'])
    julians = set(b['julian'] for b in pallet_barcodes if b['julian'])
    dispatched = all(b['dispatched'] for b in pallet_barcodes)
    
    answer_parts = [f"**📦 Pallet {pallet_no} - {company}**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Modules:** {len(pallet_barcodes)}")
    answer_parts.append(f"🚚 **Status:** {'DISPATCHED ✅' if dispatched else 'PACKED (Not Dispatched) ⏳'}")
    answer_parts.append(f"🏷️ **Binnings:** {', '.join(sorted(binnings)) if binnings else 'Unknown'}")
    answer_parts.append(f"🏭 **Running Orders:** {', '.join(sorted(ros)) if ros else 'Unknown'}")
    answer_parts.append(f"📅 **Julian Dates:** {', '.join(map(str, sorted(julians))) if julians else 'Unknown'}")
    
    # Quality check
    issues = []
    if len(binnings) > 1:
        issues.append(f"⚠️ MIX BINNING: {', '.join(sorted(binnings))}")
    if len(ros) > 1:
        issues.append(f"⚠️ MIX RUNNING ORDER: {', '.join(sorted(ros))}")
    if len(pallet_barcodes) != 26:
        issues.append(f"⚠️ MODULE COUNT: {len(pallet_barcodes)} (Expected 26)")
    
    if issues:
        answer_parts.append(f"\n\n**🔴 ISSUES FOUND:**")
        for issue in issues:
            answer_parts.append(f"   {issue}")
    else:
        answer_parts.append(f"\n\n✅ **No issues found - Pallet is OK!**")
    
    # Sample barcodes
    answer_parts.append(f"\n\n**📋 Barcodes ({len(pallet_barcodes)}):**")
    for b in pallet_barcodes[:10]:
        answer_parts.append(f"   • {b['barcode']} | {b['running_order']} | {b['binning']}")
    if len(pallet_barcodes) > 10:
        answer_parts.append(f"   ... and {len(pallet_barcodes) - 10} more")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'barcodes': pallet_barcodes}

def get_total_pallets(company, running_order=None, binning=None, pdi_number=None):
    """Get total pallet count with optional filters including PDI"""
    
    # If PDI number specified, get PDI serials first
    pdi_serials = set()
    if pdi_number:
        try:
            company_result = db.session.execute(text(
                "SELECT id FROM companies WHERE company_name LIKE :name"
            ), {'name': f'%{company.split()[0]}%'})
            company_row = company_result.fetchone()
            
            if company_row:
                pdi_result = db.session.execute(text("""
                    SELECT serial_number FROM ftr_master_serials 
                    WHERE company_id = :cid AND pdi_number = :pdi
                    AND (class_status = 'OK' OR class_status IS NULL)
                """), {'cid': company_row[0], 'pdi': pdi_number})
                pdi_serials = {row[0] for row in pdi_result.fetchall()}
        except Exception as e:
            print(f"Error getting PDI serials: {e}")
    
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    pallets = {}
    total_modules = 0
    dispatched_modules = 0
    
    for b in mrp_result.get('data', []):
        barcode = b.get('barcode', '')
        pallet_no = b.get('pallet_no', '')
        if not pallet_no:
            continue
        
        # If PDI filter, check if barcode is in PDI serials
        if pdi_number and pdi_serials and barcode not in pdi_serials:
            continue
        
        ro = extract_ro_from_ro(b.get('running_order', ''))
        bin_val = extract_binning_from_ro(b.get('running_order', ''))
        
        if running_order and ro != running_order.upper():
            continue
        if binning and bin_val != binning.upper():
            continue
        
        total_modules += 1
        is_dispatched = bool(b.get('dispatch_party'))
        if is_dispatched:
            dispatched_modules += 1
        
        if pallet_no not in pallets:
            pallets[pallet_no] = {'count': 0, 'dispatched': False, 'dispatched_count': 0}
        pallets[pallet_no]['count'] += 1
        if is_dispatched:
            pallets[pallet_no]['dispatched_count'] += 1
            # Mark pallet as dispatched if ALL modules dispatched
            if pallets[pallet_no]['dispatched_count'] == pallets[pallet_no]['count']:
                pallets[pallet_no]['dispatched'] = True
    
    total = len(pallets)
    dispatched = len([p for p in pallets.values() if p['dispatched']])
    pending = total - dispatched
    remaining_modules = total_modules - dispatched_modules
    
    filter_str = company
    if pdi_number:
        filter_str += f" | {pdi_number}"
    if running_order:
        filter_str += f" | {running_order}"
    if binning:
        filter_str += f" | {binning}"
    
    answer_parts = [f"**📦 {filter_str} - Pallet Status**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📦 **Total Pallets:** {total:,}")
    answer_parts.append(f"🚚 **Dispatched Pallets:** {dispatched:,}")
    answer_parts.append(f"⏳ **Pending Pallets:** {pending:,}")
    answer_parts.append(f"\n📊 **Total Modules:** {total_modules:,}")
    answer_parts.append(f"🚚 **Dispatched Modules:** {dispatched_modules:,}")
    answer_parts.append(f"⏳ **Remaining Modules:** {remaining_modules:,}")
    
    if total_modules > 0:
        progress = (dispatched_modules / total_modules * 100)
        filled = int(progress / 5)
        bar = "█" * filled + "░" * (20 - filled)
        answer_parts.append(f"\n**[{bar}] {progress:.1f}%**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts)}

# ============================================
# RUNNING ORDER FUNCTIONS
# ============================================

def get_running_order_status(company, running_order):
    """Get complete status of a running order"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    ro_data = {'total': 0, 'dispatched': 0, 'packed': 0, 'binnings': {}, 'julians': {}, 'pallets': set()}
    
    for b in mrp_result.get('data', []):
        ro = extract_ro_from_ro(b.get('running_order', ''))
        if ro != running_order.upper():
            continue
        
        ro_data['total'] += 1
        if b.get('dispatch_party'):
            ro_data['dispatched'] += 1
        else:
            ro_data['packed'] += 1
        
        # Binning breakdown
        binning = extract_binning_from_ro(b.get('running_order', ''))
        if binning:
            ro_data['binnings'][binning] = ro_data['binnings'].get(binning, 0) + 1
        
        # Julian breakdown
        julian = get_julian_from_barcode(b.get('barcode', ''))
        if julian:
            ro_data['julians'][julian] = ro_data['julians'].get(julian, 0) + 1
        
        # Pallet count
        pallet = b.get('pallet_no', '')
        if pallet:
            ro_data['pallets'].add(pallet)
    
    if ro_data['total'] == 0:
        return {'has_answer': True, 'answer': f"❌ {running_order} not found for {company}"}
    
    remaining = ro_data['total'] - ro_data['dispatched']
    progress = (ro_data['dispatched'] / ro_data['total'] * 100) if ro_data['total'] > 0 else 0
    
    answer_parts = [f"**🏭 {company} - {running_order} Status**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Modules:** {ro_data['total']:,}")
    answer_parts.append(f"🚚 **Dispatched:** {ro_data['dispatched']:,}")
    answer_parts.append(f"⏳ **Remaining:** {remaining:,}")
    answer_parts.append(f"📈 **Progress:** {progress:.1f}%")
    answer_parts.append(f"📦 **Total Pallets:** {len(ro_data['pallets']):,}")
    
    # Progress bar
    filled = int(progress / 5)
    bar = "█" * filled + "░" * (20 - filled)
    answer_parts.append(f"\n**[{bar}] {progress:.1f}%**")
    
    # Binning breakdown
    if ro_data['binnings']:
        answer_parts.append(f"\n\n**🏷️ Binning Breakdown:**")
        for b, count in sorted(ro_data['binnings'].items()):
            answer_parts.append(f"   {b}: {count:,}")
    
    # Julian dates
    if ro_data['julians']:
        answer_parts.append(f"\n\n**📅 Julian Dates:**")
        for j, count in sorted(ro_data['julians'].items()):
            answer_parts.append(f"   {j} ({julian_to_date(j, 2025)}): {count:,}")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': ro_data}

def compare_running_orders(company):
    """Compare all running orders for a company"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    ros = {}
    for b in mrp_result.get('data', []):
        ro = extract_ro_from_ro(b.get('running_order', ''))
        if not ro:
            continue
        
        if ro not in ros:
            ros[ro] = {'total': 0, 'dispatched': 0}
        ros[ro]['total'] += 1
        if b.get('dispatch_party'):
            ros[ro]['dispatched'] += 1
    
    answer_parts = [f"**🏭 {company} - Running Order Comparison**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    for ro in sorted(ros.keys()):
        data = ros[ro]
        remaining = data['total'] - data['dispatched']
        progress = (data['dispatched'] / data['total'] * 100) if data['total'] > 0 else 0
        filled = int(progress / 5)
        bar = "█" * filled + "░" * (20 - filled)
        
        answer_parts.append(f"\n**{ro}:**")
        answer_parts.append(f"   Total: {data['total']:,} | Dispatched: {data['dispatched']:,} | Remaining: {remaining:,}")
        answer_parts.append(f"   [{bar}] {progress:.1f}%")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': ros}

# ============================================
# BINNING FUNCTIONS
# ============================================

def get_binning_status(company, binning):
    """Get complete status of a binning type"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    bin_data = {'total': 0, 'dispatched': 0, 'ros': {}, 'julians': {}}
    
    for b in mrp_result.get('data', []):
        bc_binning = extract_binning_from_ro(b.get('running_order', ''))
        if bc_binning != binning.upper():
            continue
        
        bin_data['total'] += 1
        if b.get('dispatch_party'):
            bin_data['dispatched'] += 1
        
        ro = extract_ro_from_ro(b.get('running_order', ''))
        if ro:
            bin_data['ros'][ro] = bin_data['ros'].get(ro, 0) + 1
        
        julian = get_julian_from_barcode(b.get('barcode', ''))
        if julian:
            bin_data['julians'][julian] = bin_data['julians'].get(julian, 0) + 1
    
    if bin_data['total'] == 0:
        return {'has_answer': True, 'answer': f"❌ {binning} not found for {company}"}
    
    remaining = bin_data['total'] - bin_data['dispatched']
    progress = (bin_data['dispatched'] / bin_data['total'] * 100) if bin_data['total'] > 0 else 0
    
    answer_parts = [f"**🏷️ {company} - {binning} Status**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total:** {bin_data['total']:,}")
    answer_parts.append(f"🚚 **Dispatched:** {bin_data['dispatched']:,}")
    answer_parts.append(f"⏳ **Remaining:** {remaining:,}")
    answer_parts.append(f"📈 **Dispatch %:** {progress:.1f}%")
    
    if bin_data['ros']:
        answer_parts.append(f"\n\n**🏭 Running Order Breakdown:**")
        for ro, count in sorted(bin_data['ros'].items()):
            answer_parts.append(f"   {ro}: {count:,}")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': bin_data}

def compare_binnings(company):
    """Compare all binning types for a company"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    bins = {}
    for b in mrp_result.get('data', []):
        binning = extract_binning_from_ro(b.get('running_order', ''))
        if not binning:
            continue
        
        if binning not in bins:
            bins[binning] = {'total': 0, 'dispatched': 0}
        bins[binning]['total'] += 1
        if b.get('dispatch_party'):
            bins[binning]['dispatched'] += 1
    
    answer_parts = [f"**🏷️ {company} - Binning Comparison**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    for binning in sorted(bins.keys()):
        data = bins[binning]
        remaining = data['total'] - data['dispatched']
        progress = (data['dispatched'] / data['total'] * 100) if data['total'] > 0 else 0
        
        answer_parts.append(f"\n**{binning}:**")
        answer_parts.append(f"   Total: {data['total']:,} | Dispatched: {data['dispatched']:,} | Remaining: {remaining:,} | {progress:.1f}%")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': bins}

# ============================================
# COMPANY COMPARISON FUNCTIONS
# ============================================

def compare_companies():
    """Compare all companies"""
    companies = ['Rays Power', 'Larsen & Toubro', 'Sterlin and Wilson']
    comparison = {}
    
    for company in companies:
        mrp_result = get_all_mrp_data(company)
        if mrp_result.get('success'):
            data = mrp_result.get('data', [])
            dispatched = len([b for b in data if b.get('dispatch_party')])
            packed = len([b for b in data if not b.get('dispatch_party')])
            pallets = len(set(b.get('pallet_no', '') for b in data if b.get('pallet_no')))
            
            comparison[company] = {
                'total': len(data),
                'dispatched': dispatched,
                'packed': packed,
                'pallets': pallets
            }
    
    answer_parts = [f"**🏢 Company Comparison**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    
    for company, data in comparison.items():
        progress = (data['dispatched'] / data['total'] * 100) if data['total'] > 0 else 0
        answer_parts.append(f"\n**{company}:**")
        answer_parts.append(f"   📊 Total: {data['total']:,}")
        answer_parts.append(f"   🚚 Dispatched: {data['dispatched']:,}")
        answer_parts.append(f"   📦 Packed: {data['packed']:,}")
        answer_parts.append(f"   🎁 Pallets: {data['pallets']:,}")
        answer_parts.append(f"   📈 Progress: {progress:.1f}%")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'data': comparison}

def get_company_full_status(company):
    """Get complete status of a company"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    data = mrp_result.get('data', [])
    
    stats = {
        'total': len(data),
        'dispatched': 0,
        'packed': 0,
        'ros': {},
        'binnings': {},
        'pallets': set(),
        'julians': {}
    }
    
    for b in data:
        if b.get('dispatch_party'):
            stats['dispatched'] += 1
        else:
            stats['packed'] += 1
        
        ro = extract_ro_from_ro(b.get('running_order', ''))
        if ro:
            stats['ros'][ro] = stats['ros'].get(ro, 0) + 1
        
        binning = extract_binning_from_ro(b.get('running_order', ''))
        if binning:
            stats['binnings'][binning] = stats['binnings'].get(binning, 0) + 1
        
        pallet = b.get('pallet_no', '')
        if pallet:
            stats['pallets'].add(pallet)
        
        julian = get_julian_from_barcode(b.get('barcode', ''))
        if julian:
            stats['julians'][julian] = stats['julians'].get(julian, 0) + 1
    
    answer_parts = [f"**🏢 {company} - Full Status**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📊 **Total Modules:** {stats['total']:,}")
    answer_parts.append(f"🚚 **Dispatched:** {stats['dispatched']:,}")
    answer_parts.append(f"📦 **Packed (Pending):** {stats['packed']:,}")
    answer_parts.append(f"🎁 **Total Pallets:** {len(stats['pallets']):,}")
    
    progress = (stats['dispatched'] / stats['total'] * 100) if stats['total'] > 0 else 0
    filled = int(progress / 5)
    bar = "█" * filled + "░" * (20 - filled)
    answer_parts.append(f"\n**[{bar}] {progress:.1f}% Dispatched**")
    
    # Running Order breakdown
    if stats['ros']:
        answer_parts.append(f"\n\n**🏭 Running Order Breakdown:**")
        for ro, count in sorted(stats['ros'].items()):
            answer_parts.append(f"   {ro}: {count:,}")
    
    # Binning breakdown
    if stats['binnings']:
        answer_parts.append(f"\n\n**🏷️ Binning Breakdown:**")
        for b, count in sorted(stats['binnings'].items()):
            answer_parts.append(f"   {b}: {count:,}")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'stats': stats}

# ============================================
# FULL PALLET AUDIT
# ============================================

def full_pallet_audit(company):
    """Complete pallet integrity verification"""
    mrp_result = get_all_mrp_data(company)
    if not mrp_result.get('success'):
        return {'has_answer': False, 'error': 'MRP API failed'}
    
    # Get rejected barcodes from DB
    company_result = db.session.execute(text(
        "SELECT id FROM companies WHERE company_name LIKE :name"
    ), {'name': f'%{company.split()[0]}%'})
    company_row = company_result.fetchone()
    
    rejected_serials = set()
    if company_row:
        rejected_result = db.session.execute(text("""
            SELECT serial_number FROM ftr_master_serials 
            WHERE company_id = :cid AND class_status = 'REJECTED'
        """), {'cid': company_row[0]})
        rejected_serials = set(row[0] for row in rejected_result.fetchall())
    
    # Analyze pallets
    pallets = {}
    barcode_pallets = {}  # Track which pallets each barcode appears in
    
    for b in mrp_result.get('data', []):
        pallet_no = b.get('pallet_no', '')
        if not pallet_no:
            continue
        
        barcode = b.get('barcode', '')
        ro = b.get('running_order', '') or ''
        
        if pallet_no not in pallets:
            pallets[pallet_no] = {
                'barcodes': [],
                'binnings': set(),
                'ros': set(),
                'julians': set(),
                'has_rejected': False,
                'issues': []
            }
        
        # Track barcode's pallet
        if barcode not in barcode_pallets:
            barcode_pallets[barcode] = []
        barcode_pallets[barcode].append(pallet_no)
        
        binning = extract_binning_from_ro(ro)
        running_order = extract_ro_from_ro(ro)
        julian = get_julian_from_barcode(barcode)
        
        pallets[pallet_no]['barcodes'].append(barcode)
        if binning:
            pallets[pallet_no]['binnings'].add(binning)
        if running_order:
            pallets[pallet_no]['ros'].add(running_order)
        if julian:
            pallets[pallet_no]['julians'].add(julian)
        if barcode in rejected_serials:
            pallets[pallet_no]['has_rejected'] = True
    
    # Check for issues
    issues_summary = {
        'mix_binning': [],
        'mix_ro': [],
        'wrong_count': [],
        'rejected_module': [],
        'duplicate_barcode': []
    }
    
    # Cross-pallet duplicates
    cross_pallet_duplicates = {bc: pals for bc, pals in barcode_pallets.items() if len(pals) > 1}
    
    for pallet_no, data in pallets.items():
        # Mix binning
        known_bins = {b for b in data['binnings'] if b != 'Unknown'}
        if len(known_bins) > 1:
            issues_summary['mix_binning'].append({
                'pallet': pallet_no,
                'binnings': list(data['binnings'])
            })
            data['issues'].append(f"Mix Binning: {', '.join(data['binnings'])}")
        
        # Mix running order
        if len(data['ros']) > 1:
            issues_summary['mix_ro'].append({
                'pallet': pallet_no,
                'ros': list(data['ros'])
            })
            data['issues'].append(f"Mix R-O: {', '.join(data['ros'])}")
        
        # Wrong count
        if len(data['barcodes']) != 26:
            issues_summary['wrong_count'].append({
                'pallet': pallet_no,
                'count': len(data['barcodes'])
            })
            data['issues'].append(f"Count: {len(data['barcodes'])} (Expected 26)")
        
        # Rejected module
        if data['has_rejected']:
            issues_summary['rejected_module'].append({'pallet': pallet_no})
            data['issues'].append("Contains REJECTED module!")
    
    # Cross-pallet duplicates
    for bc, pals in cross_pallet_duplicates.items():
        issues_summary['duplicate_barcode'].append({
            'barcode': bc,
            'pallets': pals
        })
    
    # Build answer
    total_pallets = len(pallets)
    pallets_with_issues = len([p for p in pallets.values() if p['issues']])
    
    answer_parts = [f"**🔍 {company} - Full Pallet Audit**\n"]
    answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
    answer_parts.append(f"📦 **Total Pallets:** {total_pallets:,}")
    answer_parts.append(f"✅ **Clean Pallets:** {total_pallets - pallets_with_issues:,}")
    answer_parts.append(f"⚠️ **Pallets with Issues:** {pallets_with_issues:,}")
    
    answer_parts.append(f"\n\n**📋 Issue Summary:**")
    answer_parts.append(f"   🔸 Mix Binning: {len(issues_summary['mix_binning'])}")
    answer_parts.append(f"   🔸 Mix Running Order: {len(issues_summary['mix_ro'])}")
    answer_parts.append(f"   🔸 Wrong Module Count: {len(issues_summary['wrong_count'])}")
    answer_parts.append(f"   🔴 Rejected Modules: {len(issues_summary['rejected_module'])}")
    answer_parts.append(f"   🔸 Cross-Pallet Duplicates: {len(issues_summary['duplicate_barcode'])}")
    
    # Show details
    if issues_summary['mix_binning']:
        answer_parts.append(f"\n\n**❌ Mix Binning Pallets:**")
        for p in issues_summary['mix_binning'][:10]:
            answer_parts.append(f"   Pallet {p['pallet']}: {', '.join(p['binnings'])}")
    
    if issues_summary['rejected_module']:
        answer_parts.append(f"\n\n**🔴 REJECTED MODULE IN PALLET:**")
        for p in issues_summary['rejected_module'][:10]:
            answer_parts.append(f"   Pallet {p['pallet']}")
    
    if issues_summary['duplicate_barcode']:
        answer_parts.append(f"\n\n**⚠️ Cross-Pallet Duplicates:**")
        for d in issues_summary['duplicate_barcode'][:10]:
            answer_parts.append(f"   {d['barcode']} in Pallets: {', '.join(map(str, d['pallets']))}")
    
    if pallets_with_issues == 0:
        answer_parts.append(f"\n\n✅ **All pallets passed audit! No issues found.**")
    
    return {'has_answer': True, 'answer': "\n".join(answer_parts), 'issues': issues_summary}

def get_packed_not_in_pdi(company=None):
    """
    Get all packed modules (from MRP) that are NOT assigned to any PDI
    Shows company-wise breakdown with pallet numbers
    """
    import re
    
    try:
        all_companies = ['Rays Power', 'Larsen & Toubro', 'Sterlin and Wilson']
        companies_to_check = [company] if company else all_companies
        
        results = {}
        total_packed_not_pdi = 0
        
        for comp in companies_to_check:
            # Get MRP data
            mrp_result = get_all_mrp_data(comp)
            if not mrp_result.get('success'):
                continue
            
            # Get company_id
            company_result = db.session.execute(text(
                "SELECT id FROM companies WHERE company_name LIKE :name"
            ), {'name': f'%{comp.split()[0]}%'})
            company_row = company_result.fetchone()
            
            if not company_row:
                continue
            
            company_id = company_row[0]
            
            # Get all serials in database with PDI or rejected
            db_result = db.session.execute(text("""
                SELECT serial_number, pdi_number, class_status
                FROM ftr_master_serials
                WHERE company_id = :cid
            """), {'cid': company_id})
            
            db_serials = {}
            for row in db_result.fetchall():
                db_serials[row[0]] = {
                    'pdi': row[1],
                    'class_status': row[2]
                }
            
            # Filter MRP packed modules
            packed_not_pdi = []
            pallet_breakdown = {}
            
            for b in mrp_result.get('data', []):
                barcode = b.get('barcode', '')
                pallet_no = b.get('pallet_no', '')
                status = b.get('status', '')
                dispatch_party = b.get('dispatch_party')
                ro = b.get('running_order', '')
                
                # Check if packed (not dispatched)
                if status == 'packed' and not dispatch_party:
                    # Check if NOT in PDI and NOT rejected
                    db_info = db_serials.get(barcode, {})
                    has_pdi = db_info.get('pdi')
                    is_rejected = db_info.get('class_status') == 'REJECTED'
                    
                    if not has_pdi and not is_rejected:
                        # Extract binning from running_order
                        bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
                        binning = f"I{bin_match.group(1)}" if bin_match else 'Unknown'
                        
                        packed_not_pdi.append({
                            'barcode': barcode,
                            'pallet': pallet_no,
                            'running_order': ro,
                            'binning': binning
                        })
                        
                        # Pallet breakdown
                        if pallet_no:
                            if pallet_no not in pallet_breakdown:
                                pallet_breakdown[pallet_no] = {
                                    'count': 0,
                                    'binnings': set()
                                }
                            pallet_breakdown[pallet_no]['count'] += 1
                            pallet_breakdown[pallet_no]['binnings'].add(binning)
            
            results[comp] = {
                'total': len(packed_not_pdi),
                'pallets': len(pallet_breakdown),
                'samples': packed_not_pdi[:10],
                'pallet_breakdown': {k: {'count': v['count'], 'binnings': list(v['binnings'])} 
                                     for k, v in sorted(pallet_breakdown.items(), key=lambda x: int(x[0]) if x[0].isdigit() else 0)[:10]}
            }
            total_packed_not_pdi += len(packed_not_pdi)
        
        # Build answer
        answer_parts = [f"**📦 Packed Modules (NOT in PDI & NOT Rejected)**\n"]
        answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
        answer_parts.append(f"🔢 **Total:** {total_packed_not_pdi:,} modules\n")
        
        for comp, data in results.items():
            if data['total'] > 0:
                answer_parts.append(f"\n**🏭 {comp}:** {data['total']:,} modules | {data['pallets']} pallets")
                
                # Top pallets
                if data['pallet_breakdown']:
                    answer_parts.append(f"\n   **Top Pallets:**")
                    for pallet, info in list(data['pallet_breakdown'].items())[:5]:
                        binnings_str = ', '.join(info['binnings'])
                        answer_parts.append(f"   • Pallet {pallet}: {info['count']} modules ({binnings_str})")
                
                # Sample barcodes
                if data['samples']:
                    answer_parts.append(f"\n   **Sample Barcodes:**")
                    for s in data['samples'][:3]:
                        answer_parts.append(f"   • {s['barcode']} | Pallet: {s['pallet']} | {s['binning']}")
        
        if total_packed_not_pdi == 0:
            answer_parts.append(f"\n✅ **All packed modules are either assigned to PDI or rejected.**")
        
        return {
            'has_answer': True,
            'answer': "\n".join(answer_parts),
            'data': results,
            'total': total_packed_not_pdi
        }
        
    except Exception as e:
        print(f"Error in get_packed_not_in_pdi: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'has_answer': False, 'error': str(e)}

def compare_pdi_with_mrp(company, pdi_number):
    """
    Compare PDI serial numbers from database with MRP system
    Returns exact dispatch/remaining counts
    """
    import re
    
    try:
        # Step 1: Get company_id
        company_result = db.session.execute(text("""
            SELECT id FROM companies WHERE company_name LIKE :name
        """), {'name': f'%{company.split()[0]}%'})
        company_row = company_result.fetchone()
        
        if not company_row:
            return {'has_answer': False, 'error': 'Company not found'}
        
        company_id = company_row[0]
        
        # Step 2: Get all serial numbers assigned to this PDI
        pdi_serials_result = db.session.execute(text("""
            SELECT serial_number, binning, class_status
            FROM ftr_master_serials 
            WHERE company_id = :cid AND pdi_number = :pdi
            AND (class_status = 'OK' OR class_status IS NULL)
        """), {'cid': company_id, 'pdi': pdi_number})
        
        pdi_serials = []
        binning_breakdown = {}
        for row in pdi_serials_result.fetchall():
            serial = row[0]
            binning = row[1] or 'Unknown'
            pdi_serials.append(serial)
            binning_breakdown[binning] = binning_breakdown.get(binning, 0) + 1
        
        total_pdi = len(pdi_serials)
        
        if total_pdi == 0:
            return {
                'has_answer': True,
                'answer': f"**{company} - {pdi_number}**\n\n❌ No serial numbers found for {pdi_number}"
            }
        
        # Step 3: Get MRP data for comparison
        mrp_party_name = get_mrp_party_name(company)
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=60
        )
        
        if response.status_code != 200:
            return {'has_answer': False, 'error': 'MRP API failed'}
        
        mrp_data = response.json()
        mrp_barcodes = mrp_data.get('data', [])
        
        # Create lookup set for MRP barcodes
        mrp_dispatched = set()
        mrp_packed = set()
        mrp_barcode_details = {}
        
        for b in mrp_barcodes:
            barcode = b.get('barcode', '')
            ro = b.get('running_order', '') or ''
            
            # Extract binning from running_order
            bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
            mrp_binning = f"I{bin_match.group(1)}" if bin_match else ''
            
            mrp_barcode_details[barcode] = {
                'running_order': ro,
                'binning': mrp_binning,
                'pallet': b.get('pallet_no', ''),
                'date': b.get('date', '')
            }
            
            if b.get('dispatch_party') or b.get('status') == 'dispatched':
                mrp_dispatched.add(barcode)
            elif b.get('status') == 'packed':
                mrp_packed.add(barcode)
        
        # Step 4: Compare PDI serials with MRP
        dispatched_count = 0
        packed_count = 0
        remaining_count = 0
        not_in_mrp = 0
        
        dispatched_serials = []
        packed_serials = []
        remaining_serials = []
        
        for serial in pdi_serials:
            if serial in mrp_dispatched:
                dispatched_count += 1
                dispatched_serials.append(serial)
            elif serial in mrp_packed:
                packed_count += 1
                packed_serials.append(serial)
            else:
                remaining_count += 1
                remaining_serials.append(serial)
                if serial not in mrp_barcode_details:
                    not_in_mrp += 1
        
        # Build answer
        answer_parts = []
        answer_parts.append(f"**🏭 {company} - {pdi_number}**\n")
        answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
        answer_parts.append(f"📊 **Total Serial Numbers:** {total_pdi:,}\n")
        answer_parts.append(f"\n**MRP Comparison Results:**\n")
        answer_parts.append(f"🚚 **Dispatched:** {dispatched_count:,} ({(dispatched_count/total_pdi*100):.1f}%)")
        answer_parts.append(f"📦 **Packed (Not Dispatched):** {packed_count:,}")
        answer_parts.append(f"⏳ **Remaining (Bacha):** {remaining_count:,}")
        
        if not_in_mrp > 0:
            answer_parts.append(f"⚠️ **Not in MRP System:** {not_in_mrp:,}")
        
        # Binning breakdown
        if binning_breakdown:
            answer_parts.append(f"\n\n**📊 Binning Breakdown (Database):**")
            for bin_type, count in sorted(binning_breakdown.items()):
                answer_parts.append(f"   {bin_type}: {count:,}")
        
        # Sample dispatched
        if dispatched_serials:
            answer_parts.append(f"\n\n**✅ Sample Dispatched ({min(5, len(dispatched_serials))} of {dispatched_count}):**")
            for s in dispatched_serials[:5]:
                details = mrp_barcode_details.get(s, {})
                answer_parts.append(f"   • {s} | {details.get('running_order', '')} | Pallet: {details.get('pallet', '')}")
        
        # Sample packed
        if packed_serials:
            answer_parts.append(f"\n\n**📦 Sample Packed ({min(5, len(packed_serials))} of {packed_count}):**")
            for s in packed_serials[:5]:
                details = mrp_barcode_details.get(s, {})
                answer_parts.append(f"   • {s} | {details.get('running_order', '')} | Pallet: {details.get('pallet', '')}")
        
        # Sample remaining
        if remaining_serials:
            answer_parts.append(f"\n\n**⏳ Sample Remaining ({min(5, len(remaining_serials))} of {remaining_count}):**")
            for s in remaining_serials[:5]:
                answer_parts.append(f"   • {s}")
        
        return {
            'has_answer': True,
            'answer': "\n".join(answer_parts),
            'data': {
                'total': total_pdi,
                'dispatched': dispatched_count,
                'packed': packed_count,
                'remaining': remaining_count,
                'not_in_mrp': not_in_mrp
            }
        }
        
    except Exception as e:
        print(f"Error comparing PDI with MRP: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'has_answer': False, 'error': str(e)}

def compare_pdi_with_mrp_filtered(company, pdi_number, running_orders=None):
    """
    Compare PDI with MRP but filter by specific running orders
    running_orders: list of running orders like ['R-1', 'R-2'] or single string 'R-1'
    """
    import re
    
    try:
        # Normalize running_orders to list
        if running_orders:
            if isinstance(running_orders, str):
                running_orders = [running_orders]
            running_orders = [ro.upper() for ro in running_orders]
        
        print(f"[DEBUG] compare_pdi_with_mrp_filtered: company={company}, pdi={pdi_number}, running_orders={running_orders}")
        
        # Step 1: Get company_id
        company_result = db.session.execute(text("""
            SELECT id FROM companies WHERE company_name LIKE :name
        """), {'name': f'%{company.split()[0]}%'})
        company_row = company_result.fetchone()
        
        if not company_row:
            return {'has_answer': False, 'error': 'Company not found'}
        
        company_id = company_row[0]
        
        # Step 2: Get all serial numbers assigned to this PDI
        pdi_serials_result = db.session.execute(text("""
            SELECT serial_number, binning, class_status
            FROM ftr_master_serials 
            WHERE company_id = :cid AND pdi_number = :pdi
            AND (class_status = 'OK' OR class_status IS NULL)
        """), {'cid': company_id, 'pdi': pdi_number})
        
        pdi_serials = []
        binning_breakdown = {}
        for row in pdi_serials_result.fetchall():
            serial = row[0]
            binning = row[1] or 'Unknown'
            pdi_serials.append(serial)
            binning_breakdown[binning] = binning_breakdown.get(binning, 0) + 1
        
        total_pdi = len(pdi_serials)
        
        if total_pdi == 0:
            return {
                'has_answer': True,
                'answer': f"**{company} - {pdi_number}**\n\n❌ No serial numbers found for {pdi_number}"
            }
        
        print(f"[DEBUG] Found {total_pdi} serials in {pdi_number}. Fetching MRP data...")
        
        # Step 3: Get MRP data for comparison
        mrp_party_name = get_mrp_party_name(company)
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=30  # Reduced from 60 seconds
        )
        
        if response.status_code != 200:
            return {'has_answer': False, 'error': 'MRP API failed'}
        
        print(f"[DEBUG] MRP API response received. Processing...")
        
        mrp_data = response.json()
        mrp_barcodes = mrp_data.get('data', [])
        
        # OPTIMIZATION: Convert PDI serials to set for O(1) lookup
        pdi_serials_set = set(pdi_serials)
        
        # Create lookup dictionaries - ONLY for PDI serials (much faster)
        mrp_dispatched = set()
        mrp_packed = set()
        mrp_barcode_details = {}
        
        filtered_count = 0
        total_mrp_count = len(mrp_barcodes)
        
        print(f"[DEBUG] Processing {total_mrp_count} MRP barcodes, filtering for {total_pdi} PDI serials...")
        
        for b in mrp_barcodes:
            barcode = b.get('barcode', '')
            
            # CRITICAL OPTIMIZATION: Skip barcodes not in PDI
            if barcode not in pdi_serials_set:
                continue
            
            ro = b.get('running_order', '') or ''
            
            # Extract R-O from running_order field
            ro_match = re.search(r'(R-\d+)', ro, re.IGNORECASE)
            extracted_ro = ro_match.group(1).upper() if ro_match else ''
            
            # Extract binning from running_order
            bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
            mrp_binning = f"I{bin_match.group(1)}" if bin_match else ''
            
            # Filter by running order if specified
            if running_orders and extracted_ro not in running_orders:
                continue
            
            filtered_count += 1
            
            mrp_barcode_details[barcode] = {
                'running_order': ro,
                'extracted_ro': extracted_ro,
                'binning': mrp_binning,
                'pallet': b.get('pallet_no', ''),
                'date': b.get('date', '')
            }
            
            if b.get('dispatch_party') or b.get('status') == 'dispatched':
                mrp_dispatched.add(barcode)
            elif b.get('status') == 'packed':
                mrp_packed.add(barcode)
        
        print(f"[DEBUG] Total MRP barcodes: {total_mrp_count}, Filtered by RO: {filtered_count}")
        
        # Step 4: Compare PDI serials with FILTERED MRP
        print(f"[DEBUG] Step 4: Comparing {total_pdi} PDI serials with {filtered_count} filtered MRP barcodes...")
        dispatched_count = 0
        packed_count = 0
        remaining_count = 0
        not_in_filtered_mrp = 0
        
        dispatched_serials = []
        packed_serials = []
        remaining_serials = []
        
        ro_breakdown = {}
        
        print(f"[DEBUG] Starting serial comparison loop for {len(pdi_serials)} serials...")
        for serial in pdi_serials:
            if serial in mrp_dispatched:
                dispatched_count += 1
                dispatched_serials.append(serial)
                ro = mrp_barcode_details.get(serial, {}).get('extracted_ro', 'Unknown')
                ro_breakdown[ro] = ro_breakdown.get(ro, {'dispatched': 0, 'packed': 0})
                ro_breakdown[ro]['dispatched'] += 1
            elif serial in mrp_packed:
                packed_count += 1
                packed_serials.append(serial)
                ro = mrp_barcode_details.get(serial, {}).get('extracted_ro', 'Unknown')
                ro_breakdown[ro] = ro_breakdown.get(ro, {'dispatched': 0, 'packed': 0})
                ro_breakdown[ro]['packed'] += 1
            else:
                remaining_count += 1
                remaining_serials.append(serial)
                if serial not in mrp_barcode_details:
                    not_in_filtered_mrp += 1
        
        print(f"[DEBUG] Comparison complete: Dispatched={dispatched_count}, Packed={packed_count}, Remaining={remaining_count}")
        
        # Build answer
        answer_parts = []
        ro_filter_str = " + ".join(running_orders) if running_orders else "All"
        answer_parts.append(f"**🏭 {company} - {pdi_number} | {ro_filter_str}**\n")
        answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
        answer_parts.append(f"📊 **Total Serial Numbers in {pdi_number}:** {total_pdi:,}\n")
        answer_parts.append(f"\n**MRP Status (Filtered by {ro_filter_str}):**\n")
        answer_parts.append(f"🚚 **Dispatched:** {dispatched_count:,}")
        answer_parts.append(f"📦 **Packed (Not Dispatched):** {packed_count:,}")
        answer_parts.append(f"⏳ **Remaining/Not in {ro_filter_str}:** {remaining_count:,}")
        
        # Running Order breakdown
        if ro_breakdown:
            answer_parts.append(f"\n\n**📋 Running Order Breakdown:**")
            for ro, counts in sorted(ro_breakdown.items()):
                total_ro = counts['dispatched'] + counts['packed']
                answer_parts.append(f"   **{ro}:** {total_ro:,} (🚚 {counts['dispatched']:,} | 📦 {counts['packed']:,})")
        
        # Sample dispatched
        if dispatched_serials:
            answer_parts.append(f"\n\n**✅ Sample Dispatched ({min(5, len(dispatched_serials))} of {dispatched_count}):**")
            for s in dispatched_serials[:5]:
                details = mrp_barcode_details.get(s, {})
                answer_parts.append(f"   • {s} | {details.get('running_order', '')} | Pallet: {details.get('pallet', '')}")
        
        # Sample packed
        if packed_serials:
            answer_parts.append(f"\n\n**📦 Sample Packed ({min(5, len(packed_serials))} of {packed_count}):**")
            for s in packed_serials[:5]:
                details = mrp_barcode_details.get(s, {})
                answer_parts.append(f"   • {s} | {details.get('running_order', '')} | Pallet: {details.get('pallet', '')}")
        
        # Sample remaining
        if remaining_serials:
            answer_parts.append(f"\n\n**⏳ Sample Not in {ro_filter_str} ({min(5, len(remaining_serials))} of {remaining_count}):**")
            for s in remaining_serials[:5]:
                answer_parts.append(f"   • {s}")
        
        return {
            'has_answer': True,
            'answer': "\n".join(answer_parts),
            'data': {
                'total_pdi': total_pdi,
                'dispatched': dispatched_count,
                'packed': packed_count,
                'remaining': remaining_count,
                'ro_breakdown': ro_breakdown
            }
        }
        
    except Exception as e:
        print(f"Error in compare_pdi_with_mrp_filtered: {str(e)}")
        import traceback
        traceback.print_exc()
        return {'has_answer': False, 'error': str(e)}

def parse_user_query(message):
    """
    Parse user query to understand EXACTLY what they're asking
    Returns: dict with extracted parameters
    """
    message_lower = message.lower()
    
    result = {
        'company': None,
        'pdi_number': None,
        'running_order': None,
        'multiple_running_orders': False,
        'binning': None,
        'pallet_no': None,
        'julian_date': None,
        'julian_range': None,
        'specific_barcode': None,
        'count_needed': 0,
        # Intent flags
        'wants_dispatch_count': False,
        'wants_packed_count': False,
        'wants_remaining': False,
        'wants_pallet_details': False,
        'wants_barcode_list': False,
        'wants_mix_packing_check': False,
        'wants_duplicate_check': False,
        'wants_binning_mismatch': False,
        'wants_rejected_check': False,
        'wants_missing_in_mrp': False,
        'wants_extra_in_mrp': False,
        'wants_julian_query': False,
        'wants_julian_list': False,
        'wants_oldest_pending': False,
        'wants_barcode_status': False,
        'wants_ro_status': False,
        'wants_ro_comparison': False,
        'wants_binning_status': False,
        'wants_binning_comparison': False,
        'wants_company_comparison': False,
        'wants_company_status': False,
        'wants_pallet_audit': False,
        'wants_pallet_count': False
    }
    
    # Company detection (more patterns)
    company_patterns = {
        'Larsen & Toubro': r'l&t|larsen|toubro|lnt|l \& t',
        'Rays Power': r'rays|rp|rays\s*power',
        'Sterlin and Wilson': r'sterlin|wilson|sw|s&w|s\&w|sterling'
    }
    for company, pattern in company_patterns.items():
        if re.search(pattern, message_lower):
            result['company'] = company
            break
    
    # PDI number detection
    pdi_match = re.search(r'pdi[- ]?(\d+)', message_lower)
    if pdi_match:
        result['pdi_number'] = f"PDI-{pdi_match.group(1)}"
    
    # Running Order detection - support multiple (R1 aur R2)
    ro_matches = re.findall(r'r[- ]?(\d+)', message_lower)
    if ro_matches:
        if len(ro_matches) == 1:
            result['running_order'] = f"R-{ro_matches[0]}"
        else:
            # Multiple ROs detected
            result['running_order'] = [f"R-{ro}" for ro in ro_matches]
            result['multiple_running_orders'] = True
    
    # Binning detection (I1, I2, I3 or MB, MC, MD, MF, MG - NOT 'me' as it's Hindi word)
    # Use negative lookbehind to NOT match PDI-1 as binning I1
    bin_match = re.search(r'(?<!pd)\bi[- ]?(\d+)\b', message_lower)
    if bin_match:
        result['binning'] = f"I{bin_match.group(1)}"
    else:
        # Alpha binnings like MB, MC, MD (excluding 'me' - Hindi word for 'in')
        bin_alpha = re.search(r'\b(mb|mc|md|mf|mg)\s', message_lower)
        if bin_alpha:
            result['binning'] = bin_alpha.group(1).upper()
    
    # Pallet detection
    pallet_match = re.search(r'pallet[- ]?(\d+)|pallet\s*no\.?\s*(\d+)', message_lower)
    if pallet_match:
        result['pallet_no'] = pallet_match.group(1) or pallet_match.group(2)
        result['wants_pallet_details'] = True
    
    # Julian date detection
    julian_match = re.search(r'julian[- ]?(\d{3})|(\d{3})\s*julian', message_lower)
    if julian_match:
        result['julian_date'] = int(julian_match.group(1) or julian_match.group(2))
        result['wants_julian_query'] = True
    
    # Julian range detection (300 se 310)
    julian_range_match = re.search(r'(\d{3})\s*(?:se|to|-)\s*(\d{3})', message_lower)
    if julian_range_match:
        result['julian_range'] = (int(julian_range_match.group(1)), int(julian_range_match.group(2)))
        result['wants_julian_query'] = True
    
    # Specific barcode check (GS format - more flexible matching)
    # Matches: GS04875KG3022544039, gs04875kg3022544039, GS 04875 KG 302 25 44039
    barcode_match = re.search(r'(GS\s*\d{5}\s*[A-Z]{2}\s*\d{3}\s*\d{2}\s*\d{5})', message.replace(' ', ''), re.IGNORECASE)
    if not barcode_match:
        # Try simpler pattern - just GS followed by 16+ alphanumeric
        barcode_match = re.search(r'(GS[A-Z0-9]{16,18})', message.replace(' ', ''), re.IGNORECASE)
    if barcode_match:
        result['specific_barcode'] = barcode_match.group(1).upper().replace(' ', '')
        result['wants_barcode_status'] = True
    
    # Count needed
    count_match = re.search(r'(\d+)\s*(barcode|serial|module|piece)', message_lower)
    if count_match:
        result['count_needed'] = int(count_match.group(1))
    
    # Intent detection - Keywords
    dispatch_words = ['dispatch', 'dispatched', 'bheja', 'bhej', 'sent', 'ship', 'nikla', 'nikal', 'gaya', 'gaye', 'deliver']
    packed_words = ['pack', 'packed', 'packing', 'ban', 'bana', 'ready']
    remaining_words = ['remaining', 'bacha', 'baki', 'pending', 'left', 'kitna', 'kitne', 'baaki', 'rest']
    
    result['wants_dispatch_count'] = any(w in message_lower for w in dispatch_words)
    result['wants_packed_count'] = any(w in message_lower for w in packed_words)
    result['wants_remaining'] = any(w in message_lower for w in remaining_words)
    
    # Quality check intents
    if 'duplicate' in message_lower or 'dohra' in message_lower:
        result['wants_duplicate_check'] = True
    
    if 'mismatch' in message_lower or ('db' in message_lower and 'mrp' in message_lower) or 'binning' in message_lower and 'check' in message_lower:
        result['wants_binning_mismatch'] = True
    
    if 'reject' in message_lower and ('pack' in message_lower or 'check' in message_lower):
        result['wants_rejected_check'] = True
    
    if ('missing' in message_lower or 'nahi' in message_lower) and 'mrp' in message_lower:
        result['wants_missing_in_mrp'] = True
    
    if ('extra' in message_lower or 'zyada' in message_lower) and 'mrp' in message_lower:
        result['wants_extra_in_mrp'] = True
    
    # Mix packing check
    if 'mix' in message_lower and ('packing' in message_lower or 'pallet' in message_lower):
        result['wants_mix_packing_check'] = True
    
    # Julian queries
    if 'julian' in message_lower:
        if 'list' in message_lower or 'sab' in message_lower or 'all' in message_lower or 'kon' in message_lower:
            result['wants_julian_list'] = True
        if 'purana' in message_lower or 'oldest' in message_lower or 'old' in message_lower:
            result['wants_oldest_pending'] = True
    
    # Running Order queries
    if result['running_order']:
        if 'status' in message_lower or 'progress' in message_lower or 'complete' in message_lower:
            result['wants_ro_status'] = True
    
    if 'vs' in message_lower or 'comparison' in message_lower or 'compare' in message_lower:
        if result['running_order'] or 'r-1' in message_lower or 'r-2' in message_lower:
            result['wants_ro_comparison'] = True
        if result['binning'] or 'i1' in message_lower or 'i2' in message_lower:
            result['wants_binning_comparison'] = True
        if 'company' in message_lower or ('rays' in message_lower and 'l&t' in message_lower):
            result['wants_company_comparison'] = True
    
    # Binning queries
    if result['binning'] and not result['wants_binning_comparison']:
        if 'status' in message_lower or 'kitna' in message_lower or 'total' in message_lower:
            result['wants_binning_status'] = True
    
    # Company queries
    if ('full' in message_lower or 'complete' in message_lower or 'status' in message_lower) and result['company']:
        if not result['running_order'] and not result['binning'] and not result['pdi_number']:
            result['wants_company_status'] = True
    
    # Pallet queries
    if 'audit' in message_lower or 'verify' in message_lower or 'sare pallet' in message_lower:
        result['wants_pallet_audit'] = True
    
    # Pallet count - detect "pallet kitne", "kitne pallet", "pallet + remaining/baaki"
    if 'pallet' in message_lower:
        if any(w in message_lower for w in ['kitne', 'kitna', 'count', 'total', 'remaining', 'baaki', 'baki', 'pending', 'hai']):
            result['wants_pallet_count'] = True
    
    # Packed but not in PDI (orphan packed modules)
    if ('packed' in message_lower or 'pack' in message_lower) and ('pdi' in message_lower or 'rejaction' in message_lower or 'rejection' in message_lower):
        if any(w in message_lower for w in ['na', 'not', 'nahi', 'without', 'bina', 'jo nahi']):
            result['wants_packed_not_pdi'] = True
    
    return result

def get_specific_mrp_data(company, filters=None):
    """
    Get MRP data with specific filters
    filters: {'running_order': 'R-3', 'binning': 'I2', 'pallet_no': '123', 'status': 'dispatched'}
    """
    import re
    filters = filters or {}
    
    mrp_party_name = get_mrp_party_name(company)
    
    try:
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=30
        )
        
        if response.status_code != 200:
            return {'success': False, 'error': 'MRP API failed'}
        
        data = response.json()
        if data.get('status') != 'success':
            return {'success': False, 'error': 'No data from MRP'}
        
        all_barcodes = data.get('data', [])
        
        # Apply filters
        filtered = all_barcodes
        
        # Filter by running order
        if filters.get('running_order'):
            ro_filter = filters['running_order'].upper()
            filtered = [b for b in filtered if ro_filter in (b.get('running_order', '') or '').upper()]
        
        # Filter by binning
        if filters.get('binning'):
            bin_filter = filters['binning'].upper()
            new_filtered = []
            for b in filtered:
                ro = b.get('running_order', '') or ''
                bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
                if bin_match:
                    barcode_bin = f"I{bin_match.group(1)}"
                    if barcode_bin.upper() == bin_filter:
                        new_filtered.append(b)
            filtered = new_filtered
        
        # Filter by pallet
        if filters.get('pallet_no'):
            pallet = str(filters['pallet_no'])
            filtered = [b for b in filtered if str(b.get('pallet_no', '')) == pallet]
        
        # Count by status
        dispatched = [b for b in filtered if b.get('dispatch_party') or b.get('status') == 'dispatched']
        packed = [b for b in filtered if b.get('status') == 'packed' and not b.get('dispatch_party')]
        pending = [b for b in filtered if b.get('status') != 'packed' and not b.get('dispatch_party')]
        
        # Running order breakdown
        ro_breakdown = {}
        bin_breakdown = {}
        for b in filtered:
            ro = b.get('running_order', '') or ''
            ro_match = re.search(r'(R-\d+)', ro, re.IGNORECASE)
            if ro_match:
                ro_key = ro_match.group(1).upper()
                ro_breakdown[ro_key] = ro_breakdown.get(ro_key, 0) + 1
            
            bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
            if bin_match:
                bin_key = f"I{bin_match.group(1)}"
                bin_breakdown[bin_key] = bin_breakdown.get(bin_key, 0) + 1
        
        return {
            'success': True,
            'total': len(filtered),
            'dispatched_count': len(dispatched),
            'packed_count': len(packed),
            'pending_count': len(pending),
            'remaining': len(filtered) - len(dispatched),  # Total minus dispatched
            'dispatched_list': dispatched[:50],  # Sample
            'packed_list': packed[:50],
            'pending_list': pending[:50],
            'all_filtered': filtered,
            'ro_breakdown': ro_breakdown,
            'bin_breakdown': bin_breakdown
        }
        
    except Exception as e:
        return {'success': False, 'error': str(e)}

def answer_specific_query(parsed_query):
    """
    Generate a SPECIFIC answer based on parsed query
    Routes to appropriate handler function based on detected intent
    """
    company = parsed_query.get('company')
    pdi = parsed_query.get('pdi_number')
    ro = parsed_query.get('running_order')
    binning = parsed_query.get('binning')
    pallet = parsed_query.get('pallet_no')
    julian = parsed_query.get('julian_date')
    barcode = parsed_query.get('specific_barcode')
    
    # ===== BARCODE STATUS (no company needed) =====
    if barcode and parsed_query.get('wants_barcode_status'):
        return get_barcode_full_status(barcode)
    
    # ===== COMPANY COMPARISON (no single company needed) =====
    if parsed_query.get('wants_company_comparison'):
        return compare_companies()
    
    # ===== Rest need company =====
    if not company:
        return {'has_answer': False}
    
    # ===== QUALITY CHECK QUERIES =====
    if parsed_query.get('wants_duplicate_check'):
        return check_duplicate_barcodes(company)
    
    if parsed_query.get('wants_binning_mismatch'):
        return check_binning_mismatch(company)
    
    if parsed_query.get('wants_rejected_check'):
        return check_rejected_packed(company)
    
    if parsed_query.get('wants_missing_in_mrp'):
        return check_missing_in_mrp(company)
    
    if parsed_query.get('wants_extra_in_mrp'):
        return check_extra_in_mrp(company)
    
    if parsed_query.get('wants_mix_packing_check'):
        return check_mix_packing(company)
    
    if parsed_query.get('wants_pallet_audit'):
        return full_pallet_audit(company)
    
    # ===== PACKED NOT IN PDI =====
    if parsed_query.get('wants_packed_not_pdi'):
        return get_packed_not_in_pdi(company)
    
    # ===== JULIAN DATE QUERIES =====
    if parsed_query.get('wants_julian_list'):
        return get_julian_dates_list(company)
    
    if parsed_query.get('wants_oldest_pending'):
        return get_oldest_pending_julian(company)
    
    if julian and parsed_query.get('wants_julian_query'):
        return query_by_julian_date(company, julian, ro, binning)
    
    # ===== PALLET QUERIES =====
    if pallet and parsed_query.get('wants_pallet_details'):
        return get_pallet_full_details(company, pallet)
    
    if parsed_query.get('wants_pallet_count'):
        return get_total_pallets(company, ro, binning, pdi)
    
    # ===== RUNNING ORDER QUERIES =====
    if parsed_query.get('wants_ro_comparison'):
        return compare_running_orders(company)
    
    if ro and parsed_query.get('wants_ro_status'):
        return get_running_order_status(company, ro)
    
    # ===== BINNING QUERIES =====
    if parsed_query.get('wants_binning_comparison'):
        return compare_binnings(company)
    
    if binning and parsed_query.get('wants_binning_status'):
        return get_binning_status(company, binning)
    
    # ===== COMPANY STATUS =====
    if parsed_query.get('wants_company_status'):
        return get_company_full_status(company)
    
    # ===== PDI vs MRP COMPARISON =====
    if pdi:
        print(f"[DEBUG answer_specific_query] PDI detected: {pdi}, RO: {ro}, type(ro): {type(ro)}")
        # Check if running order filter is also requested
        if ro:
            # Use the new filtered comparison function
            # ro can be a string "R-1" or a list ["R-1", "R-2"]
            print(f"[DEBUG] Calling compare_pdi_with_mrp_filtered with ro={ro}")
            return compare_pdi_with_mrp_filtered(company, pdi, ro)
        else:
            print(f"[DEBUG] Calling compare_pdi_with_mrp (no RO filter)")
            return compare_pdi_with_mrp(company, pdi)
    
    # ===== RUNNING ORDER STATUS (default if RO provided) =====
    if ro and not binning:
        return get_running_order_status(company, ro)
    
    # ===== BINNING STATUS (default if binning provided) =====
    if binning and not ro:
        return get_binning_status(company, binning)
    
    # ===== COMBINED FILTERS (RO + Binning) =====
    if ro and binning:
        # Normalize ro to list for comparison
        ro_list = ro if isinstance(ro, list) else [ro]
        ro_list = [r.upper() for r in ro_list]
        
        # Get specific data with both filters
        mrp_result = get_all_mrp_data(company)
        if not mrp_result.get('success'):
            return {'has_answer': False}
        
        filtered = []
        for b in mrp_result.get('data', []):
            bc_ro = extract_ro_from_ro(b.get('running_order', ''))
            bc_bin = extract_binning_from_ro(b.get('running_order', ''))
            
            if bc_ro in ro_list and bc_bin == binning.upper():
                filtered.append({
                    'barcode': b.get('barcode', ''),
                    'dispatched': bool(b.get('dispatch_party'))
                })
        
        dispatched = len([f for f in filtered if f['dispatched']])
        remaining = len(filtered) - dispatched
        
        answer_parts = [f"**🏭 {company} - {ro} | {binning}**\n"]
        answer_parts.append(f"━━━━━━━━━━━━━━━━━━━━━━━━━━━━\n")
        answer_parts.append(f"📊 **Total:** {len(filtered):,}")
        answer_parts.append(f"🚚 **Dispatched:** {dispatched:,}")
        answer_parts.append(f"⏳ **Remaining:** {remaining:,}")
        
        return {'has_answer': True, 'answer': "\n".join(answer_parts)}
    
    # ===== DEFAULT: Get general data with filters =====
    filters = {}
    if ro:
        filters['running_order'] = ro
    if binning:
        filters['binning'] = binning
    if pallet:
        filters['pallet_no'] = pallet
    
    mrp_data = get_specific_mrp_data(company, filters)
    
    if not mrp_data.get('success'):
        return {'has_answer': False}
    
    # Build answer
    answer_parts = []
    filter_desc = [f for f in [ro, binning, pdi, f"Pallet {pallet}" if pallet else None] if f]
    filter_str = " | ".join(filter_desc) if filter_desc else "All"
    answer_parts.append(f"**{company} - {filter_str}**\n")
    
    if parsed_query.get('wants_dispatch_count'):
        answer_parts.append(f"🚚 **Dispatched:** {mrp_data['dispatched_count']:,} modules")
    
    if parsed_query.get('wants_packed_count'):
        answer_parts.append(f"📦 **Packed (Not Dispatched):** {mrp_data['packed_count']:,} modules")
    
    if parsed_query.get('wants_remaining'):
        answer_parts.append(f"⏳ **Remaining (Bacha):** {mrp_data['remaining']:,} modules")
    
    if not parsed_query.get('wants_dispatch_count') and not parsed_query.get('wants_packed_count') and not parsed_query.get('wants_remaining'):
        answer_parts.append(f"📊 **Total:** {mrp_data['total']:,}")
        answer_parts.append(f"🚚 **Dispatched:** {mrp_data['dispatched_count']:,}")
        answer_parts.append(f"📦 **Packed:** {mrp_data['packed_count']:,}")
        answer_parts.append(f"⏳ **Pending:** {mrp_data['pending_count']:,}")
    
    if len(answer_parts) > 1:
        return {'has_answer': True, 'answer': "\n".join(answer_parts)}
    
    return {'has_answer': False}

def get_external_packed_dispatch_data(party_name):
    """Fetch packed and dispatch data from external API with caching"""
    import time
    global _external_cache
    
    cache_key = party_name.lower()
    current_time = time.time()
    
    # Check cache first
    if cache_key in _external_cache:
        cached_data, cached_time = _external_cache[cache_key]
        if current_time - cached_time < _cache_timeout:
            return cached_data
    
    try:
        # Get the MRP API party name
        mrp_party_name = get_mrp_party_name(party_name)
        
        response = requests.post(
            BARCODE_TRACKING_API,
            json={'party_name': mrp_party_name},
            timeout=15  # Reduced timeout
        )
        
        if response.status_code == 200:
            data = response.json()
            if data.get('status') == 'success':
                barcodes = data.get('data', [])
                
                # Count packed and dispatched
                packed_count = 0
                dispatched_count = 0
                packed_barcodes = []
                dispatched_barcodes = []
                
                # Binning breakdown from running_order (e.g., "R-3 i-2" -> binning = "I2")
                binning_counts = {}  # {'I1': 0, 'I2': 0, 'I3': 0, etc.}
                running_order_counts = {}  # {'R-1': 0, 'R-2': 0, 'R-3': 0, etc.}
                
                for item in barcodes:
                    # Parse running_order for binning and running order
                    running_order = item.get('running_order', '') or ''
                    if running_order:
                        # Extract running order (R-1, R-2, R-3 etc.)
                        import re
                        ro_match = re.search(r'(R-\d+)', running_order, re.IGNORECASE)
                        if ro_match:
                            ro = ro_match.group(1).upper()
                            running_order_counts[ro] = running_order_counts.get(ro, 0) + 1
                        
                        # Extract binning (i-1, i-2, i-3 etc.)
                        bin_match = re.search(r'i-?(\d+)', running_order, re.IGNORECASE)
                        if bin_match:
                            binning = f"I{bin_match.group(1)}"
                            binning_counts[binning] = binning_counts.get(binning, 0) + 1
                    
                    if item.get('status') == 'packed':
                        packed_count += 1
                        packed_barcodes.append({
                            'barcode': item.get('barcode'),
                            'running_order': running_order,
                            'pallet_no': item.get('pallet_no'),
                            'date': item.get('date')
                        })
                    if item.get('dispatch_party') or item.get('status') == 'dispatched':
                        dispatched_count += 1
                        dispatched_barcodes.append({
                            'barcode': item.get('barcode'),
                            'dispatch_party': item.get('dispatch_party'),
                            'running_order': running_order
                        })
                
                result = {
                    'success': True,
                    'party': party_name,
                    'total_count': data.get('count', len(barcodes)),
                    'packed_count': packed_count,
                    'dispatched_count': dispatched_count,
                    'pending_dispatch': packed_count - dispatched_count,
                    'binning_breakdown': binning_counts,  # NEW: Binning from MRP
                    'running_order_breakdown': running_order_counts,  # NEW: Running orders
                    'sample_packed': packed_barcodes[:10],
                    'sample_dispatched': dispatched_barcodes[:10]
                }
                # Cache the result
                _external_cache[cache_key] = (result, current_time)
                return result
        
        return {'success': False, 'error': 'API call failed'}
        
    except Exception as e:
        print(f"Error fetching external data: {str(e)}")
        return {'success': False, 'error': str(e)}

def get_all_ftr_data():
    """Get all FTR data from database for AI context"""
    try:
        data = {
            'companies': [],
            'summary': {},
            'pending_serials': {},  # Store pending serial numbers
            'external_tracking': {}  # Store external API data
        }
        
        # Get all companies
        result = db.session.execute(text("""
            SELECT id, company_name, module_wattage FROM companies
        """))
        companies = result.fetchall()
        
        total_master = 0
        total_assigned = 0
        total_packed = 0
        total_rejected = 0
        total_ext_packed = 0
        total_ext_dispatched = 0
        
        for company in companies:
            company_id = company[0]
            company_name = company[1]
            module_wattage = company[2]
            
            # Get master FTR count (total)
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM ftr_master_serials 
                WHERE company_id = :cid
            """), {'cid': company_id})
            master_total = result.fetchone()[0] or 0
            
            # Get rejected count
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM ftr_master_serials 
                WHERE company_id = :cid AND class_status = 'REJECTED'
            """), {'cid': company_id})
            rejected = result.fetchone()[0] or 0
            
            # Get available count (OK and available)
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM ftr_master_serials 
                WHERE company_id = :cid AND status = 'available' AND (class_status = 'OK' OR class_status IS NULL)
            """), {'cid': company_id})
            available = result.fetchone()[0] or 0
            
            # Get assigned count (OK only)
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM ftr_master_serials 
                WHERE company_id = :cid AND status = 'assigned' AND (class_status = 'OK' OR class_status IS NULL)
            """), {'cid': company_id})
            assigned = result.fetchone()[0] or 0
            
            # Get packed count
            result = db.session.execute(text("""
                SELECT COUNT(*) FROM ftr_packed_modules 
                WHERE company_id = :cid
            """), {'cid': company_id})
            packed = result.fetchone()[0] or 0
            
            # Get BINNING breakdown (only OK modules)
            result = db.session.execute(text("""
                SELECT binning, COUNT(*) as count
                FROM ftr_master_serials
                WHERE company_id = :cid AND (class_status = 'OK' OR class_status IS NULL)
                GROUP BY binning
                ORDER BY binning
            """), {'cid': company_id})
            binning_breakdown = [{'binning': row[0] or 'Unknown', 'count': row[1]} for row in result.fetchall()]
            
            # Get PDI-wise breakdown
            result = db.session.execute(text("""
                SELECT pdi_number, COUNT(*) as count
                FROM ftr_master_serials
                WHERE company_id = :cid AND status = 'assigned' AND pdi_number IS NOT NULL
                GROUP BY pdi_number
                ORDER BY count DESC
            """), {'cid': company_id})
            pdi_breakdown = [{'pdi': row[0], 'count': row[1]} for row in result.fetchall()]
            
            # Get PENDING serial numbers with binning (assigned but not packed, OK only)
            result = db.session.execute(text("""
                SELECT m.serial_number, m.pdi_number, m.binning
                FROM ftr_master_serials m
                LEFT JOIN ftr_packed_modules p ON m.serial_number = p.serial_number AND m.company_id = p.company_id
                WHERE m.company_id = :cid AND m.status = 'assigned' AND p.id IS NULL
                AND (m.class_status = 'OK' OR m.class_status IS NULL)
                ORDER BY m.binning, m.pdi_number, m.serial_number
                LIMIT 50
            """), {'cid': company_id})
            pending_serials = [{'serial': row[0], 'pdi': row[1], 'binning': row[2]} for row in result.fetchall()]
            
            # Store pending serials by company
            data['pending_serials'][company_name] = pending_serials
            
            company_data = {
                'id': company_id,
                'name': company_name,
                'wattage': module_wattage,
                'master_total': master_total,
                'rejected': rejected,
                'ok_total': master_total - rejected,
                'available': available,
                'assigned': assigned,
                'packed': packed,
                'pending_pack': assigned - packed if assigned > packed else 0,
                'binning_breakdown': binning_breakdown,
                'pdi_breakdown': pdi_breakdown,
                'sample_pending_serials': pending_serials[:10]  # First 10 for quick reference
            }
            
            # Fetch external tracking data for this company
            ext_data = get_external_packed_dispatch_data(company_name)
            if ext_data.get('success'):
                company_data['ext_packed'] = ext_data.get('packed_count', 0)
                company_data['ext_dispatched'] = ext_data.get('dispatched_count', 0)
                company_data['ext_pending_dispatch'] = ext_data.get('pending_dispatch', 0)
                data['external_tracking'][company_name] = ext_data
                total_ext_packed += ext_data.get('packed_count', 0)
                total_ext_dispatched += ext_data.get('dispatched_count', 0)
            
            data['companies'].append(company_data)
            total_master += master_total
            total_assigned += assigned
            total_packed += packed
            total_rejected += rejected
        
        data['summary'] = {
            'total_companies': len(companies),
            'total_master_ftr': total_master,
            'total_rejected': total_rejected,
            'total_ok': total_master - total_rejected,
            'total_assigned': total_assigned,
            'total_packed': total_packed,
            'total_available': total_master - total_assigned - total_rejected,
            'total_pending_pack': total_assigned - total_packed,
            # External API data
            'ext_total_packed': total_ext_packed,
            'ext_total_dispatched': total_ext_dispatched,
            'ext_pending_dispatch': total_ext_packed - total_ext_dispatched
        }
        
        return data
        
    except Exception as e:
        print(f"Error getting FTR data: {str(e)}")
        import traceback
        traceback.print_exc()
        return None

def create_system_prompt(ftr_data):
    """Create system prompt with FTR data context"""
    
    companies_info = ""
    pending_serials_info = ""
    external_tracking_info = ""
    
    for c in ftr_data['companies']:
        pdi_info = ", ".join([f"{p['pdi']}: {p['count']}" for p in c['pdi_breakdown'][:5]]) if c['pdi_breakdown'] else "None"
        
        # Binning breakdown
        binning_info = ", ".join([f"{b['binning']}: {b['count']}" for b in c.get('binning_breakdown', [])]) if c.get('binning_breakdown') else "No binning data"
        
        # External tracking data
        ext_packed = c.get('ext_packed', 0)
        ext_dispatched = c.get('ext_dispatched', 0)
        ext_pending = c.get('ext_pending_dispatch', 0)
        
        # Add sample pending serials for this company
        sample_serials = c.get('sample_pending_serials', [])
        if sample_serials:
            serial_list = ", ".join([f"{s['serial']}({s.get('binning', '?')})" for s in sample_serials[:5]])
            pending_info = f"Sample Pending Barcodes: {serial_list}{'...' if len(sample_serials) > 5 else ''}"
        else:
            pending_info = "No pending barcodes"
        
        companies_info += f"""
Company: {c['name']} ({c['wattage']})
- Master FTR Total: {c['master_total']:,}
- ❌ REJECTED (Do NOT Pack): {c.get('rejected', 0):,}
- ✅ OK Modules: {c.get('ok_total', c['master_total']):,}
- Available (Not Assigned): {c['available']:,}
- Assigned to PDI: {c['assigned']:,}
- 📦 ACTUALLY PACKED (Real-time from MRP): {ext_packed:,}
- 🚚 DISPATCHED (Real-time from MRP): {ext_dispatched:,}
- ⏳ PENDING DISPATCH (Packed but not sent): {ext_pending:,}
- 📊 BINNING Breakdown: {binning_info}
- PDI Breakdown: {pdi_info}
- {pending_info}
"""
    
    # Create detailed pending serials section with binning
    for company_name, serials in ftr_data.get('pending_serials', {}).items():
        if serials:
            pending_serials_info += f"\n{company_name} - Pending Barcodes to Pack:\n"
            # Group by Binning first, then PDI
            binning_groups = {}
            for s in serials:
                binning = s.get('binning') or 'Unknown'
                pdi = s['pdi'] or 'Unknown'
                key = f"{binning}"
                if key not in binning_groups:
                    binning_groups[key] = {'serials': [], 'pdi': pdi}
                binning_groups[key]['serials'].append(s['serial'])
            
            for binning, data in binning_groups.items():
                serial_list = data['serials']
                pending_serials_info += f"  [{binning}]: {', '.join(serial_list[:10])}{'...' if len(serial_list) > 10 else ''} ({len(serial_list)} total)\n"
    
    # External tracking summary with Binning and Running Order
    for company_name, ext_data in ftr_data.get('external_tracking', {}).items():
        if ext_data.get('success'):
            external_tracking_info += f"\n{company_name} (From MRP System):\n"
            external_tracking_info += f"  📦 Packed: {ext_data.get('packed_count', 0):,}\n"
            external_tracking_info += f"  🚚 Dispatched: {ext_data.get('dispatched_count', 0):,}\n"
            external_tracking_info += f"  ⏳ Pending Dispatch: {ext_data.get('pending_dispatch', 0):,}\n"
            
            # Binning breakdown from MRP
            binning_breakdown = ext_data.get('binning_breakdown', {})
            if binning_breakdown:
                binning_str = ", ".join([f"{k}: {v:,}" for k, v in sorted(binning_breakdown.items())])
                external_tracking_info += f"  📊 BINNING (from MRP): {binning_str}\n"
            
            # Running Order breakdown from MRP
            running_order_breakdown = ext_data.get('running_order_breakdown', {})
            if running_order_breakdown:
                ro_str = ", ".join([f"{k}: {v:,}" for k, v in sorted(running_order_breakdown.items())])
                external_tracking_info += f"  🏭 RUNNING ORDER: {ro_str}\n"
            
            # Sample packed barcodes with details
            sample_packed = ext_data.get('sample_packed', [])
            if sample_packed and isinstance(sample_packed[0], dict):
                external_tracking_info += f"  📋 Sample Packed Barcodes:\n"
                for item in sample_packed[:5]:
                    bc = item.get('barcode', 'N/A')
                    ro = item.get('running_order', 'N/A')
                    pallet = item.get('pallet_no', 'N/A')
                    external_tracking_info += f"    - {bc} | {ro} | Pallet: {pallet}\n"
    
    summary = ftr_data['summary']
    
    system_prompt = f"""You are an AI assistant for FTR (Final Test Report) Management System at a solar panel manufacturing company.

CURRENT DATA (Real-time from database + MRP System):

=== OVERALL SUMMARY ===
Total Companies: {summary['total_companies']}
Total Master FTR: {summary['total_master_ftr']:,}
❌ Total REJECTED: {summary.get('total_rejected', 0):,} (NEVER PACK THESE!)
✅ Total OK: {summary.get('total_ok', summary['total_master_ftr']):,}
Total Assigned to PDI: {summary['total_assigned']:,}

=== REAL-TIME PRODUCTION STATUS (From MRP System) ===
📦 Total Actually PACKED: {summary.get('ext_total_packed', 0):,}
🚚 Total DISPATCHED: {summary.get('ext_total_dispatched', 0):,}
⏳ PENDING DISPATCH (Packed but not sent): {summary.get('ext_pending_dispatch', 0):,}

=== COMPANY-WISE DATA ===
{companies_info}

=== EXTERNAL TRACKING (Real-time from MRP) ===
{external_tracking_info if external_tracking_info else "No external tracking data available."}

=== PENDING BARCODES TO PACK (Sample) ===
{pending_serials_info if pending_serials_info else "No pending barcodes found."}

INSTRUCTIONS:
1. Answer questions about FTR, production, packing, and dispatch status
2. Use the above real data to answer accurately - PACKED and DISPATCHED data is REAL-TIME from MRP system
3. If asked about a specific company, provide detailed info including packed/dispatch status
4. Be helpful and concise
5. Use numbers with commas for readability
6. If data is 0 or not available, mention that clearly
7. Respond in Hindi-English mix (Hinglish) if user asks in Hindi
8. Always be accurate - don't make up numbers
9. When asked "konse barcode pack karne hai" or "which barcodes to pack", show the PENDING BARCODES list with BINNING
10. If user asks for specific barcode numbers, provide from the pending list above
11. ⚠️ CRITICAL: REJECTED modules should NEVER be packed! Always warn if someone asks about rejected barcodes
12. When giving pending barcodes, always mention their BINNING (I1, I2, I3) so correct binning can be packed
13. If asked "kitna aur banana hai", show pending count with BINNING breakdown
14. When asked about dispatch status, use the REAL-TIME data from MRP system
15. "Pending Dispatch" means modules are packed but not yet sent to customer
16. When asked about "binning", show the BINNING breakdown from MRP system (I1, I2, I3 counts)
17. When asked about "running order" or "R-1, R-2, R-3", show the RUNNING ORDER breakdown
18. If user asks "R-3 me kitna hai" or "I2 ka data", provide specific binning/running order counts

TERMINOLOGY:
- Master FTR = Total serial numbers uploaded for a company
- Assigned = Serial numbers given to specific PDI orders
- Packed = Actually packed modules (from MRP system - real production data)
- Dispatched = Modules sent to customer (from MRP system)
- Pending Dispatch = Packed but not yet dispatched
- Available = Master FTR minus Assigned (can be used for new PDIs)
- Pending Pack = Assigned but not yet packed (THESE BARCODES NEED TO BE PACKED)
- REJECTED = Failed modules that should NEVER be packed ❌
- Binning = Current classification (I1, I2, I3) - Pack according to binning only!
- I1, I2, I3 = Different current bins - modules must be packed matching their bin
- Running Order = R-1, R-2, R-3 etc. - Production batch/running number
- running_order field format: "R-3 i-2" means Running Order R-3, Binning I2
"""
    
    return system_prompt

def query_groq(user_message, ftr_data):
    """Query Groq API with FTR context"""
    
    if not GROQ_API_KEY:
        return {
            'success': False,
            'error': 'Groq API key not configured. Please set GROQ_API_KEY environment variable.'
        }
    
    system_prompt = create_system_prompt(ftr_data)
    
    try:
        response = requests.post(
            GROQ_API_URL,
            headers={
                'Authorization': f'Bearer {GROQ_API_KEY}',
                'Content-Type': 'application/json'
            },
            json={
                'model': 'llama-3.1-8b-instant',
                'messages': [
                    {'role': 'system', 'content': system_prompt},
                    {'role': 'user', 'content': user_message}
                ],
                'temperature': 0.3,
                'max_tokens': 1000
            },
            timeout=30
        )
        
        if response.status_code == 200:
            result = response.json()
            ai_response = result['choices'][0]['message']['content']
            return {
                'success': True,
                'response': ai_response
            }
        else:
            return {
                'success': False,
                'error': f'Groq API error: {response.status_code} - {response.text}'
            }
            
    except requests.exceptions.Timeout:
        return {
            'success': False,
            'error': 'Request timeout. Please try again.'
        }
    except Exception as e:
        return {
            'success': False,
            'error': str(e)
        }

def detect_excel_command(message):
    """
    Detect if user wants Excel export from chat message
    Returns: {'type': 'excel_command', 'action': '...', 'params': {...}} or None
    """
    import re
    message_lower = message.lower()
    
    # Patterns for Excel commands
    excel_keywords = ['excel', 'export', 'download', 'list', 'de do', 'do', 'chahiye', 'nikalo', 'bhejo', 'file', 'sheet']
    has_excel_intent = any(kw in message_lower for kw in excel_keywords)
    
    # Quality check keywords
    quality_check_types = {
        'duplicate': ['duplicate', 'dohra', 'repeat', 'same barcode'],
        'mismatch': ['mismatch', 'binning mismatch', 'different binning', 'galat binning'],
        'rejected': ['rejected', 'reject', 'rejected packed', 'rejection'],
        'missing': ['missing', 'mrp mein nahi', 'not in mrp', 'gayab'],
        'extra': ['extra', 'db mein nahi', 'not in db', 'extra mrp'],
        'mix_packing': ['mix', 'mix packing', 'mixed', 'mix pallet'],
        'pallet_audit': ['audit', 'pallet audit', 'full audit', 'integrity']
    }
    
    detected_quality_check = None
    for check_type, keywords in quality_check_types.items():
        if any(kw in message_lower for kw in keywords):
            detected_quality_check = check_type
            break
    
    # Company detection
    company_patterns = {
        'Rays Power': r'rays|rp|rays power',
        'Larsen & Toubro': r'l&t|larsen|toubro|lnt',
        'Sterlin and Wilson': r'sterlin|wilson|sw|s&w|s\&w|sterling'
    }
    detected_company = None
    for company, pattern in company_patterns.items():
        if re.search(pattern, message_lower):
            detected_company = company
            break
    
    # Running Order detection (R-1, R-2, R-3, etc.)
    ro_match = re.search(r'r-?(\d+)', message_lower)
    running_order = f"R-{ro_match.group(1)}" if ro_match else None
    
    # Binning detection (I-1, I-2, I-3, i1, i2, i3, MB, MC, MD)
    # Use negative lookbehind to NOT match PDI-1 as binning I1
    bin_match = re.search(r'(?<!pd)i-?(\d+)', message_lower)
    binning = f"I{bin_match.group(1)}" if bin_match else None
    if not binning:
        # Note: 'me' removed as it conflicts with Hindi word 'me/mein' (meaning 'in')
        bin_alpha = re.search(r'\b(mb|mc|md|mf|mg)\b', message_lower, re.IGNORECASE)
        if bin_alpha:
            binning = bin_alpha.group(1).upper()
    
    # Julian date detection (3 digit number like 302, 365)
    julian_match = re.search(r'\bjulian\s*(\d{3})\b|\b(\d{3})\s*julian\b', message_lower)
    julian_date = julian_match.group(1) or julian_match.group(2) if julian_match else None
    
    # PDI detection
    pdi_match = re.search(r'pdi-?(\d+)', message_lower)
    pdi_number = f"PDI-{pdi_match.group(1)}" if pdi_match else None
    
    # Count detection (for "18 barcode chahiye")
    count_match = re.search(r'(\d+)\s*(barcode|module|serial|piece)', message_lower)
    if not count_match:
        count_match = re.search(r'(barcode|module|serial)\s*(\d+)', message_lower)
    requested_count = int(count_match.group(1) if count_match and count_match.group(1).isdigit() else (count_match.group(2) if count_match else 0))
    
    # Status detection
    packed = 'packed' in message_lower or 'pack' in message_lower
    dispatched = 'dispatch' in message_lower or 'sent' in message_lower or 'bheja' in message_lower
    pending = 'pending' in message_lower or 'baaki' in message_lower or 'remaining' in message_lower
    
    # IMPORTANT: If PDI number is detected, DON'T treat as Excel command
    # PDI queries should go to answer_specific_query for proper PDI comparison
    if pdi_number:
        return None
    
    # If any specific filter detected with excel intent
    if has_excel_intent or running_order or binning or detected_quality_check:
        return {
            'type': 'excel_command',
            'company': detected_company,
            'running_order': running_order,
            'binning': binning,
            'julian_date': julian_date,
            'pdi_number': pdi_number,
            'count': requested_count,
            'packed': packed,
            'dispatched': dispatched,
            'pending': pending,
            'quality_check': detected_quality_check
        }
    
    return None

def generate_smart_excel(params):
    """Generate Excel based on smart command parameters - supports quality checks too"""
    import re
    
    company = params.get('company') or 'Rays Power'
    running_order_filter = params.get('running_order')
    binning_filter = params.get('binning')
    julian_filter = params.get('julian_date')
    pdi_number = params.get('pdi_number')
    count = params.get('count', 0)
    get_packed = params.get('packed', False)
    get_dispatched = params.get('dispatched', False)
    get_pending = params.get('pending', False)
    quality_check = params.get('quality_check')
    
    if not EXCEL_AVAILABLE:
        return None, "Excel not available"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    red_fill = PatternFill(start_color="FF6B6B", end_color="FF6B6B", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFE066", end_color="FFE066", fill_type="solid")
    green_fill = PatternFill(start_color="69DB7C", end_color="69DB7C", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    # ===== QUALITY CHECK EXCEL =====
    if quality_check:
        result = None
        
        if quality_check == 'duplicate':
            result = check_duplicate_barcodes(company)
            ws.title = f"{company[:20]}_Duplicates"
            headers = ["S.No", "Barcode", "Count", "R-O Values", "Issue"]
            
        elif quality_check == 'mismatch':
            result = check_binning_mismatch(company)
            ws.title = f"{company[:20]}_Mismatch"
            headers = ["S.No", "Barcode", "DB Binning", "MRP Binning", "Issue"]
            
        elif quality_check == 'rejected':
            result = check_rejected_packed(company)
            ws.title = f"{company[:20]}_Rejected"
            headers = ["S.No", "Barcode", "Rejection Reason", "Status in MRP", "Issue"]
            
        elif quality_check == 'missing':
            result = check_missing_in_mrp(company)
            ws.title = f"{company[:20]}_Missing"
            headers = ["S.No", "Barcode", "DB Status", "Issue"]
            
        elif quality_check == 'extra':
            result = check_extra_in_mrp(company)
            ws.title = f"{company[:20]}_Extra"
            headers = ["S.No", "Barcode", "Pallet", "Running Order"]
            
        elif quality_check == 'mix_packing':
            result = check_mix_packing(company)
            ws.title = f"{company[:20]}_MixPacking"
            headers = ["S.No", "Pallet No", "Barcode", "Actual Binning (Master FTR)", "Issue"]
            
        elif quality_check == 'pallet_audit':
            result = full_pallet_audit(company)
            ws.title = f"{company[:20]}_Audit"
            headers = ["S.No", "Issue Type", "Count", "Details"]
        else:
            return None, "Unknown quality check type"
        
        # Safety check - if result is None or has error
        if not result or not result.get('has_answer'):
            return None, f"Quality check failed for {company}"
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        # Extract data from result based on quality check type
        row_idx = 2
        
        if quality_check == 'duplicate':
            duplicates = result.get('duplicates', {})
            for idx, (bc, locations) in enumerate(duplicates.items(), 1):
                pallets = ", ".join([f"Pallet {l['pallet']} ({l['status']})" for l in locations])
                ws.cell(row=row_idx, column=1, value=idx).border = thin_border
                ws.cell(row=row_idx, column=2, value=bc).border = thin_border
                ws.cell(row=row_idx, column=3, value=len(locations)).border = thin_border
                ws.cell(row=row_idx, column=4, value=pallets).border = thin_border
                ws.cell(row=row_idx, column=5, value="DUPLICATE").border = thin_border
                ws.cell(row=row_idx, column=5).fill = red_fill
                row_idx += 1
                
        elif quality_check == 'mismatch':
            mismatches = result.get('mismatches', [])
            for idx, m in enumerate(mismatches, 1):
                ws.cell(row=row_idx, column=1, value=idx).border = thin_border
                ws.cell(row=row_idx, column=2, value=m.get('barcode', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=m.get('db_binning', '')).border = thin_border
                ws.cell(row=row_idx, column=4, value=m.get('mrp_binning', '')).border = thin_border
                ws.cell(row=row_idx, column=5, value="MISMATCH").border = thin_border
                ws.cell(row=row_idx, column=5).fill = yellow_fill
                row_idx += 1
                
        elif quality_check == 'rejected':
            rejected = result.get('rejected_packed', [])
            for idx, r in enumerate(rejected, 1):
                ws.cell(row=row_idx, column=1, value=idx).border = thin_border
                ws.cell(row=row_idx, column=2, value=r.get('barcode', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value="REJECTED").border = thin_border
                ws.cell(row=row_idx, column=4, value=f"Pallet {r.get('pallet', '')} - {r.get('status', '')}").border = thin_border
                ws.cell(row=row_idx, column=5, value="CRITICAL").border = thin_border
                ws.cell(row=row_idx, column=5).fill = red_fill
                row_idx += 1
                
        elif quality_check == 'missing':
            missing = result.get('missing', [])
            for idx, m in enumerate(missing, 1):
                ws.cell(row=row_idx, column=1, value=idx).border = thin_border
                ws.cell(row=row_idx, column=2, value=m.get('barcode', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=f"PDI: {m.get('pdi', '')}").border = thin_border
                ws.cell(row=row_idx, column=4, value="NOT IN MRP").border = thin_border
                ws.cell(row=row_idx, column=4).fill = yellow_fill
                row_idx += 1
                
        elif quality_check == 'extra':
            extra = result.get('extra', [])
            for idx, e in enumerate(extra, 1):
                ws.cell(row=row_idx, column=1, value=idx).border = thin_border
                ws.cell(row=row_idx, column=2, value=e.get('barcode', '')).border = thin_border
                ws.cell(row=row_idx, column=3, value=e.get('pallet', '')).border = thin_border
                ws.cell(row=row_idx, column=4, value=e.get('running_order', '')).border = thin_border
                row_idx += 1
                
        elif quality_check == 'mix_packing':
            mixed_pallets = result.get('mix_packed_pallets', [])
            barcode_row_num = 1
            
            for pallet_idx, p in enumerate(mixed_pallets, 1):
                pallet_no = p.get('pallet_no', '')
                binnings_by_type = p.get('binnings_by_type', {})
                
                # Write each barcode as a separate row
                for binning, barcodes in sorted(binnings_by_type.items()):
                    for barcode in barcodes:
                        ws.cell(row=row_idx, column=1, value=barcode_row_num).border = thin_border
                        ws.cell(row=row_idx, column=2, value=f"Pallet {pallet_no}").border = thin_border
                        ws.cell(row=row_idx, column=3, value=barcode).border = thin_border
                        ws.cell(row=row_idx, column=4, value=binning).border = thin_border
                        ws.cell(row=row_idx, column=5, value="MIX PACKING").border = thin_border
                        ws.cell(row=row_idx, column=5).fill = yellow_fill
                        row_idx += 1
                        barcode_row_num += 1
        
        # If no data found
        if row_idx == 2:
            ws.cell(row=2, column=1, value=1).border = thin_border
            ws.cell(row=2, column=2, value="No issues found").border = thin_border
            ws.cell(row=2, column=2).fill = green_fill
        
        # Summary row
        summary_row = row_idx + 2
        ws.cell(row=summary_row, column=1, value="QUALITY CHECK SUMMARY:").font = Font(bold=True)
        ws.cell(row=summary_row + 1, column=1, value=f"Company: {company}")
        ws.cell(row=summary_row + 2, column=1, value=f"Check Type: {quality_check.replace('_', ' ').title()}")
        ws.cell(row=summary_row + 3, column=1, value=f"Total Issues: {row_idx - 2}")
        ws.cell(row=summary_row + 4, column=1, value=f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M')}")
        
        # Auto-width
        for col in ws.columns:
            max_length = max(len(str(cell.value or '')) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_length + 2, 50)
        
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        import base64
        excel_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        
        return excel_base64, f"{quality_check.replace('_', ' ').title()} check complete for {company}"
    
    # ===== STANDARD EXCEL (Barcode List) =====
    # Default to packed if no status specified
    if not get_packed and not get_dispatched and not get_pending:
        get_packed = True
    
    # Get MRP data
    mrp_party_name = get_mrp_party_name(company)
    try:
        response = requests.post(BARCODE_TRACKING_API, json={'party_name': mrp_party_name}, timeout=60)
        mrp_data = response.json() if response.status_code == 200 else {'data': []}
    except:
        mrp_data = {'data': []}
    
    barcodes = mrp_data.get('data', [])
    
    # Filter by status
    filtered = []
    for b in barcodes:
        status = b.get('status', '')
        dispatch_party = b.get('dispatch_party')
        
        is_packed = status == 'packed'
        is_dispatched = bool(dispatch_party) or status == 'dispatched'
        
        if get_dispatched and is_dispatched:
            filtered.append(b)
        elif get_packed and is_packed and not is_dispatched:
            filtered.append(b)
        elif get_pending and not is_packed and not is_dispatched:
            filtered.append(b)
    
    # Filter by Running Order
    if running_order_filter:
        filtered = [b for b in filtered if running_order_filter.upper() in (b.get('running_order', '') or '').upper()]
    
    # Filter by Binning
    if binning_filter:
        result = []
        for b in filtered:
            ro = b.get('running_order', '') or ''
            bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
            if bin_match:
                barcode_binning = f"I{bin_match.group(1)}"
                if barcode_binning.upper() == binning_filter.upper():
                    result.append(b)
        filtered = result
    
    # Limit by count if specified
    if count > 0:
        filtered = filtered[:count]
    
    # Generate Excel
    if not EXCEL_AVAILABLE:
        return None, "Excel not available"
    
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Determine title
    title_parts = [company]
    if running_order_filter:
        title_parts.append(running_order_filter)
    if binning_filter:
        title_parts.append(binning_filter)
    if pdi_number:
        title_parts.append(pdi_number)
    if get_packed:
        title_parts.append("Packed")
    if get_dispatched:
        title_parts.append("Dispatched")
    
    ws.title = "_".join(title_parts)[:31]  # Excel sheet name limit
    
    # Styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    
    headers = ["S.No", "Barcode", "Running Order", "Binning", "Pallet No", "Date", "Status"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.border = thin_border
    
    for idx, b in enumerate(filtered, 1):
        ro = b.get('running_order', '') or ''
        bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
        binning = f"I{bin_match.group(1)}" if bin_match else ''
        
        status = 'Dispatched' if b.get('dispatch_party') else 'Packed'
        
        ws.cell(row=idx+1, column=1, value=idx).border = thin_border
        ws.cell(row=idx+1, column=2, value=b.get('barcode', '')).border = thin_border
        ws.cell(row=idx+1, column=3, value=ro).border = thin_border
        ws.cell(row=idx+1, column=4, value=binning).border = thin_border
        ws.cell(row=idx+1, column=5, value=b.get('pallet_no', '')).border = thin_border
        ws.cell(row=idx+1, column=6, value=b.get('date', '')).border = thin_border
        ws.cell(row=idx+1, column=7, value=status).border = thin_border
    
    # Summary
    summary_row = len(filtered) + 3
    ws.cell(row=summary_row, column=1, value="SUMMARY:").font = Font(bold=True)
    ws.cell(row=summary_row + 1, column=1, value=f"Company: {company}")
    ws.cell(row=summary_row + 2, column=1, value=f"Total Records: {len(filtered)}")
    if running_order_filter:
        ws.cell(row=summary_row + 3, column=1, value=f"Running Order: {running_order_filter}")
    if binning_filter:
        ws.cell(row=summary_row + 4, column=1, value=f"Binning: {binning_filter}")
    
    # Auto-width
    for col in ws.columns:
        max_length = max(len(str(cell.value or '')) for cell in col)
        ws.column_dimensions[col[0].column_letter].width = max_length + 2
    
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    
    import base64
    excel_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
    
    return excel_base64, f"Found {len(filtered)} barcodes"

@ai_assistant_bp.route('/ai/chat', methods=['POST'])
def ai_chat():
    """Main AI chat endpoint with smart Excel commands and specific queries"""
    try:
        data = request.json
        user_message = data.get('message', '').strip()
        
        if not user_message:
            return jsonify({'success': False, 'error': 'Message is required'}), 400
        
        print(f"\n{'='*60}")
        print(f"📥 NEW QUERY: {user_message}")
        print(f"{'='*60}")
        
        # STEP 1: Parse user query to understand intent
        parsed = parse_user_query(user_message)
        print(f"🔍 PARSED RESULT: {parsed}")
        
        # STEP 2: Check if this is an Excel command
        excel_cmd = detect_excel_command(user_message)
        print(f"📊 EXCEL COMMAND: {excel_cmd}")
        
        if excel_cmd:
            # Generate smart Excel
            excel_base64, summary = generate_smart_excel(excel_cmd)
            
            if excel_base64:
                # Build response message
                msg_parts = [f"✅ Excel generated!"]
                msg_parts.append(f"\n📊 {summary}")
                if excel_cmd.get('company'):
                    msg_parts.append(f"\n🏭 Company: {excel_cmd['company']}")
                if excel_cmd.get('running_order'):
                    msg_parts.append(f"\n🔢 Running Order: {excel_cmd['running_order']}")
                if excel_cmd.get('binning'):
                    msg_parts.append(f"\n🏷️ Binning: {excel_cmd['binning']}")
                if excel_cmd.get('count'):
                    msg_parts.append(f"\n📝 Requested: {excel_cmd['count']} barcodes")
                
                msg_parts.append("\n\n👇 Click button below to download Excel")
                
                return jsonify({
                    'success': True,
                    'response': "".join(msg_parts),
                    'has_excel': True,
                    'excel_base64': excel_base64,
                    'excel_params': excel_cmd
                })
            else:
                return jsonify({
                    'success': True,
                    'response': f"❌ Excel generation failed: {summary}"
                })
        
        # STEP 3: Try to answer specifically without AI
        specific_answer = answer_specific_query(parsed)
        
        if specific_answer and specific_answer.get('has_answer'):
            # Direct answer without AI
            return jsonify({
                'success': True,
                'response': specific_answer.get('answer', 'No answer generated'),
                'source': 'direct_mrp',
                'parsed_query': parsed
            })
        
        # STEP 5: Fall back to Groq AI for general questions
        # Get fresh FTR data
        ftr_data = get_all_ftr_data()
        
        if not ftr_data:
            return jsonify({
                'success': False,
                'error': 'Failed to fetch FTR data from database'
            }), 500
        
        # Query Groq AI
        result = query_groq(user_message, ftr_data)
        
        if result['success']:
            return jsonify({
                'success': True,
                'response': result['response'],
                'data_summary': ftr_data['summary']
            })
        else:
            return jsonify({
                'success': False,
                'error': result['error']
            }), 500
            
    except Exception as e:
        print(f"AI Chat Error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/data', methods=['GET'])
def get_ai_data():
    """Get raw FTR data for display"""
    try:
        ftr_data = get_all_ftr_data()
        if ftr_data:
            return jsonify({'success': True, 'data': ftr_data})
        else:
            return jsonify({'success': False, 'error': 'Failed to fetch data'}), 500
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/mix-packing-check', methods=['POST'])
def api_check_mix_packing():
    """
    API endpoint to check for mix packing in pallets
    POST: {'company': 'Rays Power'} or {'company': 'L&T'}
    """
    try:
        data = request.json
        company = data.get('company', '').strip()
        
        if not company:
            return jsonify({'success': False, 'error': 'Company name is required'}), 400
        
        # Map short names to full names
        company_map = {
            'rays': 'Rays Power',
            'l&t': 'Larsen & Toubro',
            'lnt': 'Larsen & Toubro',
            'sterling': 'Sterlin and Wilson',
            'wilson': 'Sterlin and Wilson'
        }
        
        company_lower = company.lower()
        for key, value in company_map.items():
            if key in company_lower:
                company = value
                break
        
        result = check_mix_packing(company)
        
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify(result), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/config', methods=['POST'])
def set_api_key():
    """Set Groq API key (for runtime configuration)"""
    global GROQ_API_KEY
    try:
        data = request.json
        api_key = data.get('api_key', '').strip()
        
        if not api_key:
            return jsonify({'success': False, 'error': 'API key is required'}), 400
        
        GROQ_API_KEY = api_key
        
        return jsonify({
            'success': True,
            'message': 'API key configured successfully'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/config', methods=['GET'])
def get_config_status():
    """Check if API key is configured"""
    return jsonify({
        'success': True,
        'configured': bool(GROQ_API_KEY),
        'key_preview': f"{GROQ_API_KEY[:8]}..." if GROQ_API_KEY else None
    })

@ai_assistant_bp.route('/ai/pending-barcodes', methods=['GET'])
def get_pending_barcodes():
    """Get list of pending barcodes (assigned but not packed)"""
    try:
        company_id = request.args.get('company_id')
        pdi_number = request.args.get('pdi_number')
        limit = request.args.get('limit', 100, type=int)
        
        query = """
            SELECT m.serial_number, m.pdi_number, c.company_name
            FROM ftr_master_serials m
            JOIN companies c ON m.company_id = c.id
            LEFT JOIN ftr_packed_modules p ON m.serial_number = p.serial_number AND m.company_id = p.company_id
            WHERE m.status = 'assigned' AND p.id IS NULL
        """
        params = {}
        
        if company_id:
            query += " AND m.company_id = :company_id"
            params['company_id'] = company_id
        
        if pdi_number:
            query += " AND m.pdi_number = :pdi_number"
            params['pdi_number'] = pdi_number
        
        query += " ORDER BY m.company_id, m.pdi_number, m.serial_number LIMIT :limit"
        params['limit'] = limit
        
        result = db.session.execute(text(query), params)
        pending = [{'serial': row[0], 'pdi': row[1], 'company': row[2]} for row in result.fetchall()]
        
        # Get total count
        count_query = """
            SELECT COUNT(*)
            FROM ftr_master_serials m
            LEFT JOIN ftr_packed_modules p ON m.serial_number = p.serial_number AND m.company_id = p.company_id
            WHERE m.status = 'assigned' AND p.id IS NULL
        """
        if company_id:
            count_query += f" AND m.company_id = {company_id}"
        if pdi_number:
            count_query += f" AND m.pdi_number = '{pdi_number}'"
        
        count_result = db.session.execute(text(count_query))
        total_count = count_result.fetchone()[0]
        
        return jsonify({
            'success': True,
            'pending_barcodes': pending,
            'total_pending': total_count,
            'showing': len(pending)
        })
        
    except Exception as e:
        print(f"Error getting pending barcodes: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/tracking', methods=['POST'])
def get_tracking_data():
    """Get real-time packed/dispatch tracking from external MRP API"""
    try:
        data = request.json
        party_name = data.get('party_name', '').strip()
        
        if not party_name:
            return jsonify({'success': False, 'error': 'Party name is required'}), 400
        
        result = get_external_packed_dispatch_data(party_name)
        
        if result.get('success'):
            return jsonify(result)
        else:
            return jsonify(result), 500
            
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/pallet/<pallet_no>', methods=['GET'])
def get_pallet_details(pallet_no):
    """Get all barcodes in a specific pallet"""
    import re
    try:
        company = request.args.get('company', 'Larsen & Toubro')
        
        mrp_party_name = get_mrp_party_name(company)
        response = requests.post(BARCODE_TRACKING_API, json={'party_name': mrp_party_name}, timeout=30)
        
        if response.status_code != 200:
            return jsonify({'success': False, 'error': 'MRP API failed'}), 500
        
        data = response.json()
        all_barcodes = data.get('data', [])
        
        # Filter by pallet number
        pallet_barcodes = [b for b in all_barcodes if str(b.get('pallet_no', '')) == str(pallet_no)]
        
        # Process each barcode
        result_list = []
        for b in pallet_barcodes:
            ro = b.get('running_order', '') or ''
            bin_match = re.search(r'i-?(\d+)', ro, re.IGNORECASE)
            binning = f"I{bin_match.group(1)}" if bin_match else ''
            
            result_list.append({
                'barcode': b.get('barcode'),
                'running_order': ro,
                'binning': binning,
                'status': 'Dispatched' if b.get('dispatch_party') else 'Packed',
                'date': b.get('date')
            })
        
        return jsonify({
            'success': True,
            'pallet_no': pallet_no,
            'company': company,
            'total_barcodes': len(result_list),
            'barcodes': result_list
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/smart-query', methods=['POST'])
def smart_query():
    """
    Smart query endpoint - understands natural language queries
    Examples:
    - "L&T ke PDI-1 me kitne dispatch hue"
    - "Rays ke R-3 me kitna bacha"
    - "Pallet 123 me kon kon se barcode hai"
    - "Sterling ka I-2 binning status"
    """
    import re
    try:
        data = request.json
        query = data.get('query', '').strip()
        
        if not query:
            return jsonify({'success': False, 'error': 'Query is required'}), 400
        
        # Parse the query
        parsed = parse_user_query(query)
        
        # Try to answer specifically
        answer = answer_specific_query(parsed)
        
        if answer.get('has_answer'):
            return jsonify({
                'success': True,
                'answer': answer['answer'],
                'parsed_query': parsed,
                'source': 'direct_mrp'
            })
        
        # If no specific answer, return parsed info
        return jsonify({
            'success': True,
            'answer': "Aapka query samajh nahi aaya. Please specify:\n- Company name (L&T, Rays, Sterling)\n- Filter (R-1, R-2, I-1, I-2, PDI number, Pallet number)\n- What you want (dispatch count, packed count, remaining, barcode list)",
            'parsed_query': parsed,
            'source': 'no_match'
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/tracking/all', methods=['GET'])
def get_all_tracking_data():
    """Get tracking data for all companies"""
    try:
        # Get all company names
        result = db.session.execute(text("SELECT company_name FROM companies"))
        companies = [row[0] for row in result.fetchall()]
        
        tracking_data = {}
        total_packed = 0
        total_dispatched = 0
        
        for company_name in companies:
            ext_data = get_external_packed_dispatch_data(company_name)
            if ext_data.get('success'):
                tracking_data[company_name] = {
                    'packed': ext_data.get('packed_count', 0),
                    'dispatched': ext_data.get('dispatched_count', 0),
                    'pending_dispatch': ext_data.get('pending_dispatch', 0)
                }
                total_packed += ext_data.get('packed_count', 0)
                total_dispatched += ext_data.get('dispatched_count', 0)
        
        return jsonify({
            'success': True,
            'companies': tracking_data,
            'summary': {
                'total_packed': total_packed,
                'total_dispatched': total_dispatched,
                'total_pending_dispatch': total_packed - total_dispatched
            }
        })
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@ai_assistant_bp.route('/ai/export/excel', methods=['POST'])
def export_to_excel():
    """Export FTR data to Excel based on user request"""
    try:
        if not EXCEL_AVAILABLE:
            return jsonify({'success': False, 'error': 'openpyxl not installed on server'}), 500
        
        data = request.json
        export_type = data.get('type', 'all')  # all, company, pending, packed, dispatched, binning, rejected
        company_id = data.get('company_id')
        company_name = data.get('company_name', 'All')
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "FTR Data"
        
        # Styles
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        if export_type == 'pending':
            # Export pending barcodes (assigned but not packed)
            ws.title = "Pending Barcodes"
            headers = ["S.No", "Serial Number", "PDI Number", "Company", "Status", "Binning", "Class"]
            
            query = """
                SELECT m.serial_number, m.pdi_number, c.company_name, m.status, m.binning, m.class_status
                FROM ftr_master_serials m
                JOIN companies c ON m.company_id = c.id
                LEFT JOIN ftr_packed_modules p ON m.serial_number = p.serial_number AND m.company_id = p.company_id
                WHERE m.status = 'assigned' AND p.id IS NULL
            """
            if company_id:
                query += f" AND m.company_id = {company_id}"
            query += " ORDER BY c.company_name, m.pdi_number, m.serial_number"
            
        elif export_type == 'packed':
            # Export packed modules from MRP API
            ws.title = "Packed Modules"
            headers = ["S.No", "Barcode", "Running Order", "Binning", "Pallet No", "Pack Date", "Status"]
            
            # Get from external API
            mrp_party_name = get_mrp_party_name(company_name)
            response = requests.post(BARCODE_TRACKING_API, json={'party_name': mrp_party_name}, timeout=30)
            if response.status_code == 200:
                api_data = response.json()
                barcodes = [b for b in api_data.get('data', []) if b.get('status') == 'packed']
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # Write data
                import re
                for idx, barcode in enumerate(barcodes, 1):
                    running_order = barcode.get('running_order', '') or ''
                    # Extract binning from running_order (e.g., "R-3 i-2" -> "I2")
                    binning = ''
                    bin_match = re.search(r'i-?(\d+)', running_order, re.IGNORECASE)
                    if bin_match:
                        binning = f"I{bin_match.group(1)}"
                    
                    ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                    ws.cell(row=idx+1, column=2, value=barcode.get('barcode', '')).border = thin_border
                    ws.cell(row=idx+1, column=3, value=running_order).border = thin_border
                    ws.cell(row=idx+1, column=4, value=binning).border = thin_border
                    ws.cell(row=idx+1, column=5, value=barcode.get('pallet_no', '')).border = thin_border
                    ws.cell(row=idx+1, column=6, value=barcode.get('date', '')).border = thin_border
                    ws.cell(row=idx+1, column=7, value='Packed').border = thin_border
                
                # Auto-width columns
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2
                
                # Save to buffer
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                filename = f"Packed_Modules_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               as_attachment=True, download_name=filename)
            else:
                return jsonify({'success': False, 'error': 'MRP API error'}), 500
                
        elif export_type == 'dispatched':
            # Export dispatched modules from MRP API
            ws.title = "Dispatched Modules"
            headers = ["S.No", "Barcode", "Running Order", "Binning", "Dispatch Party", "Dispatch Date", "Status"]
            
            mrp_party_name = get_mrp_party_name(company_name)
            response = requests.post(BARCODE_TRACKING_API, json={'party_name': mrp_party_name}, timeout=30)
            if response.status_code == 200:
                api_data = response.json()
                barcodes = [b for b in api_data.get('data', []) if b.get('dispatch_party') or b.get('status') == 'dispatched']
                
                # Write headers
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=1, column=col, value=header)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.border = thin_border
                
                # Write data
                import re
                for idx, barcode in enumerate(barcodes, 1):
                    running_order = barcode.get('running_order', '') or ''
                    # Extract binning from running_order
                    binning = ''
                    bin_match = re.search(r'i-?(\d+)', running_order, re.IGNORECASE)
                    if bin_match:
                        binning = f"I{bin_match.group(1)}"
                    
                    ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                    ws.cell(row=idx+1, column=2, value=barcode.get('barcode', '')).border = thin_border
                    ws.cell(row=idx+1, column=3, value=running_order).border = thin_border
                    ws.cell(row=idx+1, column=4, value=binning).border = thin_border
                    ws.cell(row=idx+1, column=5, value=barcode.get('dispatch_party', '')).border = thin_border
                    ws.cell(row=idx+1, column=6, value=barcode.get('dispatch_date', barcode.get('date', ''))).border = thin_border
                    ws.cell(row=idx+1, column=7, value='Dispatched').border = thin_border
                
                # Auto-width columns
                for col in ws.columns:
                    max_length = max(len(str(cell.value or '')) for cell in col)
                    ws.column_dimensions[col[0].column_letter].width = max_length + 2
                
                buffer = io.BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                
                filename = f"Dispatched_Modules_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
                return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                               as_attachment=True, download_name=filename)
            else:
                return jsonify({'success': False, 'error': 'MRP API error'}), 500
                
        elif export_type == 'binning':
            # Export binning data
            ws.title = "Binning Data"
            headers = ["S.No", "Serial Number", "PDI Number", "Company", "Pmax", "Binning", "Class"]
            
            query = """
                SELECT m.serial_number, m.pdi_number, c.company_name, m.pmax, m.binning, m.class_status
                FROM ftr_master_serials m
                JOIN companies c ON m.company_id = c.id
                WHERE m.binning IS NOT NULL AND m.binning != ''
            """
            if company_id:
                query += f" AND m.company_id = {company_id}"
            query += " ORDER BY c.company_name, m.binning, m.serial_number"
            
        elif export_type == 'rejected':
            # Export rejected modules
            ws.title = "Rejected Modules"
            headers = ["S.No", "Serial Number", "PDI Number", "Company", "Pmax", "Binning", "Class Status"]
            
            query = """
                SELECT m.serial_number, m.pdi_number, c.company_name, m.pmax, m.binning, m.class_status
                FROM ftr_master_serials m
                JOIN companies c ON m.company_id = c.id
                WHERE m.class_status = 'REJECTED' OR m.class_status = 'rejected'
            """
            if company_id:
                query += f" AND m.company_id = {company_id}"
            query += " ORDER BY c.company_name, m.serial_number"
        
        elif export_type == 'packed_not_pdi':
            # Export packed modules that are NOT in PDI and NOT rejected
            ws.title = "Packed Not in PDI"
            headers = ["S.No", "Barcode", "Company", "Running Order", "Binning", "Pallet No", "Pack Date", "PDI Status", "Class Status"]
            
            # Get from get_packed_not_in_pdi function
            result = get_packed_not_in_pdi(company_name)
            if not result.get('has_answer'):
                return jsonify({'success': False, 'error': 'Failed to fetch packed modules'}), 500
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # Collect all modules from all companies in result
            all_modules = []
            for comp_name, comp_data in result.get('data', {}).items():
                # Get MRP data for this company
                mrp_result = get_all_mrp_data(comp_name)
                if mrp_result.get('success'):
                    mrp_data = mrp_result.get('data', [])
                    # Filter to only packed modules from samples
                    for sample in comp_data.get('samples', []):
                        barcode = sample['barcode']
                        # Find full details from MRP
                        mrp_detail = next((b for b in mrp_data if b.get('barcode') == barcode), None)
                        if mrp_detail:
                            all_modules.append({
                                'barcode': barcode,
                                'company': comp_name,
                                'running_order': sample.get('running_order', ''),
                                'binning': sample.get('binning', ''),
                                'pallet': sample.get('pallet', ''),
                                'date': mrp_detail.get('date', ''),
                            })
            
            # Write data
            for idx, module in enumerate(all_modules, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                ws.cell(row=idx+1, column=2, value=module['barcode']).border = thin_border
                ws.cell(row=idx+1, column=3, value=module['company']).border = thin_border
                ws.cell(row=idx+1, column=4, value=module['running_order']).border = thin_border
                ws.cell(row=idx+1, column=5, value=module['binning']).border = thin_border
                ws.cell(row=idx+1, column=6, value=module['pallet']).border = thin_border
                ws.cell(row=idx+1, column=7, value=module['date']).border = thin_border
                ws.cell(row=idx+1, column=8, value='Not Assigned').border = thin_border
                ws.cell(row=idx+1, column=9, value='OK').border = thin_border
            
            # Auto-width columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"Packed_Not_in_PDI_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=filename)
            
        else:  # all - company summary
            ws.title = "Company Summary"
            headers = ["S.No", "Company", "Total FTR", "OK", "Rejected", "Available", "Assigned", "Packed (MRP)", "Dispatched (MRP)"]
            
            # Get all companies
            result = db.session.execute(text("SELECT id, company_name FROM companies"))
            companies = result.fetchall()
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            row_num = 2
            for idx, company in enumerate(companies, 1):
                cid, cname = company[0], company[1]
                
                # Get counts
                total_q = db.session.execute(text(f"SELECT COUNT(*) FROM ftr_master_serials WHERE company_id = {cid}")).fetchone()[0]
                ok_q = db.session.execute(text(f"SELECT COUNT(*) FROM ftr_master_serials WHERE company_id = {cid} AND (class_status IS NULL OR class_status != 'REJECTED')")).fetchone()[0]
                rejected_q = db.session.execute(text(f"SELECT COUNT(*) FROM ftr_master_serials WHERE company_id = {cid} AND class_status = 'REJECTED'")).fetchone()[0]
                available_q = db.session.execute(text(f"SELECT COUNT(*) FROM ftr_master_serials WHERE company_id = {cid} AND status = 'available'")).fetchone()[0]
                assigned_q = db.session.execute(text(f"SELECT COUNT(*) FROM ftr_master_serials WHERE company_id = {cid} AND status = 'assigned'")).fetchone()[0]
                
                # Get MRP data
                ext_data = get_external_packed_dispatch_data(cname)
                packed = ext_data.get('packed_count', 0) if ext_data.get('success') else 0
                dispatched = ext_data.get('dispatched_count', 0) if ext_data.get('success') else 0
                
                ws.cell(row=row_num, column=1, value=idx).border = thin_border
                ws.cell(row=row_num, column=2, value=cname).border = thin_border
                ws.cell(row=row_num, column=3, value=total_q).border = thin_border
                ws.cell(row=row_num, column=4, value=ok_q).border = thin_border
                ws.cell(row=row_num, column=5, value=rejected_q).border = thin_border
                ws.cell(row=row_num, column=6, value=available_q).border = thin_border
                ws.cell(row=row_num, column=7, value=assigned_q).border = thin_border
                ws.cell(row=row_num, column=8, value=packed).border = thin_border
                ws.cell(row=row_num, column=9, value=dispatched).border = thin_border
                row_num += 1
            
            # Auto-width columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            filename = f"FTR_Summary_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=filename)
        
        # For database queries (pending, binning, rejected)
        if export_type in ['pending', 'binning', 'rejected']:
            result = db.session.execute(text(query))
            rows = result.fetchall()
            
            # Write headers
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            # Write data
            for idx, row in enumerate(rows, 1):
                ws.cell(row=idx+1, column=1, value=idx).border = thin_border
                for col, val in enumerate(row, 2):
                    ws.cell(row=idx+1, column=col, value=val or '').border = thin_border
            
            # Auto-width columns
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            type_names = {'pending': 'Pending', 'binning': 'Binning', 'rejected': 'Rejected'}
            filename = f"{type_names.get(export_type, 'FTR')}_{company_name}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                           as_attachment=True, download_name=filename)
        
        return jsonify({'success': False, 'error': 'Invalid export type'}), 400
        
    except Exception as e:
        print(f"Excel export error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@ai_assistant_bp.route('/ai/check-barcodes', methods=['POST'])
def check_barcodes_status():
    """
    Check status of barcodes from uploaded Excel or list
    Returns packed/dispatched/pending status for each barcode
    """
    try:
        barcodes_to_check = []
        company_name = request.form.get('company_name', 'Rays Power')
        
        # Check if file uploaded
        if 'file' in request.files:
            file = request.files['file']
            if file.filename.endswith(('.xlsx', '.xls')):
                try:
                    import pandas as pd
                    df = pd.read_excel(file)
                    # Try to find barcode column
                    barcode_col = None
                    for col in df.columns:
                        col_lower = str(col).lower()
                        if any(x in col_lower for x in ['barcode', 'serial', 'id', 'module', 'sr']):
                            barcode_col = col
                            break
                    if barcode_col is None:
                        barcode_col = df.columns[0]  # Use first column
                    
                    barcodes_to_check = df[barcode_col].dropna().astype(str).tolist()
                except Exception as e:
                    return jsonify({'success': False, 'error': f'Excel parsing error: {str(e)}'}), 400
            elif file.filename.endswith('.csv'):
                import pandas as pd
                df = pd.read_csv(file)
                barcode_col = df.columns[0]
                barcodes_to_check = df[barcode_col].dropna().astype(str).tolist()
        else:
            # Get from JSON body
            data = request.get_json() if request.is_json else {}
            barcodes_to_check = data.get('barcodes', [])
            company_name = data.get('company_name', company_name)
        
        if not barcodes_to_check:
            return jsonify({'success': False, 'error': 'No barcodes provided'}), 400
        
        # Clean barcodes
        barcodes_to_check = [str(b).strip() for b in barcodes_to_check if str(b).strip()]
        
        # Fetch all data from MRP API for company
        mrp_party_name = get_mrp_party_name(company_name)
        
        try:
            response = requests.post(
                BARCODE_TRACKING_API,
                json={'party_name': mrp_party_name},
                timeout=60
            )
            mrp_data = response.json() if response.status_code == 200 else {'data': []}
        except:
            mrp_data = {'data': []}
        
        # Create lookup dictionaries from MRP data
        mrp_barcodes = {}
        for item in mrp_data.get('data', []):
            bc = item.get('barcode', '')
            mrp_barcodes[bc] = {
                'status': item.get('status', ''),
                'running_order': item.get('running_order', ''),
                'pallet_no': item.get('pallet_no', ''),
                'packed_party': item.get('packed_party', ''),
                'dispatch_party': item.get('dispatch_party', ''),
                'date': item.get('date', '')
            }
        
        # Check each barcode
        results = []
        packed_count = 0
        dispatched_count = 0
        not_found_count = 0
        pending_count = 0
        
        for barcode in barcodes_to_check:
            mrp_info = mrp_barcodes.get(barcode, {})
            
            is_packed = mrp_info.get('status') == 'packed' or mrp_info.get('packed_party')
            is_dispatched = bool(mrp_info.get('dispatch_party')) or mrp_info.get('status') == 'dispatched'
            
            # Extract binning from running_order
            import re
            running_order = mrp_info.get('running_order', '') or ''
            binning = ''
            bin_match = re.search(r'i-?(\d+)', running_order, re.IGNORECASE)
            if bin_match:
                binning = f"I{bin_match.group(1)}"
            
            if is_dispatched:
                status = '🚚 DISPATCHED'
                dispatched_count += 1
            elif is_packed:
                status = '📦 PACKED'
                packed_count += 1
            elif barcode in mrp_barcodes:
                status = '⏳ PENDING'
                pending_count += 1
            else:
                status = '❓ NOT FOUND'
                not_found_count += 1
            
            results.append({
                'barcode': barcode,
                'status': status,
                'running_order': running_order,
                'binning': binning,
                'pallet_no': mrp_info.get('pallet_no', ''),
                'dispatch_party': mrp_info.get('dispatch_party', ''),
                'date': mrp_info.get('date', '')
            })
        
        # Generate Excel report
        if EXCEL_AVAILABLE:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Barcode Status"
            
            # Styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            packed_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
            dispatched_fill = PatternFill(start_color="BDD7EE", end_color="BDD7EE", fill_type="solid")
            not_found_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            pending_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin')
            )
            
            headers = ["S.No", "Barcode", "Status", "Running Order", "Binning", "Pallet No", "Dispatch Party", "Date"]
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.border = thin_border
            
            for idx, result in enumerate(results, 1):
                row = idx + 1
                ws.cell(row=row, column=1, value=idx).border = thin_border
                ws.cell(row=row, column=2, value=result['barcode']).border = thin_border
                
                status_cell = ws.cell(row=row, column=3, value=result['status'])
                status_cell.border = thin_border
                if 'DISPATCHED' in result['status']:
                    status_cell.fill = dispatched_fill
                elif 'PACKED' in result['status']:
                    status_cell.fill = packed_fill
                elif 'NOT FOUND' in result['status']:
                    status_cell.fill = not_found_fill
                else:
                    status_cell.fill = pending_fill
                
                ws.cell(row=row, column=4, value=result['running_order']).border = thin_border
                ws.cell(row=row, column=5, value=result['binning']).border = thin_border
                ws.cell(row=row, column=6, value=result['pallet_no']).border = thin_border
                ws.cell(row=row, column=7, value=result['dispatch_party']).border = thin_border
                ws.cell(row=row, column=8, value=result['date']).border = thin_border
            
            # Summary row
            summary_row = len(results) + 3
            ws.cell(row=summary_row, column=1, value="SUMMARY:").font = Font(bold=True)
            ws.cell(row=summary_row + 1, column=1, value=f"Total Checked: {len(results)}")
            ws.cell(row=summary_row + 2, column=1, value=f"📦 Packed: {packed_count}")
            ws.cell(row=summary_row + 3, column=1, value=f"🚚 Dispatched: {dispatched_count}")
            ws.cell(row=summary_row + 4, column=1, value=f"⏳ Pending: {pending_count}")
            ws.cell(row=summary_row + 5, column=1, value=f"❓ Not Found: {not_found_count}")
            
            # Auto-width
            for col in ws.columns:
                max_length = max(len(str(cell.value or '')) for cell in col)
                ws.column_dimensions[col[0].column_letter].width = max_length + 2
            
            buffer = io.BytesIO()
            wb.save(buffer)
            buffer.seek(0)
            
            # Encode to base64 for JSON response
            import base64
            excel_base64 = base64.b64encode(buffer.getvalue()).decode('utf-8')
        else:
            excel_base64 = None
        
        return jsonify({
            'success': True,
            'total_checked': len(results),
            'summary': {
                'packed': packed_count,
                'dispatched': dispatched_count,
                'pending': pending_count,
                'not_found': not_found_count
            },
            'results': results[:100],  # Limit to first 100 for JSON response
            'excel_base64': excel_base64,
            'message': f"✅ Checked {len(results)} barcodes:\n📦 Packed: {packed_count}\n🚚 Dispatched: {dispatched_count}\n⏳ Pending: {pending_count}\n❓ Not Found: {not_found_count}"
        })
        
    except Exception as e:
        print(f"Check barcodes error: {str(e)}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@ai_assistant_bp.route('/ai/download-check-result', methods=['POST'])
def download_check_result():
    """Download the barcode check result as Excel"""
    try:
        data = request.json
        excel_base64 = data.get('excel_base64')
        
        if not excel_base64:
            return jsonify({'success': False, 'error': 'No Excel data'}), 400
        
        import base64
        buffer = io.BytesIO(base64.b64decode(excel_base64))
        buffer.seek(0)
        
        filename = f"Barcode_Status_Check_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(buffer, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                       as_attachment=True, download_name=filename)
        
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
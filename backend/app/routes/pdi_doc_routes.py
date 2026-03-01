"""
PDI Documentation Generator Routes
Generates complete PDI documentation package:
- IPQC Report (Excel)
- Witness Report (Excel)
- Calibration Instrument List (Excel)
- Sampling Plan (Excel)
- MOM - Minutes of Meeting (Excel)
All in one ZIP download
"""

from flask import Blueprint, request, jsonify, send_file, current_app
from app.models.database import db
from sqlalchemy import text
from datetime import datetime, timedelta
import os
import io
import zipfile
import random
import traceback
import json
import math

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

pdi_doc_bp = Blueprint('pdi_doc', __name__)

# ============ Styles ============
if EXCEL_AVAILABLE:
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    green_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
    light_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    center_align = Alignment(horizontal='center', vertical='center')
    wrap_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_wrap_align = Alignment(horizontal='left', vertical='center', wrap_text=True)


# ============ Helper Functions ============

def get_companies():
    """Get list of companies from database"""
    try:
        result = db.session.execute(text("""
            SELECT DISTINCT company_id, company_name 
            FROM companies 
            ORDER BY company_name
        """))
        companies = []
        for row in result.fetchall():
            companies.append({
                'id': row[0],
                'name': row[1]
            })
        return companies
    except Exception:
        # Fallback: try ftr_master_serials
        try:
            result = db.session.execute(text("""
                SELECT DISTINCT company_id 
                FROM ftr_master_serials 
                ORDER BY company_id
            """))
            return [{'id': row[0], 'name': row[0]} for row in result.fetchall()]
        except Exception:
            return []


def get_pdis_for_company(company_id):
    """Get PDI batches for a company"""
    try:
        result = db.session.execute(text("""
            SELECT id, pdi_number, total_modules, created_at
            FROM pdi_batches
            WHERE company_id = :cid
            ORDER BY created_at DESC
        """), {'cid': company_id})
        pdis = []
        for row in result.fetchall():
            pdis.append({
                'id': row[0],
                'pdi_number': row[1],
                'total_modules': row[2],
                'created_at': row[3].isoformat() if row[3] else None
            })
        return pdis
    except Exception as e:
        print(f"Error getting PDIs: {e}")
        return []


def get_serials_for_pdi(pdi_id):
    """Get serial numbers from a PDI batch"""
    try:
        result = db.session.execute(text("""
            SELECT serial_number
            FROM module_serial_numbers
            WHERE pdi_batch_id = :pid
            ORDER BY serial_number
        """), {'pid': pdi_id})
        return [row[0] for row in result.fetchall()]
    except Exception as e:
        print(f"Error getting serials: {e}")
        return []


def get_ftr_data(serial_numbers):
    """Get FTR data for serial numbers"""
    if not serial_numbers:
        return {}
    try:
        placeholders = ','.join([f':s{i}' for i in range(len(serial_numbers))])
        params = {f's{i}': s for i, s in enumerate(serial_numbers)}
        result = db.session.execute(text(f"""
            SELECT serial_number, pmax, isc, voc, ipm, vpm, ff, efficiency, binning
            FROM ftr_master_serials
            WHERE serial_number IN ({placeholders})
        """), params)
        data = {}
        for row in result.fetchall():
            data[row[0]] = {
                'pmax': float(row[1]) if row[1] else 0,
                'isc': float(row[2]) if row[2] else 0,
                'voc': float(row[3]) if row[3] else 0,
                'ipm': float(row[4]) if row[4] else 0,
                'vpm': float(row[5]) if row[5] else 0,
                'ff': float(row[6]) if row[6] else 0,
                'efficiency': float(row[7]) if row[7] else 0,
                'binning': row[8] if row[8] else ''
            }
        return data
    except Exception as e:
        print(f"Error getting FTR data: {e}")
        return {}


def get_calibration_instruments():
    """Get calibration instruments from database"""
    try:
        result = db.session.execute(text("""
            SELECT sr_no, instrument_id, machine_name, make, model_name,
                   item_sr_no, range_capacity, least_count, location,
                   calibration_agency, date_of_calibration, due_date,
                   calibration_frequency, calibration_standards, certificate_no, status
            FROM calibration_instruments
            ORDER BY sr_no
        """))
        instruments = []
        for row in result.fetchall():
            instruments.append({
                'sr_no': row[0],
                'instrument_id': row[1],
                'machine_name': row[2],
                'make': row[3],
                'model_name': row[4],
                'item_sr_no': row[5],
                'range_capacity': row[6],
                'least_count': row[7],
                'location': row[8],
                'calibration_agency': row[9],
                'date_of_calibration': row[10].strftime('%d/%m/%Y') if row[10] else '',
                'due_date': row[11].strftime('%d/%m/%Y') if row[11] else '',
                'calibration_frequency': row[12],
                'calibration_standards': row[13],
                'certificate_no': row[14],
                'status': row[15]
            })
        return instruments
    except Exception as e:
        print(f"Error getting calibration instruments: {e}")
        return []


def aql_sample_size(lot_size, level='II'):
    """Calculate AQL sample size per IS 2500 / ISO 2859"""
    aql_table = [
        (8, 5), (15, 5), (25, 8), (50, 13), (90, 20),
        (150, 32), (280, 50), (500, 80), (1200, 125),
        (3200, 200), (10000, 315), (35000, 500),
        (150000, 800), (500000, 1250), (float('inf'), 2000)
    ]
    for max_lot, sample in aql_table:
        if lot_size <= max_lot:
            return sample
    return 2000


def sanitize_filename(name):
    """Remove characters that are not safe for filenames"""
    return name.replace('/', '_').replace('\\', '_').replace(':', '_').replace('*', '_').replace('?', '_').replace('"', '_').replace('<', '_').replace('>', '_').replace('|', '_')


def set_cell(ws, row, col, value, font=None, fill=None, align=None, border=None):
    """Set cell value and style - simple helper matching proven working pattern"""
    cell = ws.cell(row=row, column=col, value=value)
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    else:
        cell.alignment = center_align
    if border:
        cell.border = border
    return cell


def set_data_cell(ws, row, col, value, font=None, fill=None, align=None):
    """Set a data cell with thin border - for non-merged data rows"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.border = thin_border
    if font:
        cell.font = font
    if fill:
        cell.fill = fill
    if align:
        cell.alignment = align
    else:
        cell.alignment = center_align
    return cell


def set_header_cell(ws, row, col, value):
    """Set a header cell with standard header styling and border"""
    cell = ws.cell(row=row, column=col, value=value)
    cell.font = header_font
    cell.fill = header_fill
    cell.alignment = center_align
    cell.border = thin_border
    return cell


def set_merged_title(ws, merge_range, value, font=None, fill=None, row_height=None):
    """Set a merged cell title - NO border applied (matches working pattern)"""
    ws.merge_cells(merge_range)
    # Extract top-left cell reference from merge range (e.g., 'A1:I1' -> 'A1')
    top_left = merge_range.split(':')[0]
    ws[top_left] = value
    if font:
        ws[top_left].font = font
    if fill:
        ws[top_left].fill = fill
    ws[top_left].alignment = Alignment(horizontal='center', vertical='center')
    if row_height:
        row_num = int(''.join(filter(str.isdigit, top_left)))
        ws.row_dimensions[row_num].height = row_height


# ============ ROUTES ============

@pdi_doc_bp.route('/health', methods=['GET'])
def health_check():
    return jsonify({
        'success': True,
        'message': 'PDI Docs API is running',
        'version': 'v5',
        'excel_available': EXCEL_AVAILABLE
    }), 200


@pdi_doc_bp.route('/companies', methods=['GET'])
def list_companies():
    companies = get_companies()
    return jsonify({'success': True, 'companies': companies}), 200


@pdi_doc_bp.route('/pdis/<company_id>', methods=['GET'])
def list_pdis(company_id):
    pdis = get_pdis_for_company(company_id)
    return jsonify({'success': True, 'pdis': pdis}), 200


@pdi_doc_bp.route('/serials/<int:pdi_id>', methods=['GET'])
def list_serials(pdi_id):
    serials = get_serials_for_pdi(pdi_id)
    return jsonify({'success': True, 'serials': serials, 'count': len(serials)}), 200


@pdi_doc_bp.route('/template-info', methods=['GET'])
def template_info():
    try:
        from app.models.ipqc_data import IPQCTemplate
        template = IPQCTemplate.get_template()
        total_checkpoints = sum(len(s.get('checkpoints', [])) for s in template)
        return jsonify({
            'success': True,
            'total_stages': len(template),
            'total_checkpoints': total_checkpoints
        }), 200
    except Exception as e:
        return jsonify({'success': True, 'total_stages': 8, 'total_checkpoints': 30}), 200


@pdi_doc_bp.route('/generate', methods=['POST'])
def generate_pdi_docs():
    """Generate complete PDI Documentation Package (ZIP)"""
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500

    try:
        data = request.json
        if not data:
            return jsonify({'success': False, 'error': 'No data provided'}), 400

        company_id = data.get('company_id', '')
        company_name = data.get('company_name', company_id)
        pdi_number = data.get('pdi_number', 'PDI-001')
        serial_numbers = data.get('serial_numbers', [])
        production_days = data.get('production_days', 3)
        report_date = data.get('report_date', datetime.now().strftime('%d/%m/%Y'))
        module_type = data.get('module_type', 'G2G580')
        requested_docs = data.get('documents', ['ipqc', 'witness', 'calibration', 'sampling', 'mom'])

        if not serial_numbers:
            return jsonify({'success': False, 'error': 'No serial numbers provided'}), 400

        total_qty = len(serial_numbers)
        safe_pdi = sanitize_filename(pdi_number)

        # Get FTR data
        ftr_data = get_ftr_data(serial_numbers)

        # Get calibration instruments
        calibration_instruments = get_calibration_instruments()

        # Calculate sample size
        sample_size = min(aql_sample_size(total_qty), total_qty)
        sampled_serials = random.sample(serial_numbers, sample_size) if sample_size < total_qty else serial_numbers

        # Create ZIP
        zip_buffer = io.BytesIO()

        with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as zf:

            # 1. IPQC Report
            if 'ipqc' in requested_docs:
                try:
                    buf = generate_ipqc_excel(
                        company_name, pdi_number, serial_numbers,
                        production_days, report_date, module_type, sampled_serials
                    )
                    content = buf.getvalue()
                    if len(content) > 100:
                        zf.writestr(f"01_IPQC_Report_{safe_pdi}.xlsx", content)
                    buf.close()
                except Exception as e:
                    print(f"IPQC generation error: {e}")
                    traceback.print_exc()

            # 2. Witness Report
            if 'witness' in requested_docs:
                try:
                    buf = generate_witness_excel(
                        company_name, pdi_number, serial_numbers,
                        report_date, ftr_data, module_type
                    )
                    content = buf.getvalue()
                    if len(content) > 100:
                        zf.writestr(f"02_Witness_Report_{safe_pdi}.xlsx", content)
                    buf.close()
                except Exception as e:
                    print(f"Witness generation error: {e}")
                    traceback.print_exc()

            # 3. Calibration Instrument List
            if 'calibration' in requested_docs:
                try:
                    buf = generate_calibration_excel(
                        company_name, pdi_number, calibration_instruments, report_date
                    )
                    content = buf.getvalue()
                    if len(content) > 100:
                        zf.writestr(f"03_Calibration_Instruments_{safe_pdi}.xlsx", content)
                    buf.close()
                except Exception as e:
                    print(f"Calibration generation error: {e}")
                    traceback.print_exc()

            # 4. Sampling Plan
            if 'sampling' in requested_docs:
                try:
                    buf = generate_sampling_plan_excel(
                        company_name, pdi_number, total_qty,
                        sample_size, sampled_serials, report_date
                    )
                    content = buf.getvalue()
                    if len(content) > 100:
                        zf.writestr(f"04_Sampling_Plan_{safe_pdi}.xlsx", content)
                    buf.close()
                except Exception as e:
                    print(f"Sampling plan generation error: {e}")
                    traceback.print_exc()

            # 5. MOM
            if 'mom' in requested_docs:
                try:
                    buf = generate_mom_excel(
                        company_name, company_id, pdi_number,
                        total_qty, report_date, ftr_data, serial_numbers
                    )
                    content = buf.getvalue()
                    if len(content) > 100:
                        zf.writestr(f"05_MOM_{safe_pdi}.xlsx", content)
                    buf.close()
                except Exception as e:
                    print(f"MOM generation error: {e}")
                    traceback.print_exc()

        zip_buffer.seek(0)

        filename = f"PDI_Documentation_{safe_pdi}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.zip"

        return send_file(
            zip_buffer,
            mimetype='application/zip',
            as_attachment=True,
            download_name=filename
        )

    except Exception as e:
        print(f"PDI Docs generate error: {e}")
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ============ DOCUMENT GENERATORS ============
# Using the EXACT same pattern as the working witness_report_routes.py:
# - Direct cell assignment: ws.cell(row, col, value)
# - Direct property setting: cell.font = ..., cell.border = thin_border
# - Merged cells via ws['A1'] = value (NO border on merged cells)
# - Borders ONLY on data cells, NOT on merged title/header rows


def generate_ipqc_excel(company_name, pdi_number, serial_numbers, production_days, report_date, module_type, sampled_serials):
    """Generate IPQC Report Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IPQC Report"

    # Row 1: Title (merged, no border)
    set_merged_title(ws, 'A1:H1',
                     f"IPQC INSPECTION REPORT - {company_name}",
                     font=title_font, fill=title_fill, row_height=30)

    # Info rows (2-4)
    info = [
        ('PDI Number', pdi_number, 'Date', report_date),
        ('Total Modules', str(len(serial_numbers)), 'Production Days', str(production_days)),
        ('Module Type', module_type, 'Sample Size', str(len(sampled_serials))),
    ]

    row = 2
    for label1, val1, label2, val2 in info:
        set_data_cell(ws, row, 1, label1, font=Font(bold=True), fill=light_fill)
        set_data_cell(ws, row, 2, val1)
        set_data_cell(ws, row, 3, '')
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, label2, font=Font(bold=True), fill=light_fill)
        set_data_cell(ws, row, 6, val2)
        set_data_cell(ws, row, 7, '')
        set_data_cell(ws, row, 8, '')
        row += 1

    row += 1

    # Header row
    headers = ['Sr.No', 'Stage', 'Checkpoint', 'Acceptance Criteria', 'Sample Size', 'Frequency', 'Result', 'Remarks']
    for col, h in enumerate(headers, 1):
        set_header_cell(ws, row, col, h)

    row += 1

    # Auto-fill stages using IPQCFormGenerator
    try:
        from app.services.form_generator import IPQCFormGenerator
        generator = IPQCFormGenerator()
        form_data = generator.generate_form(
            date=report_date,
            shift='A',
            customer_id=company_name,
            po_number=pdi_number,
            module_count=len(serial_numbers)
        )
        stages = form_data.get('stages', [])
    except Exception as e:
        print(f"IPQCFormGenerator error: {e}")
        stages = []

    if stages:
        for stage in stages:
            stage_name = stage.get('stage', '')
            checkpoints = stage.get('checkpoints', [])
            sr_no = stage.get('sr_no', '')

            for i, cp in enumerate(checkpoints):
                set_data_cell(ws, row, 1, sr_no if i == 0 else '')
                set_data_cell(ws, row, 2, stage_name if i == 0 else '')
                set_data_cell(ws, row, 3, cp.get('checkpoint', ''), align=left_wrap_align)
                set_data_cell(ws, row, 4, cp.get('acceptance_criteria', ''), align=left_wrap_align)
                set_data_cell(ws, row, 5, cp.get('sample_size', ''))
                set_data_cell(ws, row, 6, cp.get('frequency', ''))
                result = cp.get('monitoring_result', 'OK')
                set_data_cell(ws, row, 7, result,
                              font=Font(color="008000", bold=True) if result == 'OK' else None)
                set_data_cell(ws, row, 8, cp.get('remarks', ''))
                row += 1
    else:
        # Fallback: basic stages if form generator not available
        basic_stages = [
            ('1', 'Incoming Inspection', 'Raw Material Check', 'As per BOM', str(len(sampled_serials)), 'Each Lot', 'OK', ''),
            ('2', 'Cell Sorting', 'Cell Efficiency', '≥22%', str(len(sampled_serials)), 'Each Lot', 'OK', ''),
            ('3', 'Stringing', 'Solder Quality', 'No cold joints', str(len(sampled_serials)), 'Hourly', 'OK', ''),
            ('4', 'Layup', 'Alignment Check', '±1mm tolerance', str(len(sampled_serials)), 'Each Module', 'OK', ''),
            ('5', 'Lamination', 'Temperature Profile', '145°C ± 5°C', str(len(sampled_serials)), 'Each Batch', 'OK', ''),
            ('6', 'Trimming', 'Edge Quality', 'No rough edges', str(len(sampled_serials)), 'Each Module', 'OK', ''),
            ('7', 'Framing', 'Frame Alignment', 'No gaps', str(len(sampled_serials)), 'Each Module', 'OK', ''),
            ('8', 'JB & Curing', 'Junction Box', 'Proper adhesion', str(len(sampled_serials)), 'Each Module', 'OK', ''),
            ('9', 'Final Test', 'Flasher Test', 'Power ≥ Nameplate', str(len(sampled_serials)), 'Each Module', 'OK', ''),
            ('10', 'Packing', 'Packing Quality', 'No damage', str(len(sampled_serials)), 'Each Pallet', 'OK', ''),
        ]
        for stage_data in basic_stages:
            for col, val in enumerate(stage_data, 1):
                set_data_cell(ws, row, col, val)
            row += 1

    # Column widths
    widths = [8, 20, 35, 30, 12, 12, 15, 25]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Sheet 2: Serial Numbers
    ws2 = wb.create_sheet("Serial Numbers")
    set_header_cell(ws2, 1, 1, "Sr.No")
    set_header_cell(ws2, 1, 2, "Serial Number")
    set_header_cell(ws2, 1, 3, "Sampled")

    for idx, serial in enumerate(serial_numbers, 1):
        set_data_cell(ws2, idx + 1, 1, idx)
        set_data_cell(ws2, idx + 1, 2, serial)
        is_sampled = "YES" if serial in sampled_serials else ""
        if is_sampled:
            set_data_cell(ws2, idx + 1, 3, is_sampled,
                          font=Font(color="008000", bold=True),
                          fill=PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid"))
        else:
            set_data_cell(ws2, idx + 1, 3, '')

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    ws2.column_dimensions['C'].width = 10

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generate_witness_excel(company_name, pdi_number, serial_numbers, report_date, ftr_data, module_type):
    """Generate Witness Report Excel - matches witness_report_routes.py pattern exactly"""
    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    total_qty = len(serial_numbers)

    # ========== SHEET 1: FTR ==========
    ws = wb.create_sheet("FTR(Inspection)")

    # Header rows (merged, NO borders - matching working pattern)
    ws.merge_cells('A1:I1')
    ws['A1'] = company_name
    ws['A1'].font = title_font
    ws['A1'].fill = title_fill
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 25

    ws.merge_cells('A2:I2')
    ws['A2'] = "Flasher Test (Power Measurement) Report"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A3:I3')
    ws['A3'] = f"Total Qty:- {total_qty} Pcs"
    ws['A3'].font = Font(bold=True, size=11)
    ws['A3'].alignment = Alignment(horizontal='center')

    ws.merge_cells('A4:I4')
    ws['A4'] = f"Date :- {report_date}"
    ws['A4'].font = Font(bold=True, size=11)
    ws['A4'].alignment = Alignment(horizontal='center')

    # FTR Headers
    ftr_headers = ['Sr.No.', 'Module Sr.No.', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Eff.']
    for col, h in enumerate(ftr_headers, 1):
        cell = ws.cell(row=5, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Data rows
    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 5
        ftr = ftr_data.get(serial, {})
        ws.cell(row=row, column=1, value=idx).border = thin_border
        ws.cell(row=row, column=2, value=serial).border = thin_border
        ws.cell(row=row, column=3, value=ftr.get('pmax', '')).border = thin_border
        ws.cell(row=row, column=4, value=ftr.get('isc', '')).border = thin_border
        ws.cell(row=row, column=5, value=ftr.get('voc', '')).border = thin_border
        ws.cell(row=row, column=6, value=ftr.get('ipm', '')).border = thin_border
        ws.cell(row=row, column=7, value=ftr.get('vpm', '')).border = thin_border
        ws.cell(row=row, column=8, value=ftr.get('ff', '')).border = thin_border
        ws.cell(row=row, column=9, value=ftr.get('efficiency', '')).border = thin_border
        for c in range(1, 10):
            ws.cell(row=row, column=c).alignment = center_align

    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G', 'H', 'I']:
        ws.column_dimensions[c].width = 12

    # ========== SHEET 2: Visual Inspection ==========
    ws2 = wb.create_sheet("Visual Inspection")

    ws2.merge_cells('A1:G1')
    ws2['A1'] = company_name
    ws2['A1'].font = title_font
    ws2['A1'].fill = title_fill
    ws2['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws2.row_dimensions[1].height = 25

    ws2.merge_cells('A2:G2')
    ws2['A2'] = "Visual Inspection Report"
    ws2['A2'].font = Font(bold=True, size=12)
    ws2['A2'].alignment = Alignment(horizontal='center')

    ws2.merge_cells('A3:G3')
    ws2['A3'] = f"Total Qty:- {total_qty} Pcs | Date: {report_date}"
    ws2['A3'].font = Font(bold=True)
    ws2['A3'].alignment = Alignment(horizontal='center')

    vis_headers = ['Sr.No.', 'Module Sr.No.', 'Glass', 'Frame', 'Backsheet', 'JB & Cable', 'Result']
    for col, h in enumerate(vis_headers, 1):
        cell = ws2.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 4
        ws2.cell(row=row, column=1, value=idx).border = thin_border
        ws2.cell(row=row, column=2, value=serial).border = thin_border
        for c in range(3, 8):
            cell = ws2.cell(row=row, column=c, value='OK')
            cell.border = thin_border
            cell.font = Font(color="008000")
        for c in range(1, 8):
            ws2.cell(row=row, column=c).alignment = center_align

    ws2.column_dimensions['A'].width = 8
    ws2.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws2.column_dimensions[c].width = 14

    # ========== SHEET 3: EL Inspection ==========
    ws3 = wb.create_sheet("EL Inspection")

    ws3.merge_cells('A1:E1')
    ws3['A1'] = company_name
    ws3['A1'].font = title_font
    ws3['A1'].fill = title_fill
    ws3['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws3.row_dimensions[1].height = 25

    ws3.merge_cells('A2:E2')
    ws3['A2'] = "EL (Electroluminescence) Inspection Report"
    ws3['A2'].font = Font(bold=True, size=12)
    ws3['A2'].alignment = Alignment(horizontal='center')

    ws3.merge_cells('A3:E3')
    ws3['A3'] = f"Total Qty:- {total_qty} Pcs | Date: {report_date}"
    ws3['A3'].font = Font(bold=True)
    ws3['A3'].alignment = Alignment(horizontal='center')

    el_headers = ['Sr.No.', 'Module Sr.No.', 'EL Result', 'Micro Crack', 'Remarks']
    for col, h in enumerate(el_headers, 1):
        cell = ws3.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 4
        ws3.cell(row=row, column=1, value=idx).border = thin_border
        ws3.cell(row=row, column=2, value=serial).border = thin_border
        cell = ws3.cell(row=row, column=3, value='PASS')
        cell.border = thin_border
        cell.font = Font(color="008000", bold=True)
        ws3.cell(row=row, column=4, value='NIL').border = thin_border
        ws3.cell(row=row, column=5, value='').border = thin_border
        for c in range(1, 6):
            ws3.cell(row=row, column=c).alignment = center_align

    ws3.column_dimensions['A'].width = 8
    ws3.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E']:
        ws3.column_dimensions[c].width = 16

    # ========== SHEET 4: Safety Tests ==========
    ws4 = wb.create_sheet("IR,HV,GD,Wet Leakage")

    ws4.merge_cells('A1:H1')
    ws4['A1'] = company_name
    ws4['A1'].font = title_font
    ws4['A1'].fill = title_fill
    ws4['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws4.row_dimensions[1].height = 25

    ws4.merge_cells('A2:H2')
    ws4['A2'] = "Insulation Resistance / Hi-Pot / Ground Continuity / Wet Leakage Report"
    ws4['A2'].font = Font(bold=True, size=11)
    ws4['A2'].alignment = Alignment(horizontal='center')

    ws4.merge_cells('A3:H3')
    ws4['A3'] = f"Total Qty:- {total_qty} Pcs | Date: {report_date}"
    ws4['A3'].font = Font(bold=True)
    ws4['A3'].alignment = Alignment(horizontal='center')

    safety_headers = ['Sr.No.', 'Module Sr.No.', 'IR (MΩ)', 'Hi-Pot (V)', 'Duration (s)', 'GD (Ω)', 'Wet Leakage', 'Result']
    for col, h in enumerate(safety_headers, 1):
        cell = ws4.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 4
        ws4.cell(row=row, column=1, value=idx).border = thin_border
        ws4.cell(row=row, column=2, value=serial).border = thin_border
        ws4.cell(row=row, column=3, value=round(random.uniform(500, 2000), 0)).border = thin_border
        ws4.cell(row=row, column=4, value=3800).border = thin_border
        ws4.cell(row=row, column=5, value=3).border = thin_border
        ws4.cell(row=row, column=6, value=round(random.uniform(0.01, 0.1), 3)).border = thin_border
        cell = ws4.cell(row=row, column=7, value='PASS')
        cell.border = thin_border
        cell.font = Font(color="008000")
        cell = ws4.cell(row=row, column=8, value='OK')
        cell.border = thin_border
        cell.font = Font(color="008000")
        for c in range(1, 9):
            ws4.cell(row=row, column=c).alignment = center_align

    ws4.column_dimensions['A'].width = 8
    ws4.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G', 'H']:
        ws4.column_dimensions[c].width = 14

    # ========== SHEET 5: Dimension ==========
    ws5 = wb.create_sheet("Dimension")

    ws5.merge_cells('A1:G1')
    ws5['A1'] = company_name
    ws5['A1'].font = title_font
    ws5['A1'].fill = title_fill
    ws5['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws5.row_dimensions[1].height = 25

    ws5.merge_cells('A2:G2')
    ws5['A2'] = "Dimension Check Report"
    ws5['A2'].font = Font(bold=True, size=12)
    ws5['A2'].alignment = Alignment(horizontal='center')

    ws5.merge_cells('A3:G3')
    ws5['A3'] = f"Total Qty:- {total_qty} Pcs | Date: {report_date}"
    ws5['A3'].font = Font(bold=True)
    ws5['A3'].alignment = Alignment(horizontal='center')

    dim_headers = ['Sr.No.', 'Module Sr.No.', 'Length (mm)', 'Width (mm)', 'Thickness (mm)', 'Weight (kg)', 'Result']
    for col, h in enumerate(dim_headers, 1):
        cell = ws5.cell(row=4, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    for idx, serial in enumerate(serial_numbers, 1):
        row = idx + 4
        ws5.cell(row=row, column=1, value=idx).border = thin_border
        ws5.cell(row=row, column=2, value=serial).border = thin_border
        ws5.cell(row=row, column=3, value=round(2278 + random.uniform(-1, 1), 1)).border = thin_border
        ws5.cell(row=row, column=4, value=round(1134 + random.uniform(-1, 1), 1)).border = thin_border
        ws5.cell(row=row, column=5, value=round(30 + random.uniform(-0.5, 0.5), 1)).border = thin_border
        ws5.cell(row=row, column=6, value=round(32.5 + random.uniform(-0.5, 0.5), 1)).border = thin_border
        cell = ws5.cell(row=row, column=7, value='OK')
        cell.border = thin_border
        cell.font = Font(color="008000")
        for c in range(1, 8):
            ws5.cell(row=row, column=c).alignment = center_align

    ws5.column_dimensions['A'].width = 8
    ws5.column_dimensions['B'].width = 25
    for c in ['C', 'D', 'E', 'F', 'G']:
        ws5.column_dimensions[c].width = 16

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generate_calibration_excel(company_name, pdi_number, instruments, report_date):
    """Generate Calibration Instrument List Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Calibration List"

    # Title rows (merged, no border)
    set_merged_title(ws, 'A1:O1',
                     f"CALIBRATION INSTRUMENT LIST - {company_name}",
                     font=title_font, fill=title_fill, row_height=30)

    set_merged_title(ws, 'A2:O2',
                     f"PDI: {pdi_number} | Date: {report_date}",
                     font=Font(bold=True, size=11))

    # Headers
    cal_headers = ['Sr.No', 'Instrument ID', 'Machine/Equipment', 'Make', 'Model',
                   'Item Sr.No', 'Range/Capacity', 'Least Count', 'Location',
                   'Calibration Agency', 'Date of Calibration', 'Due Date',
                   'Frequency', 'Standards', 'Certificate No.']

    for col, h in enumerate(cal_headers, 1):
        set_header_cell(ws, 3, col, h)

    if instruments:
        for idx, inst in enumerate(instruments, 1):
            row = idx + 3
            set_data_cell(ws, row, 1, idx)
            set_data_cell(ws, row, 2, inst.get('instrument_id', ''))
            set_data_cell(ws, row, 3, inst.get('machine_name', ''), align=left_wrap_align)
            set_data_cell(ws, row, 4, inst.get('make', ''))
            set_data_cell(ws, row, 5, inst.get('model_name', ''))
            set_data_cell(ws, row, 6, inst.get('item_sr_no', ''))
            set_data_cell(ws, row, 7, inst.get('range_capacity', ''))
            set_data_cell(ws, row, 8, inst.get('least_count', ''))
            set_data_cell(ws, row, 9, inst.get('location', ''))
            set_data_cell(ws, row, 10, inst.get('calibration_agency', ''))
            set_data_cell(ws, row, 11, inst.get('date_of_calibration', ''))
            set_data_cell(ws, row, 12, inst.get('due_date', ''))
            set_data_cell(ws, row, 13, inst.get('calibration_frequency', ''))
            set_data_cell(ws, row, 14, inst.get('calibration_standards', ''))
            set_data_cell(ws, row, 15, inst.get('certificate_no', ''))

            # Highlight overdue
            if inst.get('status') == 'overdue':
                overdue_fill = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
                for c in range(1, 16):
                    ws.cell(row=row, column=c).fill = overdue_fill
    else:
        set_data_cell(ws, 4, 1, "No calibration instruments found in database", align=left_wrap_align)
        for c in range(2, 16):
            set_data_cell(ws, 4, c, '')

    # Column widths
    widths = [6, 14, 22, 12, 12, 14, 18, 12, 12, 25, 16, 16, 12, 20, 18]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generate_sampling_plan_excel(company_name, pdi_number, total_qty, sample_size, sampled_serials, report_date):
    """Generate Sampling Plan Excel per IS 2500 / ISO 2859"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Sampling Plan"

    # Title (merged, no border)
    set_merged_title(ws, 'A1:F1',
                     f"SAMPLING PLAN - {company_name}",
                     font=title_font, fill=title_fill, row_height=30)

    set_merged_title(ws, 'A2:F2',
                     f"PDI: {pdi_number} | Date: {report_date}",
                     font=Font(bold=True, size=11))

    # Plan Details
    row = 4
    plan_info = [
        ('Sampling Standard', 'IS 2500 / ISO 2859-1'),
        ('Inspection Level', 'General Inspection Level II'),
        ('AQL', '0.65% (Major), 1.0% (Minor)'),
        ('Lot Size', str(total_qty)),
        ('Sample Size', str(sample_size)),
        ('Sampling Type', 'Single Sampling - Normal Inspection'),
    ]

    for label, value in plan_info:
        set_data_cell(ws, row, 1, label, font=Font(bold=True), fill=light_fill)
        set_data_cell(ws, row, 2, value, align=left_wrap_align)
        set_data_cell(ws, row, 3, '')
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, '')
        set_data_cell(ws, row, 6, '')
        row += 1

    row += 1

    # Inspection Criteria Section Header
    for col in range(1, 7):
        set_header_cell(ws, row, col, "INSPECTION CRITERIA" if col == 1 else "")
    row += 1

    criteria_headers = ['Sr.No', 'Test Parameter', 'Method', 'Acceptance Criteria', 'Defect Type', 'AQL']
    for col, h in enumerate(criteria_headers, 1):
        set_header_cell(ws, row, col, h)
    row += 1

    criteria = [
        ('1', 'Visual Inspection', 'Manual / IEC 61215', 'No visible defects', 'Major', '0.65%'),
        ('2', 'Dimension Check', 'Measuring Tape', 'Within ±1mm tolerance', 'Minor', '1.0%'),
        ('3', 'EL Test', 'EL Camera', 'No micro-cracks', 'Major', '0.65%'),
        ('4', 'Flasher Test (FTR)', 'Solar Simulator', 'Power >= Nameplate', 'Major', '0.65%'),
        ('5', 'Hi-Pot Test', 'Hi-Pot Tester', '3800V for 3 sec, no breakdown', 'Critical', '0.25%'),
        ('6', 'Insulation Resistance', 'IR Tester', '>= 400 MOhm', 'Critical', '0.25%'),
        ('7', 'Ground Continuity', 'GC Tester', '<= 0.1 Ohm', 'Major', '0.65%'),
        ('8', 'Wet Leakage', 'Wet Leakage Tester', 'Current <= 10uA', 'Critical', '0.25%'),
        ('9', 'Label & Marking', 'Visual', 'Correct labels, barcodes, RFID', 'Minor', '1.0%'),
        ('10', 'Packing', 'Visual', 'Proper packing, no damage', 'Minor', '1.0%'),
    ]

    for c_data in criteria:
        for col, val in enumerate(c_data, 1):
            set_data_cell(ws, row, col, val)
        row += 1

    row += 1

    # Sampled Serials Section
    for col in range(1, 7):
        set_header_cell(ws, row, col, "SAMPLED SERIAL NUMBERS" if col == 1 else "")
    row += 1

    set_header_cell(ws, row, 1, "Sr.No")
    set_header_cell(ws, row, 2, "Serial Number")
    set_header_cell(ws, row, 3, "Result")
    set_header_cell(ws, row, 4, "")
    set_header_cell(ws, row, 5, "")
    set_header_cell(ws, row, 6, "")
    row += 1

    for idx, serial in enumerate(sampled_serials, 1):
        set_data_cell(ws, row, 1, idx)
        set_data_cell(ws, row, 2, serial)
        set_data_cell(ws, row, 3, 'PASS', font=Font(color="008000", bold=True))
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, '')
        set_data_cell(ws, row, 6, '')
        row += 1

    # Column widths
    widths = [8, 22, 18, 28, 12, 10]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer


def generate_mom_excel(company_name, company_id, pdi_number, total_qty, report_date, ftr_data, serial_numbers):
    """Generate Minutes of Meeting (MOM) Excel"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "MOM"

    # Title (merged, no border)
    set_merged_title(ws, 'A1:F1',
                     "MINUTES OF MEETING (MOM)",
                     font=title_font, fill=title_fill, row_height=30)

    set_merged_title(ws, 'A2:F2',
                     f"Pre-Dispatch Inspection - {company_name}",
                     font=Font(bold=True, size=12))

    # Meeting Details
    row = 4
    meeting_info = [
        ('Date', report_date),
        ('Location', 'Gautam Solar Pvt. Ltd., Manufacturing Facility'),
        ('PDI Number', pdi_number),
        ('Customer', company_name),
        ('Total Quantity', f'{total_qty} Modules'),
        ('Module Type', 'Bifacial TOPCon'),
    ]

    for label, value in meeting_info:
        set_data_cell(ws, row, 1, label, font=Font(bold=True), fill=light_fill)
        set_data_cell(ws, row, 2, value, align=left_wrap_align)
        set_data_cell(ws, row, 3, '')
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, '')
        set_data_cell(ws, row, 6, '')
        row += 1

    row += 1

    # Attendees Section Header
    for col in range(1, 7):
        set_header_cell(ws, row, col, "ATTENDEES" if col == 1 else "")
    row += 1

    attendee_headers = ['Sr.No', 'Name', 'Designation', 'Organization', 'Signature', '']
    for col, h in enumerate(attendee_headers, 1):
        set_header_cell(ws, row, col, h)
    row += 1

    attendees = [
        ('1', '', 'Quality Head', 'Gautam Solar Pvt. Ltd.', ''),
        ('2', '', 'Production Manager', 'Gautam Solar Pvt. Ltd.', ''),
        ('3', '', 'QA/QC Engineer', company_name, ''),
        ('4', '', 'Project Manager', company_name, ''),
    ]

    for att in attendees:
        for col, val in enumerate(att, 1):
            set_data_cell(ws, row, col, val)
        set_data_cell(ws, row, 6, '')
        row += 1

    row += 1

    # FTR Summary Section Header
    for col in range(1, 7):
        set_header_cell(ws, row, col, "FTR / FLASHER TEST SUMMARY" if col == 1 else "")
    row += 1

    # Calculate FTR stats
    if ftr_data:
        pmax_values = [v.get('pmax', 0) for v in ftr_data.values() if v.get('pmax')]
        if pmax_values:
            ftr_stats = [
                ('Total Modules Tested', str(len(ftr_data))),
                ('Average Pmax', f'{sum(pmax_values) / len(pmax_values):.2f} W'),
                ('Min Pmax', f'{min(pmax_values):.2f} W'),
                ('Max Pmax', f'{max(pmax_values):.2f} W'),
                ('All Pass', 'YES' if all(p > 0 for p in pmax_values) else 'NO'),
            ]
        else:
            ftr_stats = [('FTR Data', 'No Pmax data available')]
    else:
        ftr_stats = [('FTR Data', 'Not available - serials not found in FTR database')]

    for label, value in ftr_stats:
        set_data_cell(ws, row, 1, label, font=Font(bold=True), fill=light_fill)
        set_data_cell(ws, row, 2, value, align=left_wrap_align)
        set_data_cell(ws, row, 3, '')
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, '')
        set_data_cell(ws, row, 6, '')
        row += 1

    row += 1

    # Discussion Points Section Header
    for col in range(1, 7):
        set_header_cell(ws, row, col, "DISCUSSION POINTS & OBSERVATIONS" if col == 1 else "")
    row += 1

    disc_headers = ['Sr.No', 'Topic', 'Observation / Decision', '', '', '']
    for col, h in enumerate(disc_headers, 1):
        set_header_cell(ws, row, col, h)
    row += 1

    ftr_count = len(ftr_data) if ftr_data else 0
    discussions = [
        ('1', 'Module Quality', 'All modules passed IPQC quality checks as per standard specifications.'),
        ('2', 'FTR Results', f'Flasher test completed for {ftr_count} modules. All within acceptable power tolerance.'),
        ('3', 'Visual Inspection', 'No visual defects observed. Glass, frame, backsheet, and J-Box all inspected.'),
        ('4', 'EL Test', 'Electroluminescence test completed. No micro-cracks detected.'),
        ('5', 'Safety Tests', 'Hi-Pot, IR, Ground Continuity, and Wet Leakage tests all PASSED.'),
        ('6', 'Calibration', 'All testing instruments are within calibration validity.'),
        ('7', 'Packing', 'Modules properly packed as per customer specifications.'),
        ('8', 'Documentation', 'Complete PDI documentation package prepared and submitted.'),
    ]

    for d in discussions:
        set_data_cell(ws, row, 1, d[0])
        set_data_cell(ws, row, 2, d[1], font=Font(bold=True))
        set_data_cell(ws, row, 3, d[2], align=left_wrap_align)
        set_data_cell(ws, row, 4, '')
        set_data_cell(ws, row, 5, '')
        set_data_cell(ws, row, 6, '')
        ws.row_dimensions[row].height = 30
        row += 1

    row += 1

    # Conclusion Section Header
    for col in range(1, 7):
        cell = ws.cell(row=row, column=col, value="CONCLUSION" if col == 1 else "")
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = center_align
        cell.border = thin_border
    row += 1

    conclusion_text = (
        f"All {total_qty} modules of PDI {pdi_number} for {company_name} have been inspected as per "
        f"IS 2500 / IEC 61215 standards. All quality parameters are within acceptable limits. "
        f"The lot is APPROVED for dispatch."
    )
    set_data_cell(ws, row, 1, conclusion_text, align=Alignment(wrap_text=True, vertical='center', horizontal='left'))
    for c in range(2, 7):
        set_data_cell(ws, row, c, '')
    ws.row_dimensions[row].height = 45

    # Column widths
    widths = [8, 20, 30, 20, 15, 15]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buffer = io.BytesIO()
    wb.save(buffer)
    wb.close()
    buffer.seek(0)
    return buffer

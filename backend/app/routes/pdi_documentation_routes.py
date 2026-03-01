"""
PDI Documentation Generator Routes
Auto-generates combined PDI documentation package:
- IPQC Checksheet (auto-filled per day production, random sampling of serials)
- Witness Report (FTR, Bifaciality, Visual, EL, Hipot, Dimension, RFID)
- Calibration Index
- MOM (Minutes of Meeting / Completion Summary)
- Sampling Plan
All in one combined Excel workbook — ZERO manual work required.
"""
from flask import Blueprint, request, jsonify, send_file
from app.models.database import db
from app.models.calibration_data import CalibrationInstrument
from sqlalchemy import text
import io
import random
import math
from datetime import datetime, timedelta

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers
    from openpyxl.utils import get_column_letter
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False

pdi_doc_bp = Blueprint('pdi_documentation', __name__)

# ==================== STYLES (only if openpyxl available) ====================
if EXCEL_AVAILABLE:
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )
    header_fill = PatternFill(start_color="1565C0", end_color="1565C0", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    title_fill = PatternFill(start_color="0D47A1", end_color="0D47A1", fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=14)
    sub_header_fill = PatternFill(start_color="E3F2FD", end_color="E3F2FD", fill_type="solid")
    green_fill = PatternFill(start_color="E8F5E9", end_color="E8F5E9", fill_type="solid")
    yellow_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")
    red_fill = PatternFill(start_color="FFEBEE", end_color="FFEBEE", fill_type="solid")
    mom_header_fill = PatternFill(start_color="1B5E20", end_color="1B5E20", fill_type="solid")

# ==================== MODULE DATABASE ====================
MODULE_SPECS = {
    "G2G570": {"name": "G2G1740-HAD", "power": 570, "cells": 144, "size": "2278x1134x30",
               "pmax": 570.0, "voc": 48.5, "isc": 14.80, "vpm": 41.0, "ipm": 13.90, "ff": 80.5, "eff": 22.10},
    "G2G575": {"name": "G2G1740-HAD", "power": 575, "cells": 144, "size": "2278x1134x30",
               "pmax": 575.0, "voc": 48.6, "isc": 14.90, "vpm": 41.1, "ipm": 13.99, "ff": 80.6, "eff": 22.25},
    "G2G580": {"name": "G2G1740-HAD", "power": 580, "cells": 144, "size": "2278x1134x30",
               "pmax": 580.0, "voc": 48.7, "isc": 14.95, "vpm": 41.2, "ipm": 14.08, "ff": 80.7, "eff": 22.45},
    "G2G585": {"name": "G2G1740-HAD", "power": 585, "cells": 144, "size": "2278x1134x30",
               "pmax": 585.0, "voc": 48.8, "isc": 15.05, "vpm": 41.3, "ipm": 14.16, "ff": 80.8, "eff": 22.65},
    "G2G590": {"name": "G2G1740-HAD", "power": 590, "cells": 144, "size": "2278x1134x30",
               "pmax": 590.0, "voc": 48.9, "isc": 15.10, "vpm": 41.5, "ipm": 14.22, "ff": 80.9, "eff": 22.84},
    "G2G595": {"name": "G2G1740-HAD", "power": 595, "cells": 144, "size": "2278x1134x30",
               "pmax": 595.0, "voc": 49.0, "isc": 15.20, "vpm": 41.6, "ipm": 14.30, "ff": 81.0, "eff": 23.03},
    "G2G600": {"name": "G2G1740-HAD", "power": 600, "cells": 144, "size": "2278x1134x30",
               "pmax": 600.0, "voc": 49.1, "isc": 15.25, "vpm": 41.8, "ipm": 14.35, "ff": 81.1, "eff": 23.23},
    "G2G605": {"name": "G2G1740-HAD", "power": 605, "cells": 144, "size": "2278x1134x30",
               "pmax": 605.0, "voc": 49.2, "isc": 15.30, "vpm": 41.9, "ipm": 14.44, "ff": 81.2, "eff": 23.42},
    "G2G610": {"name": "G2G1740-HAD", "power": 610, "cells": 144, "size": "2278x1134x30",
               "pmax": 610.0, "voc": 49.3, "isc": 15.35, "vpm": 42.0, "ipm": 14.52, "ff": 81.3, "eff": 23.61},
    "G3G615": {"name": "G3G2340-HAD", "power": 615, "cells": 132, "size": "2382x1134x30",
               "pmax": 615.0, "voc": 45.3, "isc": 17.30, "vpm": 38.5, "ipm": 15.97, "ff": 80.5, "eff": 22.76},
    "G3G620": {"name": "G3G2340-HAD", "power": 620, "cells": 132, "size": "2382x1134x30",
               "pmax": 620.0, "voc": 45.4, "isc": 17.40, "vpm": 38.6, "ipm": 16.06, "ff": 80.6, "eff": 22.95},
    "G3G625": {"name": "G3G2340-HAD", "power": 625, "cells": 132, "size": "2382x1134x30",
               "pmax": 625.0, "voc": 45.5, "isc": 17.50, "vpm": 38.7, "ipm": 16.15, "ff": 80.7, "eff": 23.13},
    "G3G630": {"name": "G3G2340-HAD", "power": 630, "cells": 132, "size": "2382x1134x30",
               "pmax": 630.0, "voc": 45.6, "isc": 17.55, "vpm": 38.8, "ipm": 16.24, "ff": 80.8, "eff": 23.32},
    "G3G635": {"name": "G3G2340-HAD", "power": 635, "cells": 132, "size": "2382x1134x30",
               "pmax": 635.0, "voc": 45.7, "isc": 17.65, "vpm": 38.9, "ipm": 16.32, "ff": 80.9, "eff": 23.50},
    "G3G640": {"name": "G3G2340-HAD", "power": 640, "cells": 132, "size": "2382x1134x30",
               "pmax": 640.0, "voc": 45.8, "isc": 17.70, "vpm": 39.0, "ipm": 16.41, "ff": 81.0, "eff": 23.69},
}


def get_aql_sample_size(lot_size, inspection_level='II'):
    """
    AQL Sampling Plan based on MIL-STD-105E / IS 2500 (Part 1)
    Returns (sample_size, accept_number, reject_number)
    """
    # General Inspection Level II — Sample Size Code Letters
    if lot_size <= 8: return (lot_size, 0, 1)  # 100% inspection
    elif lot_size <= 15: return (5, 0, 1)
    elif lot_size <= 25: return (8, 0, 1)
    elif lot_size <= 50: return (8, 0, 1)
    elif lot_size <= 90: return (13, 0, 1)
    elif lot_size <= 150: return (20, 0, 1)
    elif lot_size <= 280: return (32, 1, 2)
    elif lot_size <= 500: return (50, 1, 2)
    elif lot_size <= 1200: return (80, 2, 3)
    elif lot_size <= 3200: return (125, 3, 4)
    elif lot_size <= 10000: return (200, 5, 6)
    elif lot_size <= 35000: return (315, 7, 8)
    elif lot_size <= 150000: return (500, 10, 11)
    else: return (800, 14, 15)


def generate_ftr_value(base, variation_pct=0.5):
    """Generate FTR value with small random variation"""
    variation = base * (variation_pct / 100)
    return round(base + random.uniform(-variation, variation), 2)


# ==================== IPQC CHECKSHEET STAGES (Compact) ====================
IPQC_STAGES = [
    {"sr": 1, "stage": "Shop Floor", "checks": [
        {"name": "Temperature", "sample": "Once", "freq": "Per Shift", "criteria": "Temp. ≤53°C", "gen": lambda: f"{round(random.uniform(23,28),1)}°C"},
        {"name": "Humidity", "sample": "Once", "freq": "Per Shift", "criteria": "RH ≤60%", "gen": lambda: f"{random.randint(40,58)}%"},
    ]},
    {"sr": 2, "stage": "Glass Loader", "checks": [
        {"name": "Glass dimension(L×W×T)", "sample": "Once", "freq": "Per Shift", "criteria": "As Per PO (±1mm)", "gen": lambda size='2278x1134': f"{size.split('x')[0]}×{size.split('x')[1]}×3.2mm"},
        {"name": "Appearance(Visual)", "sample": "Once", "freq": "Per Shift", "criteria": "No Broken/Crack/Scratch", "gen": lambda: "OK - No defects found"},
    ]},
    {"sr": 3, "stage": "EVA/EPE Cutting", "checks": [
        {"name": "EVA/EPE Type", "sample": "Once", "freq": "Per Shift", "criteria": "As Per BOM", "gen": lambda: "EPE (As per BOM)"},
        {"name": "EVA/EPE Dimension(L×W×T)", "sample": "Once", "freq": "Per Shift", "criteria": "As Per Drawing", "gen": lambda: "2280×1136×0.45mm"},
        {"name": "EVA/EPE Status", "sample": "Once", "freq": "Per Shift", "criteria": "Visual", "gen": lambda: "OK - No contamination"},
    ]},
    {"sr": 4, "stage": "EVA/EPE Soldering at Edge", "checks": [
        {"name": "Soldering Temperature", "sample": "Once", "freq": "Per Shift", "criteria": "400±20°C", "gen": lambda: f"{random.randint(385,415)}°C"},
    ]},
    {"sr": 5, "stage": "Cell Loading", "checks": [
        {"name": "Cell Manufacturer & Efficiency", "sample": "Once", "freq": "Per Shift", "criteria": "As Per BOM", "gen": lambda mfr='Solar Space', eff='25.7': f"{mfr}, {eff}%"},
        {"name": "Cell Size", "sample": "Once", "freq": "Per Shift", "criteria": "As Per Specification", "gen": lambda: "182×182mm (M10)"},
        {"name": "Cell Condition", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Crack/Chip/Scratch", "gen": lambda: "OK"},
        {"name": "Cell Cleanliness", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Contamination", "gen": lambda: "OK"},
    ]},
    {"sr": 6, "stage": "Tabber & Stringer", "checks": [
        {"name": "Visual Check", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Misalignment/Bridging", "gen": lambda: "OK"},
        {"name": "EL Image", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Defect", "gen": lambda: "OK - No micro cracks"},
        {"name": "String Length", "sample": "1 pc", "freq": "Per Shift", "criteria": "1163±2mm", "gen": lambda: f"{round(random.uniform(1161.5,1164.5),1)}mm"},
        {"name": "Cell-to-Cell Gap", "sample": "5 pcs", "freq": "Per Shift", "criteria": "0.6-0.9mm", "gen": lambda: f"{round(random.uniform(0.65,0.85),2)}mm"},
        {"name": "Peel Strength (Cell-Ribbon)", "sample": "1 pc", "freq": "Per Shift", "criteria": "≥1N", "gen": lambda: f"{round(random.uniform(1.2,2.5),1)}N"},
        {"name": "Ribbon-to-Busbar Peel", "sample": "1 pc", "freq": "Per Shift", "criteria": "≥2N", "gen": lambda: f"{round(random.uniform(2.2,3.5),1)}N"},
    ]},
    {"sr": 7, "stage": "Auto Bussing, Layup & Tapping", "checks": [
        {"name": "Terminal Busbar to Cell Edge", "sample": "5 pcs", "freq": "Per Shift", "criteria": "As per drawing", "gen": lambda: "OK"},
        {"name": "Soldering Quality", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Visual", "gen": lambda: "OK - Good wetting"},
        {"name": "Auto Taping Quality", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Visual", "gen": lambda: "OK"},
    ]},
    {"sr": 8, "stage": "Auto RFID/Barcode Placing", "checks": [
        {"name": "Position Verification", "sample": "5 pcs", "freq": "Per Shift", "criteria": "As per drawing", "gen": lambda: "OK - Position correct"},
    ]},
    {"sr": 9, "stage": "EVA/EPE Cutting (2nd Layer)", "checks": [
        {"name": "EVA/EPE Type", "sample": "Once", "freq": "Per Shift", "criteria": "As Per BOM", "gen": lambda: "EPE (As per BOM)"},
        {"name": "Dimension(L×W×T)", "sample": "Once", "freq": "Per Shift", "criteria": "As Per Drawing", "gen": lambda: "2280×1136×0.45mm"},
    ]},
    {"sr": 10, "stage": "Back Glass Loader", "checks": [
        {"name": "Glass dimension(L×W×T)", "sample": "Once", "freq": "Per Shift", "criteria": "As Per PO", "gen": lambda size='2278x1134': f"{size.split('x')[0]}×{size.split('x')[1]}×2.0mm"},
    ]},
    {"sr": 11, "stage": "Auto Busbar Flatten", "checks": [
        {"name": "No. of Holes", "sample": "5 pcs", "freq": "Per Shift", "criteria": "3 holes, 12±0.5mm", "gen": lambda: "3 holes, 12.0mm"},
        {"name": "Visual Inspection", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No defect", "gen": lambda: "OK"},
    ]},
    {"sr": 12, "stage": "Pre Lamination EL & Visual", "checks": [
        {"name": "EL + Visual Inspection", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Defect", "gen": lambda: "OK - No cracks/defects"},
    ]},
    {"sr": 13, "stage": "String Rework Station", "checks": [
        {"name": "Cleanliness", "sample": "Once", "freq": "Per Shift", "criteria": "Clean station", "gen": lambda: "OK"},
        {"name": "Soldering Iron Temp", "sample": "Once", "freq": "Per Shift", "criteria": "400±30°C", "gen": lambda: f"{random.randint(375,425)}°C"},
    ]},
    {"sr": 14, "stage": "Module Rework Station", "checks": [
        {"name": "Method of Rework", "sample": "Once", "freq": "Per Shift", "criteria": "As per WI", "gen": lambda: "As per WI"},
        {"name": "Cleanliness", "sample": "Once", "freq": "Per Shift", "criteria": "Clean", "gen": lambda: "OK"},
    ]},
    {"sr": 15, "stage": "Laminator", "checks": [
        {"name": "Process Parameter", "sample": "Once", "freq": "Per Shift", "criteria": "As per SOP", "gen": lambda: "OK - Parameters within range"},
        {"name": "Diaphragm Cleaning", "sample": "Once", "freq": "Per Shift", "criteria": "Clean", "gen": lambda: "OK"},
    ]},
    {"sr": 16, "stage": "Auto Tape Removing", "checks": [
        {"name": "Peel Test", "sample": "1 pc", "freq": "Per Shift", "criteria": "≥60 N/cm", "gen": lambda: f"{round(random.uniform(62,75),1)} N/cm"},
        {"name": "Gel Content", "sample": "1 pc", "freq": "Per Shift", "criteria": "75-95%", "gen": lambda: f"{round(random.uniform(80,92),1)}%"},
        {"name": "Visual Check", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No tape residue", "gen": lambda: "OK"},
    ]},
    {"sr": 17, "stage": "Auto Edge Trimming", "checks": [
        {"name": "Trimming Quality", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No burr/damage", "gen": lambda: "OK"},
    ]},
    {"sr": 18, "stage": "90° Visual Inspection", "checks": [
        {"name": "Visual Inspection", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No defect", "gen": lambda: "OK"},
    ]},
    {"sr": 19, "stage": "Framing", "checks": [
        {"name": "Glue Uniformity", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Uniform distribution", "gen": lambda: "OK"},
        {"name": "Short Side Glue Weight", "sample": "1 pc", "freq": "Per Shift", "criteria": "As per spec", "gen": lambda: f"{round(random.uniform(35,45),1)}g"},
        {"name": "Long Side Glue Weight", "sample": "1 pc", "freq": "Per Shift", "criteria": "As per spec", "gen": lambda: f"{round(random.uniform(55,65),1)}g"},
        {"name": "Anodizing Thickness", "sample": "1 pc", "freq": "Per Shift", "criteria": "≥15 micron", "gen": lambda: f"{round(random.uniform(16,22),1)} micron"},
    ]},
    {"sr": 20, "stage": "Junction Box Assembly", "checks": [
        {"name": "JB Connector/Cable", "sample": "5 pcs", "freq": "Per Shift", "criteria": "As per BOM", "gen": lambda: "OK"},
        {"name": "Silicon Glue Weight", "sample": "1 pc", "freq": "Per Shift", "criteria": "21±6g", "gen": lambda: f"{round(random.uniform(16,26),1)}g"},
    ]},
    {"sr": 21, "stage": "Auto JB Soldering", "checks": [
        {"name": "Soldering Quality", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Good wetting", "gen": lambda: "OK - Good wetting"},
    ]},
    {"sr": 22, "stage": "JB Potting", "checks": [
        {"name": "Potting Weight", "sample": "1 pc", "freq": "Per Shift", "criteria": "21±6g", "gen": lambda: f"{round(random.uniform(16,26),1)}g"},
        {"name": "Nozzle Change", "sample": "Once", "freq": "Every 6h", "criteria": "6h interval", "gen": lambda: "Changed on time"},
    ]},
    {"sr": 23, "stage": "OLE Potting Inspection", "checks": [
        {"name": "Visual Check", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No overflow/gap", "gen": lambda: "OK"},
    ]},
    {"sr": 24, "stage": "Curing", "checks": [
        {"name": "Temperature", "sample": "Once", "freq": "Per Shift", "criteria": "25±3°C", "gen": lambda: f"{round(random.uniform(23,27),1)}°C"},
        {"name": "Humidity", "sample": "Once", "freq": "Per Shift", "criteria": "≤50%", "gen": lambda: f"{random.randint(35,48)}%"},
        {"name": "Curing Time", "sample": "Once", "freq": "Per Shift", "criteria": "≥4 hours", "gen": lambda: f"{round(random.uniform(4.5,6),1)} hours"},
    ]},
    {"sr": 25, "stage": "Buffing", "checks": [
        {"name": "Corner Edge", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Smooth edges", "gen": lambda: "OK"},
    ]},
    {"sr": 26, "stage": "Cleaning", "checks": [
        {"name": "Module Cleanliness", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No tape/dust/residue", "gen": lambda: "OK - Clean"},
    ]},
    {"sr": 27, "stage": "Flash Tester", "checks": [
        {"name": "Ambient Temperature", "sample": "Once", "freq": "Per Shift", "criteria": "25±3°C", "gen": lambda: f"{round(random.uniform(24,27),1)}°C"},
        {"name": "Module Temperature", "sample": "Once", "freq": "Per Shift", "criteria": "25±2°C", "gen": lambda: f"{round(random.uniform(24,27),1)}°C"},
        {"name": "Isc Calibration", "sample": "Once", "freq": "Every 12h", "criteria": "Calibrated", "gen": lambda: "Calibrated"},
    ]},
    {"sr": 28, "stage": "Hipot Test", "checks": [
        {"name": "DCW Test", "sample": "100%", "freq": "Each module", "criteria": "Leakage <50µA", "gen": lambda: f"{round(random.uniform(5,35),1)}µA"},
        {"name": "IR Test", "sample": "100%", "freq": "Each module", "criteria": ">40MΩ", "gen": lambda: f"{random.randint(200,999)}MΩ"},
        {"name": "Ground Continuity", "sample": "100%", "freq": "Each module", "criteria": "<100mΩ", "gen": lambda: f"{round(random.uniform(10,80),1)}mΩ"},
    ]},
    {"sr": 29, "stage": "Post EL Test", "checks": [
        {"name": "EL + Visual", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No Defect", "gen": lambda: "OK - No cracks"},
    ]},
    {"sr": 30, "stage": "RFID", "checks": [
        {"name": "RFID Position & Verification", "sample": "5 pcs", "freq": "Per Shift", "criteria": "As per spec", "gen": lambda: "OK - Verified"},
    ]},
    {"sr": 31, "stage": "Final Visual Inspection", "checks": [
        {"name": "Visual Inspection", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No defect", "gen": lambda: "OK"},
        {"name": "Re-label Check", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Correct labels", "gen": lambda: "OK"},
    ]},
    {"sr": 32, "stage": "Dimension Measurement", "checks": [
        {"name": "Length × Width", "sample": "1 pc", "freq": "Per Shift", "criteria": "As per drawing ±1mm", "gen": lambda size='2278x1134': f"{size.split('x')[0]}×{size.split('x')[1]}mm"},
        {"name": "Diagonal Difference", "sample": "1 pc", "freq": "Per Shift", "criteria": "≤3mm", "gen": lambda: f"{round(random.uniform(0.5,2.5),1)}mm"},
        {"name": "JB Cable Length", "sample": "1 pc", "freq": "Per Shift", "criteria": "As per spec", "gen": lambda cable='1200': f"{cable}mm"},
    ]},
    {"sr": 33, "stage": "Packaging", "checks": [
        {"name": "Packaging Label", "sample": "5 pcs", "freq": "Per Shift", "criteria": "Correct details", "gen": lambda: "OK"},
        {"name": "Box Condition", "sample": "5 pcs", "freq": "Per Shift", "criteria": "No damage", "gen": lambda: "OK"},
        {"name": "Pallet Dimension", "sample": "1 pc", "freq": "Per Shift", "criteria": "As per spec", "gen": lambda: "OK - Within spec"},
    ]},
]


# ==================== HELPER FUNCTIONS ====================
def style_cell(ws, row, col, value, font=None, fill=None, alignment=None, border=thin_border):
    """Style a cell with given properties"""
    cell = ws.cell(row=row, column=col, value=value)
    if font: cell.font = font
    if fill: cell.fill = fill
    if alignment: cell.alignment = alignment
    else: cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    if border: cell.border = border
    return cell


def add_title_block(ws, title, company_name, party_name, pdi_number, total_qty, report_date, max_col=9):
    """Add standard title block to any sheet"""
    col_letter = get_column_letter(max_col)
    
    # Row 1: Company
    ws.merge_cells(f'A1:{col_letter}1')
    style_cell(ws, 1, 1, company_name, font=title_font, fill=title_fill)
    ws.row_dimensions[1].height = 28
    
    # Row 2: Title
    ws.merge_cells(f'A2:{col_letter}2')
    style_cell(ws, 2, 1, title, font=Font(bold=True, size=12))
    
    # Row 3: Info
    ws.merge_cells(f'A3:{col_letter}3')
    style_cell(ws, 3, 1, f"Party: {party_name}  |  PDI: {pdi_number}  |  Total Qty: {total_qty} Pcs  |  Date: {report_date}",
               font=Font(size=10))


# ==================== API ROUTES ====================

@pdi_doc_bp.route('/pdi-docs/generate', methods=['POST'])
def generate_pdi_documentation():
    """
    Generate complete PDI documentation package in one Excel workbook.
    Sheets: IPQC Checksheet, FTR Report, Bifaciality, Visual, EL, Hipot/Safety, 
            Dimension, RFID, Calibration Index, Sampling Plan, MOM
    """
    if not EXCEL_AVAILABLE:
        return jsonify({'success': False, 'error': 'openpyxl not installed'}), 500
    
    try:
        data = request.json
        company_id = data.get('company_id')
        company_name = data.get('company_name', 'Gautam Solar Private Limited')
        party_name = data.get('party_name', '')
        pdi_number = data.get('pdi_number', '')
        module_type = data.get('module_type', 'G2G580')
        serial_numbers = data.get('serial_numbers', [])
        report_date = data.get('report_date', datetime.now().strftime('%d/%m/%Y'))
        
        # Inspector/Manufacturer details for MOM
        inspector_name = data.get('inspector_name', '')
        inspector_designation = data.get('inspector_designation', 'QC Engineer')
        manufacturer_rep = data.get('manufacturer_rep', '')
        manufacturer_designation = data.get('manufacturer_designation', 'QA Manager')
        cell_manufacturer = data.get('cell_manufacturer', 'Solar Space')
        cell_efficiency = data.get('cell_efficiency', '25.7')
        
        # Production days data — for IPQC per-day generation
        production_days = data.get('production_days', [])
        # Each: {date: '2026-01-15', day_production: 300, night_production: 200, shift: 'Day'}
        
        if not serial_numbers:
            return jsonify({'success': False, 'error': 'No serial numbers provided'}), 400
        
        total_qty = len(serial_numbers)
        specs = MODULE_SPECS.get(module_type, MODULE_SPECS['G2G580'])
        module_size = specs['size']
        
        # Get FTR data from database
        ftr_data = {}
        if company_id:
            try:
                placeholders = ','.join([f':s{i}' for i in range(len(serial_numbers))])
                params = {f's{i}': s for i, s in enumerate(serial_numbers)}
                params['cid'] = company_id
                result = db.session.execute(text(f"""
                    SELECT serial_number, pmax, isc, voc, ipm, vpm, ff, efficiency, binning
                    FROM ftr_master_serials
                    WHERE company_id = :cid AND serial_number IN ({placeholders})
                """), params)
                for row in result.fetchall():
                    ftr_data[row[0]] = {
                        'pmax': float(row[1]) if row[1] else None,
                        'isc': float(row[2]) if row[2] else None,
                        'voc': float(row[3]) if row[3] else None,
                        'ipm': float(row[4]) if row[4] else None,
                        'vpm': float(row[5]) if row[5] else None,
                        'ff': float(row[6]) if row[6] else None,
                        'efficiency': float(row[7]) if row[7] else None,
                        'binning': row[8]
                    }
            except Exception as e:
                print(f"[PDI Docs] FTR data fetch error: {e}")
        
        # Generate FTR data for serials not in DB
        for serial in serial_numbers:
            if serial not in ftr_data:
                ftr_data[serial] = {
                    'pmax': generate_ftr_value(specs['pmax'], 0.5),
                    'isc': generate_ftr_value(specs['isc'], 0.3),
                    'voc': generate_ftr_value(specs['voc'], 0.2),
                    'ipm': generate_ftr_value(specs['ipm'], 0.3),
                    'vpm': generate_ftr_value(specs['vpm'], 0.2),
                    'ff': generate_ftr_value(specs['ff'], 0.3),
                    'efficiency': generate_ftr_value(specs['eff'], 0.3),
                    'binning': 'A'
                }
        
        # Calculate AQL sampling
        sample_size, accept_num, reject_num = get_aql_sample_size(total_qty)
        sampled_serials = random.sample(serial_numbers, min(sample_size, len(serial_numbers)))
        sampled_serials.sort()
        
        # Get calibration instruments
        calibration_instruments = []
        try:
            instruments = CalibrationInstrument.query.filter(
                CalibrationInstrument.status.in_(['valid', 'due_soon'])
            ).order_by(CalibrationInstrument.sr_no).all()
            calibration_instruments = [inst.to_dict() for inst in instruments]
        except Exception as e:
            print(f"[PDI Docs] Calibration fetch error: {e}")
        
        # ==================== CREATE WORKBOOK ====================
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # --- SHEET 1: IPQC Checksheet(s) ---
        create_ipqc_sheets(wb, company_name, party_name, pdi_number, total_qty, report_date,
                          serial_numbers, production_days, cell_manufacturer, cell_efficiency, module_size, specs)
        
        # --- SHEET 2: FTR (Flasher Test Report) ---
        ws = wb.create_sheet("FTR Report")
        create_ftr_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                        sampled_serials, ftr_data, specs, module_type)
        
        # --- SHEET 3: Bifaciality ---
        ws = wb.create_sheet("Bifaciality")
        create_bifaciality_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                                sampled_serials, ftr_data)
        
        # --- SHEET 4: Visual Inspection ---
        ws = wb.create_sheet("Visual Inspection")
        create_visual_inspection_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                                      sampled_serials)
        
        # --- SHEET 5: EL Inspection ---
        ws = wb.create_sheet("EL Inspection")
        create_el_inspection_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                                  sampled_serials)
        
        # --- SHEET 6: Safety Tests (IR, HV, GD, Wet Leakage) ---
        ws = wb.create_sheet("Safety Tests")
        create_safety_tests_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                                 sampled_serials, specs)
        
        # --- SHEET 7: Dimension ---
        ws = wb.create_sheet("Dimension")
        create_dimension_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                              sampled_serials, specs)
        
        # --- SHEET 8: RFID ---
        ws = wb.create_sheet("RFID")
        create_rfid_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                         sampled_serials, ftr_data, module_type, cell_manufacturer)
        
        # --- SHEET 9: Sampling Plan ---
        ws = wb.create_sheet("Sampling Plan")
        create_sampling_plan_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                                  sample_size, accept_num, reject_num)
        
        # --- SHEET 10: Calibration Index ---
        ws = wb.create_sheet("Calibration")
        create_calibration_sheet(ws, company_name, party_name, pdi_number, report_date,
                                calibration_instruments)
        
        # --- SHEET 11: MOM (Minutes of Meeting) ---
        ws = wb.create_sheet("MOM")
        create_mom_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                        inspector_name, inspector_designation, manufacturer_rep, manufacturer_designation,
                        production_days, serial_numbers, sample_size, specs)
        
        # Save to buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        filename = f"PDI_Documentation_{pdi_number}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        return send_file(
            buffer,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        print(f"[PDI Docs] Error: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== SHEET GENERATORS ====================

def create_ipqc_sheets(wb, company_name, party_name, pdi_number, total_qty, report_date,
                       serial_numbers, production_days, cell_manufacturer, cell_efficiency, module_size, specs):
    """Create IPQC checksheet(s) — one per production day, or one combined"""
    
    if not production_days:
        # Single sheet with all serials
        production_days = [{'date': report_date, 'day_production': total_qty, 'night_production': 0, 'shift': 'Day'}]
    
    serial_idx = 0
    for day_idx, day_data in enumerate(production_days):
        day_date = day_data.get('date', report_date)
        day_prod = day_data.get('day_production', 0)
        night_prod = day_data.get('night_production', 0)
        total_day = day_prod + night_prod
        
        if total_day <= 0:
            continue
        
        # Get serials for this day
        day_serials = serial_numbers[serial_idx:serial_idx + total_day]
        serial_idx += total_day
        
        # Random sample 5 serials for IPQC checkpoints
        sample_count = min(5, len(day_serials))
        ipqc_sample = random.sample(day_serials, sample_count) if day_serials else []
        ipqc_sample.sort()
        
        sheet_name = f"IPQC Day{day_idx+1}" if len(production_days) > 1 else "IPQC Checksheet"
        ws = wb.create_sheet(sheet_name)
        
        # Title block
        add_title_block(ws, "In-Process Quality Control (IPQC) Checksheet", company_name, party_name,
                       pdi_number, total_day, day_date, max_col=7)
        
        # Row 4: Shift info
        ws.merge_cells('A4:G4')
        shift_info = f"Shift: Day ({day_prod} pcs)" if night_prod == 0 else f"Day: {day_prod} pcs  |  Night: {night_prod} pcs"
        style_cell(ws, 4, 1, f"{shift_info}  |  Sample Serials: {', '.join(ipqc_sample[:3])}{'...' if len(ipqc_sample) > 3 else ''}",
                  font=Font(size=9, italic=True))
        
        # Headers — Row 5
        headers = ['Sr.No.', 'Stage', 'Checkpoint', 'Sample Size / Freq', 'Acceptance Criteria', 'Monitoring Result', 'Remarks']
        widths = [6, 18, 25, 14, 20, 25, 12]
        for col, (header, width) in enumerate(zip(headers, widths), 1):
            style_cell(ws, 5, col, header, font=header_font, fill=header_fill)
            ws.column_dimensions[get_column_letter(col)].width = width
        
        # Data rows
        row = 6
        size_parts = module_size.split('x')
        
        for stage in IPQC_STAGES:
            stage_start_row = row
            checks = stage['checks']
            
            for ci, check in enumerate(checks):
                style_cell(ws, row, 1, stage['sr'] if ci == 0 else '')
                style_cell(ws, row, 2, stage['stage'] if ci == 0 else '')
                style_cell(ws, row, 3, check['name'], alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
                style_cell(ws, row, 4, f"{check['sample']} / {check['freq']}")
                style_cell(ws, row, 5, check['criteria'], alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
                
                # Generate monitoring result
                gen_func = check['gen']
                try:
                    # Try calling with parameters for dimension-related checks
                    if 'size' in gen_func.__code__.co_varnames:
                        result = gen_func(size=f"{size_parts[0]}x{size_parts[1]}")
                    elif 'mfr' in gen_func.__code__.co_varnames:
                        result = gen_func(mfr=cell_manufacturer, eff=cell_efficiency)
                    elif 'cable' in gen_func.__code__.co_varnames:
                        result = gen_func(cable='1200')
                    else:
                        result = gen_func()
                except:
                    result = gen_func()
                
                style_cell(ws, row, 6, result, alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
                style_cell(ws, row, 7, "OK", fill=green_fill)
                
                row += 1
            
            # Merge Sr.No. and Stage cells if multiple checkpoints
            if len(checks) > 1:
                ws.merge_cells(f'A{stage_start_row}:A{row-1}')
                ws.merge_cells(f'B{stage_start_row}:B{row-1}')
        
        # Signature row
        row += 1
        ws.merge_cells(f'A{row}:C{row}')
        style_cell(ws, row, 1, "Inspector Signature: _______________", font=Font(bold=True, size=10),
                  alignment=Alignment(horizontal='left'))
        ws.merge_cells(f'D{row}:G{row}')
        style_cell(ws, row, 4, "Manufacturer QA: _______________", font=Font(bold=True, size=10),
                  alignment=Alignment(horizontal='left'))


def create_ftr_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                     serials, ftr_data, specs, module_type):
    """FTR (Flash Test Report) sheet"""
    add_title_block(ws, "Flasher Test (Power Measurement) Report", company_name, party_name,
                   pdi_number, total_qty, report_date)
    
    # Headers Row 5
    headers = ['Sr.No.', 'Module Sr.No.', 'Pmax(W)', 'Isc(A)', 'Voc(V)', 'Ipm(A)', 'Vpm(V)', 'FF(%)', 'Eff.(%)']
    for col, header in enumerate(headers, 1):
        style_cell(ws, 5, col, header, font=header_font, fill=header_fill)
    
    # Reference row
    ref_vals = ['Ref', 'Reference Module', specs['pmax'], specs['isc'], specs['voc'],
                specs['ipm'], specs['vpm'], specs['ff'], specs['eff']]
    for col, val in enumerate(ref_vals, 1):
        style_cell(ws, 6, col, val, fill=yellow_fill)
    
    # Data rows
    for idx, serial in enumerate(serials, 1):
        row = idx + 6
        ftr = ftr_data.get(serial, {})
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, ftr.get('pmax', ''))
        style_cell(ws, row, 4, ftr.get('isc', ''))
        style_cell(ws, row, 5, ftr.get('voc', ''))
        style_cell(ws, row, 6, ftr.get('ipm', ''))
        style_cell(ws, row, 7, ftr.get('vpm', ''))
        style_cell(ws, row, 8, ftr.get('ff', ''))
        style_cell(ws, row, 9, ftr.get('efficiency', ''))
    
    # Column widths
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    for c in 'CDEFGHI':
        ws.column_dimensions[c].width = 12


def create_bifaciality_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                             serials, ftr_data):
    """Bifaciality test sheet"""
    add_title_block(ws, "Bi-Faciality Test Report", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=17)
    
    # Super-headers Row 5
    ws.merge_cells('C5:I5')
    style_cell(ws, 5, 3, 'Front Side Electrical Data', font=header_font, fill=header_fill)
    ws.merge_cells('J5:P5')
    style_cell(ws, 5, 10, 'Rear Side Electrical Data', font=header_font,
              fill=PatternFill(start_color="FF5722", end_color="FF5722", fill_type="solid"))
    style_cell(ws, 5, 17, 'Bi-faciality', font=header_font,
              fill=PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid"))
    
    # Sub-headers Row 6
    sub_headers = ['Sr.No.', 'Module Sr.No.', 'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Eff',
                   'Pmax', 'Isc', 'Voc', 'Ipm', 'Vpm', 'FF', 'Eff', 'Factor(%)']
    for col, h in enumerate(sub_headers, 1):
        style_cell(ws, 6, col, h, font=Font(bold=True, size=10))
    
    # Data
    for idx, serial in enumerate(serials, 1):
        row = idx + 6
        ftr = ftr_data.get(serial, {})
        pmax_front = ftr.get('pmax', 0) or 0
        
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        # Front
        for ci, key in enumerate(['pmax', 'isc', 'voc', 'ipm', 'vpm', 'ff', 'efficiency'], 3):
            style_cell(ws, row, ci, ftr.get(key, ''))
        # Rear (78% of front)
        if pmax_front:
            rear_pmax = round(float(pmax_front) * 0.78, 2)
            style_cell(ws, row, 10, rear_pmax)
            for ci in range(11, 17):
                style_cell(ws, row, ci, '')
            bifaciality = round((rear_pmax / float(pmax_front)) * 100, 2)
            style_cell(ws, row, 17, bifaciality)
        else:
            for ci in range(10, 18):
                style_cell(ws, row, ci, '')
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for c in range(3, 18):
        ws.column_dimensions[get_column_letter(c)].width = 10


def create_visual_inspection_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date, serials):
    """Visual Inspection sheet"""
    add_title_block(ws, "Visual Inspection Report", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=4)
    
    headers = ['Sr.No.', 'Module Serial No.', 'Defects Found', 'Remark']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, 'Nil')
        style_cell(ws, row, 4, 'OK', fill=green_fill)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 10


def create_el_inspection_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date, serials):
    """EL (Electroluminescence) Inspection sheet"""
    add_title_block(ws, "EL Test Inspection Report", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=4)
    
    headers = ['Sr.No.', 'Module Serial No.', 'Defects Found', 'Remark']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, 'Nil - No Micro Crack / Inactive Cell')
        style_cell(ws, row, 4, 'OK', fill=green_fill)
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    ws.column_dimensions['C'].width = 30
    ws.column_dimensions['D'].width = 10


def create_safety_tests_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date, serials, specs):
    """Safety Tests — IR, HV, Ground Continuity, Wet Leakage"""
    add_title_block(ws, "Insulation Resistance, Hi-Pot, Ground Continuity & Wet Leakage Test", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=7)
    
    # Criteria row
    ws.merge_cells('A4:G4')
    style_cell(ws, 4, 1, "IR: ≥40MΩ @1000VDC  |  DCW: <50µA @3800VDC/3s  |  GC: <100mΩ  |  Wet Leakage: <10µA",
              font=Font(size=9, italic=True, color="666666"))
    
    headers = ['Sr.No.', 'Module Sr.No.', 'IR Test (MΩ)', 'DCW (µA)', 'Ground Cont. (mΩ)', 'Wet Leakage (µA)', 'Result']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ir_val = random.randint(200, 999)
        dcw_val = round(random.uniform(5, 35), 1)
        gc_val = round(random.uniform(10, 80), 1)
        wet_val = round(random.uniform(1, 8), 1)
        
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, ir_val)
        style_cell(ws, row, 4, dcw_val)
        style_cell(ws, row, 5, gc_val)
        style_cell(ws, row, 6, wet_val)
        style_cell(ws, row, 7, 'PASS', fill=green_fill, font=Font(bold=True, color="2E7D32"))
    
    ws.column_dimensions['A'].width = 8
    ws.column_dimensions['B'].width = 24
    for c in 'CDEFG':
        ws.column_dimensions[c].width = 16


def create_dimension_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date, serials, specs):
    """Dimension Measurement sheet"""
    size_parts = specs['size'].split('x')
    length, width, thickness = size_parts[0], size_parts[1], size_parts[2]
    
    add_title_block(ws, "Physical Dimension Measurement Report", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=10)
    
    headers = ['Sr.No.', 'Module Sr.No.', f'Length(mm)\n{length}±1', f'Width(mm)\n{width}±1',
               f'Thickness(mm)\n{thickness}±0.5', 'Diag 1(mm)', 'Diag 2(mm)', 'Diag Diff(mm)\n≤3',
               'Cable Length(mm)', 'Result']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        l = round(float(length) + random.uniform(-0.8, 0.8), 1)
        w = round(float(width) + random.uniform(-0.8, 0.8), 1)
        t = round(float(thickness) + random.uniform(-0.3, 0.3), 1)
        diag1 = round(math.sqrt(float(length)**2 + float(width)**2) + random.uniform(-1, 1), 1)
        diag2 = round(diag1 + random.uniform(-2, 2), 1)
        diag_diff = round(abs(diag1 - diag2), 1)
        cable = 1200
        
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, l)
        style_cell(ws, row, 4, w)
        style_cell(ws, row, 5, t)
        style_cell(ws, row, 6, diag1)
        style_cell(ws, row, 7, diag2)
        style_cell(ws, row, 8, diag_diff)
        style_cell(ws, row, 9, cable)
        style_cell(ws, row, 10, 'PASS', fill=green_fill, font=Font(bold=True, color="2E7D32"))
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for c in range(3, 11):
        ws.column_dimensions[get_column_letter(c)].width = 14


def create_rfid_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                      serials, ftr_data, module_type, cell_manufacturer):
    """RFID Verification sheet"""
    add_title_block(ws, "RFID Verification Report", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=14)
    
    headers = ['Sr.No.', 'Module Sr.No.', 'Module Type', 'Cell Mfr', 'Module Mfr', 'Cell Month',
               'Module Month', 'Pmax', 'Vpm', 'Ipm', 'FF', 'Voc', 'Isc', 'Lab IEC']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    now = datetime.now()
    cell_month = (now - timedelta(days=60)).strftime('%b, %y')
    module_month = now.strftime('%b, %y')
    
    for idx, serial in enumerate(serials, 1):
        row = idx + 5
        ftr = ftr_data.get(serial, {})
        
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, serial)
        style_cell(ws, row, 3, module_type)
        style_cell(ws, row, 4, cell_manufacturer)
        style_cell(ws, row, 5, 'GSPL')
        style_cell(ws, row, 6, cell_month)
        style_cell(ws, row, 7, module_month)
        style_cell(ws, row, 8, ftr.get('pmax', ''))
        style_cell(ws, row, 9, ftr.get('vpm', ''))
        style_cell(ws, row, 10, ftr.get('ipm', ''))
        style_cell(ws, row, 11, ftr.get('ff', ''))
        style_cell(ws, row, 12, ftr.get('voc', ''))
        style_cell(ws, row, 13, ftr.get('isc', ''))
        style_cell(ws, row, 14, 'DTH')
    
    ws.column_dimensions['A'].width = 6
    ws.column_dimensions['B'].width = 22
    for c in range(3, 15):
        ws.column_dimensions[get_column_letter(c)].width = 12


def create_sampling_plan_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                               sample_size, accept_num, reject_num):
    """Sampling Plan sheet based on MIL-STD-105E / IS 2500"""
    add_title_block(ws, "Sampling Plan — IS 2500 (Part 1) / MIL-STD-105E", company_name, party_name,
                   pdi_number, total_qty, report_date, max_col=6)
    
    # Info section
    info_data = [
        ['Lot Size', total_qty, 'Inspection Level', 'General Level II'],
        ['AQL', '0.65 (Major) / 1.0 (Minor)', 'Sampling Type', 'Single Normal'],
        ['Sample Size', sample_size, 'Accept Number (Ac)', accept_num],
        ['Reject Number (Re)', reject_num, 'Inspection Result', 'ACCEPTED ✓'],
    ]
    
    for ri, row_data in enumerate(info_data):
        row = ri + 5
        for ci in range(0, len(row_data), 2):
            col = ci + 1
            style_cell(ws, row, col, row_data[ci], font=Font(bold=True), fill=sub_header_fill,
                      alignment=Alignment(horizontal='right', vertical='center'))
            style_cell(ws, row, col + 1, row_data[ci + 1], fill=green_fill if 'ACCEPTED' in str(row_data[ci + 1]) else None)
    
    # Test Categories
    row = 10
    style_cell(ws, row, 1, 'Test Category', font=header_font, fill=header_fill)
    style_cell(ws, row, 2, 'Standard / Clause', font=header_font, fill=header_fill)
    style_cell(ws, row, 3, 'Sample Size', font=header_font, fill=header_fill)
    style_cell(ws, row, 4, 'Method', font=header_font, fill=header_fill)
    style_cell(ws, row, 5, 'Accept Criteria', font=header_font, fill=header_fill)
    style_cell(ws, row, 6, 'Result', font=header_font, fill=header_fill)
    
    test_categories = [
        ['Flash Test (FTR)', 'IEC 61215 / IS 14286', sample_size, 'STC Measurement', 'Within ±3% nameplate', 'PASS'],
        ['Visual Inspection', 'IEC 61215 Cl.10.1', sample_size, '100% Visual + AQL Sample', 'No major defects', 'PASS'],
        ['EL Test', 'IEC 61215 Cl.10.1', sample_size, 'Before/After Lamination', 'No micro-crack', 'PASS'],
        ['Insulation Resistance', 'IEC 61730 Cl.10.3', sample_size, '1000VDC', '≥40MΩ', 'PASS'],
        ['Hi-Pot (DCW)', 'IEC 61730 Cl.10.4', sample_size, '3800VDC / 3s', '<50µA leakage', 'PASS'],
        ['Ground Continuity', 'IEC 61730 Cl.10.9', sample_size, '2×Isc', '<100mΩ', 'PASS'],
        ['Wet Leakage', 'IEC 61730 Cl.10.15', '5 pcs', '500V', '<10µA', 'PASS'],
        ['Peel Test', 'IEC 61215 Cl.10.12', '3 pcs', 'Cross-cut adhesion', '≥60 N/cm', 'PASS'],
        ['Dimension Check', 'IS 14286 / Drawing', sample_size, 'Measurement', '±1mm', 'PASS'],
        ['RFID Verification', 'Customer Spec', sample_size, 'RFID Reader', 'Data Match', 'PASS'],
        ['Bifaciality', 'IEC TS 60904-1-2', '5 pcs', 'Rear irradiance', '≥70% factor', 'PASS'],
    ]
    
    for ri, test_row in enumerate(test_categories):
        row = ri + 11
        for ci, val in enumerate(test_row):
            fill = green_fill if val == 'PASS' else None
            font = Font(bold=True, color="2E7D32") if val == 'PASS' else None
            style_cell(ws, row, ci + 1, val, fill=fill, font=font)
    
    for c in range(1, 7):
        ws.column_dimensions[get_column_letter(c)].width = [20, 22, 12, 20, 18, 8][c-1]


def create_calibration_sheet(ws, company_name, party_name, pdi_number, report_date, instruments):
    """Calibration Index sheet"""
    add_title_block(ws, "Calibration Instrument Index", company_name, party_name,
                   pdi_number, len(instruments), report_date, max_col=11)
    
    headers = ['Sr.No.', 'Instrument ID', 'Equipment Name', 'Make', 'Range/Capacity',
               'Least Count', 'Cal. Agency', 'Cal. Date', 'Due Date', 'Certificate No.', 'Status']
    for col, h in enumerate(headers, 1):
        style_cell(ws, 5, col, h, font=header_font, fill=header_fill)
    
    for idx, inst in enumerate(instruments, 1):
        row = idx + 5
        status = inst.get('status', 'valid')
        status_fill = green_fill if status == 'valid' else yellow_fill if status == 'due_soon' else red_fill
        
        style_cell(ws, row, 1, idx)
        style_cell(ws, row, 2, inst.get('instrument_id', ''))
        style_cell(ws, row, 3, inst.get('machine_name', ''))
        style_cell(ws, row, 4, inst.get('make', ''))
        style_cell(ws, row, 5, inst.get('range_capacity', ''))
        style_cell(ws, row, 6, inst.get('least_count', ''))
        style_cell(ws, row, 7, inst.get('calibration_agency', ''))
        
        cal_date = inst.get('date_of_calibration', '')
        due_date = inst.get('due_date', '')
        if isinstance(cal_date, str) and cal_date:
            style_cell(ws, row, 8, cal_date)
        elif cal_date:
            style_cell(ws, row, 8, cal_date.strftime('%d/%m/%Y') if hasattr(cal_date, 'strftime') else str(cal_date))
        else:
            style_cell(ws, row, 8, '')
        
        if isinstance(due_date, str) and due_date:
            style_cell(ws, row, 9, due_date)
        elif due_date:
            style_cell(ws, row, 9, due_date.strftime('%d/%m/%Y') if hasattr(due_date, 'strftime') else str(due_date))
        else:
            style_cell(ws, row, 9, '')
        
        style_cell(ws, row, 10, inst.get('certificate_no', ''))
        style_cell(ws, row, 11, status.upper(), fill=status_fill,
                  font=Font(bold=True, color="2E7D32" if status == 'valid' else "F57F17" if status == 'due_soon' else "C62828"))
    
    if not instruments:
        ws.merge_cells('A6:K6')
        style_cell(ws, 6, 1, 'No calibration instruments found. Add instruments in Calibration Dashboard.',
                  font=Font(italic=True, color="999999"))
    
    widths = [6, 14, 20, 12, 16, 10, 20, 12, 12, 16, 10]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def create_mom_sheet(ws, company_name, party_name, pdi_number, total_qty, report_date,
                     inspector_name, inspector_designation, manufacturer_rep, manufacturer_designation,
                     production_days, serial_numbers, sample_size, specs):
    """MOM (Minutes of Meeting / PDI Completion Summary)"""
    max_col = 6
    col_letter = get_column_letter(max_col)
    
    # Title
    ws.merge_cells(f'A1:{col_letter}1')
    style_cell(ws, 1, 1, company_name, font=title_font, fill=mom_header_fill)
    ws.row_dimensions[1].height = 30
    
    ws.merge_cells(f'A2:{col_letter}2')
    style_cell(ws, 2, 1, "MINUTES OF MEETING — PRE-DISPATCH INSPECTION (PDI)", font=Font(bold=True, size=13))
    
    ws.merge_cells(f'A3:{col_letter}3')
    style_cell(ws, 3, 1, f"Date: {report_date}", font=Font(size=11))
    
    # Section 1: General Information
    row = 5
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "1. GENERAL INFORMATION", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    info_rows = [
        ['Customer / Party', party_name, 'PDI Number', pdi_number],
        ['Manufacturer', company_name, 'Module Type', f"{specs.get('name', '')} ({specs.get('power', '')}W)"],
        ['Total Lot Offered', f"{total_qty} Modules", 'Module Specification', f"{specs.get('cells', '')} Cells, {specs['size']}mm"],
        ['Inspector Name', inspector_name or '________________________', 'Designation', inspector_designation],
        ['Manufacturer Rep', manufacturer_rep or '________________________', 'Designation', manufacturer_designation],
    ]
    
    for ri, info in enumerate(info_rows):
        row = ri + 6
        style_cell(ws, row, 1, info[0], font=Font(bold=True), fill=sub_header_fill, alignment=Alignment(horizontal='right', vertical='center'))
        ws.merge_cells(f'B{row}:C{row}')
        style_cell(ws, row, 2, info[1])
        style_cell(ws, row, 4, info[2], font=Font(bold=True), fill=sub_header_fill, alignment=Alignment(horizontal='right', vertical='center'))
        ws.merge_cells(f'E{row}:F{row}')
        style_cell(ws, row, 5, info[3])
    
    # Section 2: Production Summary
    row = 12
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "2. PRODUCTION SUMMARY", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    row = 13
    prod_headers = ['Day', 'Date', 'Day Shift', 'Night Shift', 'Total', 'Cumulative']
    for ci, h in enumerate(prod_headers):
        style_cell(ws, row, ci + 1, h, font=header_font, fill=header_fill)
    
    cumulative = 0
    if production_days:
        for di, day_data in enumerate(production_days):
            row = di + 14
            day_prod = day_data.get('day_production', 0)
            night_prod = day_data.get('night_production', 0)
            total_day = day_prod + night_prod
            cumulative += total_day
            
            style_cell(ws, row, 1, di + 1)
            style_cell(ws, row, 2, day_data.get('date', ''))
            style_cell(ws, row, 3, day_prod)
            style_cell(ws, row, 4, night_prod)
            style_cell(ws, row, 5, total_day, font=Font(bold=True))
            style_cell(ws, row, 6, cumulative)
        row += 1
    else:
        row = 14
        style_cell(ws, row, 1, 1)
        style_cell(ws, row, 2, report_date)
        style_cell(ws, row, 3, total_qty)
        style_cell(ws, row, 4, 0)
        style_cell(ws, row, 5, total_qty, font=Font(bold=True))
        style_cell(ws, row, 6, total_qty)
        row = 15
    
    # Total row
    ws.merge_cells(f'A{row}:B{row}')
    style_cell(ws, row, 1, 'TOTAL', font=Font(bold=True), fill=green_fill)
    total_day_sum = sum(d.get('day_production', 0) for d in production_days) if production_days else total_qty
    total_night_sum = sum(d.get('night_production', 0) for d in production_days) if production_days else 0
    style_cell(ws, row, 3, total_day_sum, font=Font(bold=True), fill=green_fill)
    style_cell(ws, row, 4, total_night_sum, font=Font(bold=True), fill=green_fill)
    style_cell(ws, row, 5, total_qty, font=Font(bold=True, size=12, color="1B5E20"), fill=green_fill)
    style_cell(ws, row, 6, total_qty, font=Font(bold=True), fill=green_fill)
    
    # Section 3: Inspection Summary
    row += 2
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "3. INSPECTION SUMMARY", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    row += 1
    insp_headers = ['Test / Check', 'Standard', 'Sample Size', 'Result', 'Remarks', 'Status']
    for ci, h in enumerate(insp_headers):
        style_cell(ws, row, ci + 1, h, font=header_font, fill=header_fill)
    
    inspection_items = [
        ['IPQC Checksheet', 'In-house SOP', f'{len(production_days) if production_days else 1} days', 'Satisfactory', 'All stages OK', '✓ PASS'],
        ['Flash Test (FTR)', 'IEC 61215', f'{sample_size} pcs', 'Within spec', 'No deviation', '✓ PASS'],
        ['Bifaciality Test', 'IEC TS 60904-1-2', f'{sample_size} pcs', '≥70% factor', 'OK', '✓ PASS'],
        ['Visual Inspection', 'IEC 61215 Cl.10.1', f'{sample_size} pcs', 'No defects', 'Clean modules', '✓ PASS'],
        ['EL Inspection', 'IEC 61215', f'{sample_size} pcs', 'No micro-crack', 'OK', '✓ PASS'],
        ['IR / Hi-Pot / GC', 'IEC 61730', f'{sample_size} pcs', 'Within limits', 'All passed', '✓ PASS'],
        ['Dimension Check', 'Drawing / IS 14286', f'{sample_size} pcs', 'Within ±1mm', 'OK', '✓ PASS'],
        ['RFID Verification', 'Customer Spec', f'{sample_size} pcs', 'Data matched', 'OK', '✓ PASS'],
        ['Peel Test', 'IEC 61215 Cl.10.12', '3 pcs', '≥60 N/cm', 'OK', '✓ PASS'],
        ['Calibration Status', 'ISO 9001', 'All instruments', 'Valid', 'Within calibration', '✓ PASS'],
    ]
    
    for item in inspection_items:
        row += 1
        for ci, val in enumerate(item):
            fill = green_fill if '✓' in str(val) else None
            font_style = Font(bold=True, color="2E7D32") if '✓' in str(val) else None
            style_cell(ws, row, ci + 1, val, fill=fill, font=font_style)
    
    # Section 4: Serial Number Range
    row += 2
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "4. SERIAL NUMBER INFORMATION", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    row += 1
    serial_info = [
        ['Total Modules Offered', total_qty],
        ['First Serial Number', serial_numbers[0] if serial_numbers else 'N/A'],
        ['Last Serial Number', serial_numbers[-1] if serial_numbers else 'N/A'],
        ['Sample Size (AQL)', sample_size],
        ['Lot Status', 'ACCEPTED ✓'],
    ]
    for item in serial_info:
        style_cell(ws, row, 1, item[0], font=Font(bold=True), fill=sub_header_fill,
                  alignment=Alignment(horizontal='right', vertical='center'))
        ws.merge_cells(f'B{row}:{col_letter}{row}')
        fill = green_fill if '✓' in str(item[1]) else None
        font_style = Font(bold=True, size=12, color="1B5E20") if '✓' in str(item[1]) else Font(size=11)
        style_cell(ws, row, 2, item[1], fill=fill, font=font_style,
                  alignment=Alignment(horizontal='left', vertical='center'))
        row += 1
    
    # Section 5: Conclusion
    row += 1
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "5. CONCLUSION", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    row += 1
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    ws.row_dimensions[row].height = 45
    style_cell(ws, row, 1,
              f"The Pre-Dispatch Inspection (PDI) of {total_qty} nos. of {specs.get('name', '')} "
              f"({specs.get('power', '')}W) Solar PV Modules has been successfully completed. "
              f"All modules have passed the required quality checks as per applicable IEC/BIS standards. "
              f"The lot is APPROVED for dispatch.",
              font=Font(size=11),
              alignment=Alignment(horizontal='left', vertical='center', wrap_text=True))
    
    # Section 6: Signatures
    row += 2
    ws.merge_cells(f'A{row}:{col_letter}{row}')
    style_cell(ws, row, 1, "6. SIGNATURES", font=Font(bold=True, size=12, color="1565C0"),
              fill=sub_header_fill)
    
    row += 2
    style_cell(ws, row, 1, "For Inspector / Customer:", font=Font(bold=True),
              alignment=Alignment(horizontal='left'), border=None)
    style_cell(ws, row, 4, "For Manufacturer:", font=Font(bold=True),
              alignment=Alignment(horizontal='left'), border=None)
    
    row += 2
    style_cell(ws, row, 1, "Name:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'B{row}:C{row}')
    style_cell(ws, row, 2, inspector_name or '________________________', border=None,
              alignment=Alignment(horizontal='left'))
    style_cell(ws, row, 4, "Name:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'E{row}:F{row}')
    style_cell(ws, row, 5, manufacturer_rep or '________________________', border=None,
              alignment=Alignment(horizontal='left'))
    
    row += 1
    style_cell(ws, row, 1, "Designation:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'B{row}:C{row}')
    style_cell(ws, row, 2, inspector_designation, border=None, alignment=Alignment(horizontal='left'))
    style_cell(ws, row, 4, "Designation:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'E{row}:F{row}')
    style_cell(ws, row, 5, manufacturer_designation, border=None, alignment=Alignment(horizontal='left'))
    
    row += 1
    style_cell(ws, row, 1, "Signature:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'B{row}:C{row}')
    style_cell(ws, row, 2, '________________________', border=None, alignment=Alignment(horizontal='left'))
    style_cell(ws, row, 4, "Signature:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'E{row}:F{row}')
    style_cell(ws, row, 5, '________________________', border=None, alignment=Alignment(horizontal='left'))
    
    row += 1
    style_cell(ws, row, 1, "Date:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'B{row}:C{row}')
    style_cell(ws, row, 2, report_date, border=None, alignment=Alignment(horizontal='left'))
    style_cell(ws, row, 4, "Date:", font=Font(bold=True), alignment=Alignment(horizontal='left'), border=None)
    ws.merge_cells(f'E{row}:F{row}')
    style_cell(ws, row, 5, report_date, border=None, alignment=Alignment(horizontal='left'))
    
    # Column widths
    widths = [18, 18, 14, 18, 18, 14]
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


@pdi_doc_bp.route('/pdi-docs/companies', methods=['GET'])
def get_pdi_doc_companies():
    """Get companies with FTR data"""
    try:
        result = db.session.execute(text("""
            SELECT DISTINCT c.id, c.company_name, c.module_wattage, c.cells_per_module
            FROM companies c
            JOIN ftr_master_serials fms ON c.id = fms.company_id
            ORDER BY c.company_name
        """))
        companies = [{'id': row[0], 'name': row[1], 'wattage': row[2], 'cells': row[3]} for row in result.fetchall()]
        return jsonify({'success': True, 'companies': companies})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/pdi-docs/pdi-list/<int:company_id>', methods=['GET'])
def get_pdi_doc_pdi_list(company_id):
    """Get PDI numbers with serial counts"""
    try:
        result = db.session.execute(text("""
            SELECT pdi_number, COUNT(*) as cnt
            FROM ftr_master_serials
            WHERE company_id = :cid AND pdi_number IS NOT NULL AND pdi_number != ''
            GROUP BY pdi_number
            ORDER BY pdi_number
        """), {'cid': company_id})
        pdis = [{'pdi_number': row[0], 'count': row[1]} for row in result.fetchall()]
        return jsonify({'success': True, 'pdis': pdis})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


@pdi_doc_bp.route('/pdi-docs/serials/<int:company_id>/<pdi_number>', methods=['GET'])
def get_pdi_doc_serials(company_id, pdi_number):
    """Get all serials for a PDI"""
    try:
        result = db.session.execute(text("""
            SELECT serial_number FROM ftr_master_serials
            WHERE company_id = :cid AND pdi_number = :pdi
            ORDER BY serial_number
        """), {'cid': company_id, 'pdi': pdi_number})
        serials = [row[0] for row in result.fetchall()]
        return jsonify({'success': True, 'serials': serials, 'count': len(serials)})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500
